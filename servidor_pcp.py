from __future__ import annotations

from datetime import datetime, date
from pathlib import Path
import json
import os
import re
import sqlite3
import tempfile
import threading
import traceback
import uuid
import shutil
import hashlib
import hmac
import unicodedata

from flask import Flask, Response, has_request_context, jsonify, request, send_file

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, DoughnutChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None

import db_manutencao


def _env_path(var: str, default: Path) -> Path:
    """Permite sobrepor caminhos de arquivo via variável de ambiente.

    Usado pelos testes automatizados para apontar para uma cópia isolada do
    banco/Excel/log/backups sem jamais tocar nos arquivos reais. Sem a
    variável de ambiente, o comportamento é idêntico ao de antes.
    """
    value = os.environ.get(var)
    return Path(value) if value else default


ROOT = Path(__file__).resolve().parent
HTML_FILE = ROOT / "pcp_prototype.html"
DB_FILE = _env_path("PCP_DB_FILE", ROOT / "pcp.sqlite3")
EXCEL_FILE = _env_path("PCP_EXCEL_FILE", ROOT / "dados_pcp.xlsx")
EXCEL_PENDING_FILE = _env_path("PCP_EXCEL_PENDING_FILE", ROOT / "dados_pcp_pendente.xlsx")
BASE_CPD_FILE = ROOT / "base_cpds.json"
LOG_FILE = _env_path("PCP_LOG_FILE", ROOT / "servidor_pcp.log")
BACKUP_DIR = _env_path("PCP_BACKUP_DIR", ROOT / "backups")
PORT = int(os.environ.get("PCP_PORT", "8080"))

app = Flask(__name__)
lock = threading.RLock()

AUTH_SECRET = os.environ.get("PCP_AUTH_SECRET", "mtr-topura-pcp-interno")
AUTH_USERS = {
    "VENDAS": {"role": "view"},
    "EXPEDICAO": {"role": "view"},
    "PCP": {"role": "edit"},
}

AUTH_ALIASES = {
    "VENDAS": "VENDAS",
    "VENDA": "VENDAS",
    "COMERCIAL": "VENDAS",
    "EXPEDICAO": "EXPEDICAO",
    "EXP": "EXPEDICAO",
    "PCP": "PCP",
}

def normalize_login(value: str) -> str:
    text = str(value or "").strip().upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace(" ", "").replace("-", "").replace("_", "")
    return AUTH_ALIASES.get(text, text)

def make_token(login: str, role: str) -> str:
    raw = f"{normalize_login(login)}|{role}".encode("utf-8")
    return hmac.new(AUTH_SECRET.encode("utf-8"), raw, hashlib.sha256).hexdigest()

def get_auth_info() -> dict:
    # Login temporariamente SEM SENHA e SEM TOKEN obrigatório.
    # A tela envia apenas X-PCP-Login; PCP edita, VENDAS/EXPEDICAO visualizam.
    login = normalize_login(request.headers.get("X-PCP-Login", "") or "")
    user = AUTH_USERS.get(login)
    if not user:
        return {"login": "", "role": "", "ok": False}
    return {"login": login, "role": user["role"], "ok": True}

def require_editor():
    info = get_auth_info()
    if not info.get("ok") or info.get("role") != "edit":
        return jsonify({"error": "Acesso somente leitura. Entre como PCP para editar."}), 403
    return None



def log(msg: str) -> None:
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        with LOG_FILE.open("a", encoding="utf-8") as f:
            f.write(f"[{stamp}] {msg}\n")
    except Exception:
        pass


def agora() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def hoje_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def current_actor(default: str = "Não informado") -> str:
    """Nome da pessoa usando a tela. Vem do campo 'Usuário' no navegador.

    Fora de uma requisição HTTP (ex.: rotinas de inicialização do servidor)
    não existe `request` para consultar, então retorna `default` direto sem
    tocar em `request`/`get_auth_info()`.
    """
    if not has_request_context():
        return default
    actor = ""
    try:
        actor = (request.headers.get("X-PCP-User", "") or "").strip()
        if not actor and request.is_json:
            data = request.get_json(silent=True) or {}
            actor = str(data.get("_actor") or data.get("usuario") or "").strip()
    except Exception:
        actor = ""
    return (actor or get_auth_info().get("login") or default)[:80]


def uid(prefix: str = "") -> str:
    return prefix + uuid.uuid4().hex[:14]


def json_dumps(obj) -> str:
    return json.dumps(obj, ensure_ascii=False, separators=(",", ":"))


def json_loads(text: str, fallback):
    try:
        return json.loads(text) if text else fallback
    except Exception:
        return fallback


def to_int(value, fallback=None):
    """Converte valores vindos do Excel/tela em inteiro positivo."""
    if value is None or value == "":
        return fallback
    try:
        if isinstance(value, float):
            value = int(value)
        text = str(value).strip().replace(".", "").replace(",", "")
        number = int(text)
        return number if number > 0 else fallback
    except Exception:
        return fallback


def load_base_cpds() -> list[dict]:
    """Lê a base inicial de CPDs embutida no pacote.

    Ela serve só para popular o banco pela primeira vez; depois a aba
    'Base CPDs' do Excel e a tela passam a manter a base central.
    """
    try:
        if not BASE_CPD_FILE.exists():
            return []
        data = json.loads(BASE_CPD_FILE.read_text(encoding="utf-8"))
        if isinstance(data, list):
            return [x for x in data if isinstance(x, dict) and x.get("codigo") and x.get("descricao")]
    except Exception:
        log("Não consegui ler base_cpds.json:\n" + traceback.format_exc())
    return []


def seed_base_cpds(c: sqlite3.Connection) -> int:
    """Popula CPDs padrão sem sobrescrever descrições já alteradas pela equipe."""
    count = 0
    for item in load_base_cpds():
        codigo = str(item.get("codigo") or "").strip()
        descricao = str(item.get("descricao") or "").strip()
        if not codigo or not descricao:
            continue
        cur = c.execute(
            "INSERT OR IGNORE INTO manual_cpd(codigo,descricao,updated_at) VALUES(?,?,?)",
            (codigo, descricao, agora()),
        )
        count += cur.rowcount or 0
    return count


def get_next_order_number(c: sqlite3.Connection) -> int:
    c.execute("INSERT OR IGNORE INTO meta(key,value) VALUES('next_order_number','1')")
    row = c.execute("SELECT value FROM meta WHERE key='next_order_number'").fetchone()
    return max(1, to_int(row["value"] if row else None, 1) or 1)


def set_next_order_number(c: sqlite3.Connection, value: int) -> None:
    c.execute("INSERT OR REPLACE INTO meta(key,value) VALUES('next_order_number',?)", (str(max(1, int(value))),))


def order_id_by_number(c: sqlite3.Connection, number: int | None) -> str | None:
    if not number:
        return None
    for row in c.execute("SELECT id,payload FROM orders").fetchall():
        payload = json_loads(row["payload"], {})
        if to_int(payload.get("numero")) == int(number):
            return str(row["id"])
    return None


def assign_order_number(c: sqlite3.Connection, payload: dict) -> dict:
    """Garante número sequencial visível sem mudar o ID técnico interno."""
    payload = dict(payload or {})
    current = to_int(payload.get("numero"))
    if current is None:
        current = get_next_order_number(c)
        set_next_order_number(c, current + 1)
    else:
        next_number = get_next_order_number(c)
        if current >= next_number:
            set_next_order_number(c, current + 1)
    payload["numero"] = current
    return payload


def ensure_existing_order_numbers() -> None:
    """Numera pedidos antigos que ainda não tinham campo 'numero'."""
    with lock, conn() as c:
        changed = False
        max_seen = 0
        rows = c.execute("SELECT id,payload,updated_at FROM orders ORDER BY updated_at, id").fetchall()
        for row in rows:
            payload = normalize_order(json_loads(row["payload"], {}))
            number = to_int(payload.get("numero"))
            if number is None:
                payload = assign_order_number(c, payload)
                c.execute("UPDATE orders SET payload=?, updated_at=? WHERE id=?", (json_dumps(payload), row["updated_at"], row["id"]))
                changed = True
                number = payload.get("numero")
            if number:
                max_seen = max(max_seen, int(number))
        next_number = get_next_order_number(c)
        if max_seen + 1 > next_number:
            set_next_order_number(c, max_seen + 1)
        if changed:
            bump_version(c)
        c.commit()


def conn() -> sqlite3.Connection:
    c = sqlite3.connect(DB_FILE)
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA foreign_keys=ON")
    c.execute("PRAGMA journal_mode=WAL")
    c.execute("PRAGMA synchronous=NORMAL")
    return c


def ensure_audit_actor_column(c: sqlite3.Connection) -> None:
    """Garante compatibilidade com bancos criados por versões anteriores."""
    cols = {row["name"] for row in c.execute("PRAGMA table_info(audit_log)").fetchall()}
    if "actor" not in cols:
        c.execute("ALTER TABLE audit_log ADD COLUMN actor TEXT")


def init_db() -> None:
    with lock, conn() as c:
        c.executescript(
            """
            CREATE TABLE IF NOT EXISTS meta (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS orders (
                id TEXT PRIMARY KEY,
                payload TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS sections (
                id TEXT PRIMARY KEY,
                payload TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS manual_cpd (
                codigo TEXT PRIMARY KEY,
                descricao TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                action TEXT NOT NULL,
                entity TEXT NOT NULL,
                entity_id TEXT,
                payload TEXT,
                ip TEXT,
                user_agent TEXT,
                created_at TEXT NOT NULL
            );
            """
        )
        ensure_audit_actor_column(c)
        c.execute("INSERT OR IGNORE INTO meta(key,value) VALUES('version','0')")
        c.execute("INSERT OR IGNORE INTO meta(key,value) VALUES('updated_at','')")
        c.execute("INSERT OR IGNORE INTO meta(key,value) VALUES('next_order_number','1')")
        seed_base_cpds(c)
        # Cria a chave schema_version (começando em 1) se ainda não existir.
        # Nesta etapa não há migrações registradas: só prepara a estrutura.
        db_manutencao.get_schema_version(c)
        c.commit()


def run_pending_migrations() -> dict:
    """Aplica migrações de schema pendentes, sempre com backup validado antes.

    Chamar esta função só é seguro quando alguém decidiu deliberadamente
    aplicar a(s) migração(ões) pendente(s); ver deve_aplicar_migracoes_automaticamente().
    """
    with lock, conn() as c:
        resultado = db_manutencao.run_migrations(c, DB_FILE, BACKUP_DIR, log=log)
        if resultado.get("migracoes_aplicadas"):
            log(f"[migracao] Versão final do schema: {resultado['versao_final']}")
        return resultado


def deve_aplicar_migracoes_automaticamente() -> bool:
    """Só roda migrações de schema sozinho se PCP_APLICAR_MIGRACOES=1.

    Por padrão (variável ausente ou com qualquer outro valor) iniciar o
    servidor NÃO aplica migrações pendentes: alguém precisa decidir e setar
    essa variável explicitamente antes de reiniciar o servidor. Isso evita
    que registrar uma nova migração em db_manutencao.MIGRATIONS altere o
    schema do banco de produção na próxima vez que o servidor for iniciado,
    sem revisão humana explícita naquele momento.
    """
    return os.environ.get("PCP_APLICAR_MIGRACOES") == "1"


def get_meta(c: sqlite3.Connection) -> dict:
    rows = c.execute("SELECT key,value FROM meta").fetchall()
    d = {r["key"]: r["value"] for r in rows}
    return {"version": int(d.get("version") or 0), "updated_at": d.get("updated_at") or ""}


def bump_version(c: sqlite3.Connection) -> dict:
    m = get_meta(c)
    version = m["version"] + 1
    updated_at = agora()
    c.execute("INSERT OR REPLACE INTO meta(key,value) VALUES('version',?)", (str(version),))
    c.execute("INSERT OR REPLACE INTO meta(key,value) VALUES('updated_at',?)", (updated_at,))
    return {"version": version, "updated_at": updated_at}


def audit(c: sqlite3.Connection, action: str, entity: str, entity_id: str | None, payload=None, actor: str | None = None) -> None:
    try:
        ip = request.remote_addr
        user_agent = request.headers.get("User-Agent", "")
    except Exception:
        ip = None
        user_agent = ""
    c.execute(
        """
        INSERT INTO audit_log(action, entity, entity_id, payload, ip, user_agent, created_at, actor)
        VALUES(?,?,?,?,?,?,?,?)
        """,
        (
            action,
            entity,
            entity_id,
            json_dumps(payload) if payload is not None else None,
            ip,
            user_agent,
            agora(),
            actor or current_actor(),
        ),
    )


def normalize_order(data: dict) -> dict:
    if not isinstance(data, dict):
        data = {}
    order_id = str(data.get("id") or uid("ped"))
    out = dict(data)
    out["id"] = order_id
    out.setdefault("cpd", "")
    out.setdefault("descricao", "")
    out.setdefault("setor", "Prensa")
    out.setdefault("setorOutro", "")
    # Versão atual: os fluxos como V/ TÉRMICO, V/ SUPERFICIAL, Aguardando retorno,
    # Aguardando escolha, Concluído e Cancelado ficam no próprio campo Setor.
    legacy_status = normalize_status(out.get("status"), out.get("setor"))
    setor_norm = normalize_status(out.get("setor"), None)
    if legacy_status != "Pendente":
        out["setor"] = legacy_status
    elif setor_norm in STATUS_OPTIONS and setor_norm != "Pendente":
        out["setor"] = setor_norm
    out["status"] = ""
    out.setdefault("qtd", "")
    out.setdefault("dataEntregaOriginal", "")
    out.setdefault("dataEntregaAtual", out.get("dataEntregaOriginal", ""))
    out.setdefault("historico", [])
    if not isinstance(out.get("historico"), list):
        out["historico"] = []
    out.setdefault("previsao", "")
    out.setdefault("obs", "")
    out.setdefault("cliente", "")
    out.setdefault("secaoId", "")
    out.setdefault("dataConclusao", hoje_iso() if normalize_status(out.get("setor"), None) == "Concluído" else None)
    out.setdefault("criadoPor", "")
    out.setdefault("criadoEm", "")
    out.setdefault("alteradoPor", "")
    out.setdefault("alteradoEm", "")
    numero = to_int(out.get("numero"))
    if numero is not None:
        out["numero"] = numero
    else:
        out.pop("numero", None)
    out["deleted"] = bool(out.get("deleted", False))
    out.setdefault("deletedAt", "")
    out.setdefault("deletedBy", "")
    out.setdefault("deleteReason", "")
    return out


def normalize_section(data: dict) -> dict:
    if not isinstance(data, dict):
        data = {}
    out = dict(data)
    out["id"] = str(out.get("id") or uid("sec"))
    out.setdefault("name", "Data")
    out.setdefault("defaultDate", hoje_iso())
    out["active"] = bool(out.get("active", False))
    out["deleted"] = bool(out.get("deleted", False))
    out.setdefault("deletedAt", "")
    out.setdefault("deletedBy", "")
    out.setdefault("deleteReason", "")
    return out


def load_state_from_conn(c: sqlite3.Connection) -> dict:
    meta = get_meta(c)
    all_orders = [normalize_order(json_loads(r["payload"], {})) for r in c.execute("SELECT payload FROM orders ORDER BY updated_at DESC, id DESC")]
    all_sections = [normalize_section(json_loads(r["payload"], {})) for r in c.execute("SELECT payload FROM sections ORDER BY updated_at, id")]
    orders = [o for o in all_orders if not o.get("deleted")]
    sections = [s for s in all_sections if not s.get("deleted")]
    sections.sort(key=lambda s: (s.get("defaultDate", ""), s.get("id", "")))
    cpd_rows = c.execute("SELECT codigo, descricao FROM manual_cpd ORDER BY codigo").fetchall()
    manual_cpd = {r["codigo"]: r["descricao"] for r in cpd_rows}
    return {"version": meta["version"], "updated_at": meta["updated_at"], "orders": orders, "sections": sections, "manualCPD": manual_cpd}


def load_state() -> dict:
    with lock, conn() as c:
        return load_state_from_conn(c)


def ensure_first_section(c: sqlite3.Connection) -> str:
    for row in c.execute("SELECT payload FROM sections ORDER BY updated_at DESC").fetchall():
        sec = normalize_section(json_loads(row["payload"], {}))
        if not sec.get("deleted"):
            return sec.get("id") or ""
    sec = {"id": uid("sec"), "name": "Importação Excel", "defaultDate": hoje_iso(), "active": True}
    c.execute("INSERT OR REPLACE INTO sections(id,payload,updated_at) VALUES(?,?,?)", (sec["id"], json_dumps(sec), agora()))
    return sec["id"]



def run_daily_backup(force: bool = False) -> None:
    """Cria backup diário do banco (via API oficial do SQLite) e do Excel.

    O banco usa sqlite3.Connection.backup() através de db_manutencao, que
    funciona corretamente com o banco em modo WAL e só considera o backup
    válido se ele abrir e o PRAGMA integrity_check retornar "ok". O Excel
    continua copiado com shutil.copy2 (não é um banco SQLite, não usa WAL).
    Nenhum backup existente é apagado.
    """
    try:
        BACKUP_DIR.mkdir(exist_ok=True)
        today = datetime.now().strftime("%Y-%m-%d")
        stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        marker_key = "last_backup_date"
        with conn() as c:
            c.execute("INSERT OR IGNORE INTO meta(key,value) VALUES(?,?)", (marker_key, ""))
            last = c.execute("SELECT value FROM meta WHERE key=?", (marker_key,)).fetchone()
            if (not force) and last and last["value"] == today:
                return
            c.execute("INSERT OR REPLACE INTO meta(key,value) VALUES(?,?)", (marker_key, today))
            c.commit()

        backup_path = db_manutencao.criar_backup_sqlite(DB_FILE, BACKUP_DIR, prefixo="pcp_backup", log=log)
        if backup_path is None:
            log("Backup automático do SQLite FALHOU (integrity_check inválido ou erro). Veja o log acima.")
        if EXCEL_FILE.exists():
            shutil.copy2(EXCEL_FILE, BACKUP_DIR / f"dados_pcp_backup_{stamp}.xlsx")
        log(f"Backup automático gerado em {BACKUP_DIR}")
    except Exception:
        log("Erro ao gerar backup automático:\n" + traceback.format_exc())


def save_excel_snapshot_safely() -> None:
    if Workbook is None:
        log("openpyxl não está instalado; Excel não foi atualizado.")
        return
    try:
        state = load_state()
        tmp = EXCEL_FILE.with_suffix(".tmp.xlsx")
        export_state_to_xlsx(state, tmp)
        try:
            os.replace(tmp, EXCEL_FILE)
        except PermissionError:
            # Se o Excel estiver aberto, o Windows pode bloquear a substituição.
            os.replace(tmp, EXCEL_PENDING_FILE)
            log(f"Excel principal estava aberto. Snapshot salvo como {EXCEL_PENDING_FILE.name}")
    except Exception:
        log("Erro ao atualizar Excel:\n" + traceback.format_exc())


def after_write_state(c: sqlite3.Connection, update_excel: bool = True) -> dict:
    bump_version(c)
    state = load_state_from_conn(c)
    c.commit()
    if update_excel:
        # Atualiza o arquivo de Excel que funciona como relatório/base exportada.
        save_excel_snapshot_safely()
        run_daily_backup(False)
    return state



@app.post("/api/login")
def api_login():
    data = request.get_json(silent=True) or {}
    login = normalize_login(data.get("usuario") or data.get("login") or "")
    user = AUTH_USERS.get(login)
    if not user:
        return jsonify({"error": "Usuário inválido"}), 401
    role = user["role"]
    return jsonify({"login": login, "role": role, "token": "sem-senha"})

@app.get("/api/meta")
def api_meta():
    with lock, conn() as c:
        return jsonify(get_meta(c))


@app.get("/api/state")
def api_state():
    return jsonify(load_state())


@app.route("/api/orders/<order_id>", methods=["PUT", "PATCH", "DELETE"])
def api_order(order_id: str):
    denied = require_editor()
    if denied:
        return denied
    actor = current_actor()
    with lock, conn() as c:
        if request.method == "DELETE":
            old = c.execute("SELECT payload FROM orders WHERE id=?", (order_id,)).fetchone()
            if old:
                payload = normalize_order(json_loads(old["payload"], {}))
                data = request.get_json(silent=True) or {}
                payload["deleted"] = True
                payload["deletedAt"] = agora()
                payload["deletedBy"] = actor
                payload["deleteReason"] = str(data.get("motivo") or data.get("reason") or "Removido pela tela").strip()
                c.execute("UPDATE orders SET payload=?, updated_at=? WHERE id=?", (json_dumps(payload), agora(), order_id))
                audit(c, "trash", "order", order_id, payload, actor)
            return jsonify(after_write_state(c))

        incoming = request.get_json(silent=True) or {}
        incoming.pop("_actor", None)
        timestamp = agora()

        if request.method == "PATCH":
            old = c.execute("SELECT payload FROM orders WHERE id=?", (order_id,)).fetchone()
            payload = normalize_order(json_loads(old["payload"], {}) if old else {"id": order_id})
            if not payload.get("criadoPor"):
                payload["criadoPor"] = "Sistema/legado"
            if not payload.get("criadoEm"):
                payload["criadoEm"] = payload.get("alteradoEm") or timestamp
            payload.update(incoming)
            payload["id"] = order_id
            payload["alteradoPor"] = actor
            payload["alteradoEm"] = timestamp
        else:
            payload = normalize_order(incoming)
            payload["id"] = order_id
            payload["criadoPor"] = payload.get("criadoPor") or actor
            payload["criadoEm"] = payload.get("criadoEm") or timestamp
            payload["alteradoPor"] = actor
            payload["alteradoEm"] = timestamp

        payload = assign_order_number(c, payload)
        c.execute("INSERT OR REPLACE INTO orders(id,payload,updated_at) VALUES(?,?,?)", (order_id, json_dumps(payload), timestamp))
        audit(c, "upsert" if request.method == "PUT" else "patch", "order", order_id, incoming, actor)
        return jsonify(after_write_state(c))


@app.route("/api/sections/<section_id>", methods=["PUT", "PATCH", "DELETE"])

def api_section(section_id: str):
    denied = require_editor()
    if denied:
        return denied
    actor = current_actor()
    with lock, conn() as c:
        if request.method == "DELETE":
            data = request.get_json(silent=True) or {}
            mode = str(data.get("mode") or "block")
            old_row = c.execute("SELECT payload FROM sections WHERE id=?", (section_id,)).fetchone()
            if not old_row:
                return jsonify({"error": "Data não encontrada"}), 404

            old_section = normalize_section(json_loads(old_row["payload"], {}))

            linked_orders = []
            for row in c.execute("SELECT id,payload FROM orders").fetchall():
                payload = normalize_order(json_loads(row["payload"], {}))
                if payload.get("secaoId") == section_id:
                    linked_orders.append((row["id"], payload))

            if linked_orders and mode != "delete_orders":
                return jsonify({
                    "error": "Esta data possui pedidos vinculados. Confirme para apagar os pedidos junto ou mova os pedidos antes.",
                    "orders_count": len(linked_orders),
                }), 409

            if mode == "delete_orders":
                for order_id, order_payload in linked_orders:
                    order_payload["deleted"] = True
                    order_payload["deletedAt"] = agora()
                    order_payload["deletedBy"] = actor
                    order_payload["deleteReason"] = "Data apagada"
                    c.execute("UPDATE orders SET payload=?, updated_at=? WHERE id=?", (json_dumps(order_payload), agora(), order_id))
                    audit(c, "trash", "order", order_id, {"motivo": "Data apagada", "section_id": section_id, "order": order_payload}, actor)

            old_section["deleted"] = True
            old_section["deletedAt"] = agora()
            old_section["deletedBy"] = actor
            old_section["deleteReason"] = "Data apagada"
            old_section["active"] = False
            c.execute("UPDATE sections SET payload=?, updated_at=? WHERE id=?", (json_dumps(old_section), agora(), section_id))
            audit(c, "trash", "section", section_id, {"section": old_section, "orders_deleted": len(linked_orders), "mode": mode}, actor)

            remaining = []
            has_active = False
            for row in c.execute("SELECT id,payload FROM sections").fetchall():
                sec = normalize_section(json_loads(row["payload"], {}))
                if sec.get("deleted"):
                    continue
                remaining.append((row["id"], sec))
                if sec.get("active"):
                    has_active = True

            if remaining and not has_active:
                chosen_id, chosen = sorted(remaining, key=lambda item: (item[1].get("defaultDate", ""), item[0]))[-1]
                chosen["active"] = True
                chosen["alteradoPor"] = actor
                chosen["alteradoEm"] = agora()
                c.execute("UPDATE sections SET payload=?, updated_at=? WHERE id=?", (json_dumps(chosen), agora(), chosen_id))
                audit(c, "patch", "section", chosen_id, {"active": True, "motivo": "Ativada automaticamente após exclusão de data"}, actor)

            return jsonify(after_write_state(c))

        incoming = request.get_json(silent=True) or {}
        incoming.pop("_actor", None)
        if request.method == "PATCH":
            old = c.execute("SELECT payload FROM sections WHERE id=?", (section_id,)).fetchone()
            payload = normalize_section(json_loads(old["payload"], {}) if old else {"id": section_id})
            payload.update(incoming)
            payload["id"] = section_id
        else:
            payload = normalize_section(incoming)
            payload["id"] = section_id

        timestamp = agora()
        if not payload.get("criadoPor"):
            payload["criadoPor"] = actor
        if not payload.get("criadoEm"):
            payload["criadoEm"] = timestamp
        payload["alteradoPor"] = actor
        payload["alteradoEm"] = timestamp

        if payload.get("active"):
            for row in c.execute("SELECT id,payload FROM sections WHERE id<>?", (section_id,)).fetchall():
                sec = normalize_section(json_loads(row["payload"], {}))
                sec["active"] = False
                c.execute("UPDATE sections SET payload=?, updated_at=? WHERE id=?", (json_dumps(sec), agora(), row["id"]))

        c.execute("INSERT OR REPLACE INTO sections(id,payload,updated_at) VALUES(?,?,?)", (section_id, json_dumps(payload), timestamp))
        audit(c, "upsert" if request.method == "PUT" else "patch", "section", section_id, incoming, actor)
        return jsonify(after_write_state(c))



@app.get("/api/trash")
def api_trash():
    with lock, conn() as c:
        orders = []
        sections = []
        for row in c.execute("SELECT id,payload,updated_at FROM orders ORDER BY updated_at DESC").fetchall():
            payload = normalize_order(json_loads(row["payload"], {}))
            if payload.get("deleted"):
                orders.append(payload)
        for row in c.execute("SELECT id,payload,updated_at FROM sections ORDER BY updated_at DESC").fetchall():
            payload = normalize_section(json_loads(row["payload"], {}))
            if payload.get("deleted"):
                sections.append(payload)
        return jsonify({"orders": orders, "sections": sections})


@app.post("/api/trash/restore/<entity>/<item_id>")
def api_trash_restore(entity: str, item_id: str):
    denied = require_editor()
    if denied:
        return denied
    actor = current_actor()
    include_orders = bool((request.get_json(silent=True) or {}).get("include_orders", False))
    with lock, conn() as c:
        if entity == "order":
            row = c.execute("SELECT payload FROM orders WHERE id=?", (item_id,)).fetchone()
            if not row:
                return jsonify({"error": "Pedido não encontrado na lixeira"}), 404
            payload = normalize_order(json_loads(row["payload"], {}))
            payload["deleted"] = False
            payload["deletedAt"] = ""
            payload["deletedBy"] = ""
            payload["deleteReason"] = ""
            # Se a data do pedido também estiver excluída, restaura a data junto.
            sec_id = payload.get("secaoId")
            if sec_id:
                sec_row = c.execute("SELECT payload FROM sections WHERE id=?", (sec_id,)).fetchone()
                if sec_row:
                    sec = normalize_section(json_loads(sec_row["payload"], {}))
                    if sec.get("deleted"):
                        sec["deleted"] = False
                        sec["deletedAt"] = ""
                        sec["deletedBy"] = ""
                        sec["deleteReason"] = ""
                        c.execute("UPDATE sections SET payload=?, updated_at=? WHERE id=?", (json_dumps(sec), agora(), sec_id))
            c.execute("UPDATE orders SET payload=?, updated_at=? WHERE id=?", (json_dumps(payload), agora(), item_id))
            audit(c, "restore", "order", item_id, payload, actor)
            return jsonify(after_write_state(c))
        if entity == "section":
            row = c.execute("SELECT payload FROM sections WHERE id=?", (item_id,)).fetchone()
            if not row:
                return jsonify({"error": "Data não encontrada na lixeira"}), 404
            payload = normalize_section(json_loads(row["payload"], {}))
            payload["deleted"] = False
            payload["deletedAt"] = ""
            payload["deletedBy"] = ""
            payload["deleteReason"] = ""
            c.execute("UPDATE sections SET payload=?, updated_at=? WHERE id=?", (json_dumps(payload), agora(), item_id))
            restored_orders = 0
            if include_orders:
                for orow in c.execute("SELECT id,payload FROM orders").fetchall():
                    op = normalize_order(json_loads(orow["payload"], {}))
                    if op.get("secaoId") == item_id and op.get("deleted"):
                        op["deleted"] = False
                        op["deletedAt"] = ""
                        op["deletedBy"] = ""
                        op["deleteReason"] = ""
                        c.execute("UPDATE orders SET payload=?, updated_at=? WHERE id=?", (json_dumps(op), agora(), orow["id"]))
                        restored_orders += 1
            audit(c, "restore", "section", item_id, {"section": payload, "orders_restored": restored_orders}, actor)
            return jsonify(after_write_state(c))
        return jsonify({"error": "Tipo inválido"}), 400


@app.post("/api/backup-now")
def api_backup_now():
    denied = require_editor()
    if denied:
        return denied
    run_daily_backup(True)
    return jsonify({"ok": True, "folder": str(BACKUP_DIR)})


@app.route("/api/cpd/<codigo>", methods=["PUT", "DELETE"])
def api_cpd(codigo: str):
    denied = require_editor()
    if denied:
        return denied
    actor = current_actor()
    with lock, conn() as c:
        if request.method == "DELETE":
            c.execute("DELETE FROM manual_cpd WHERE codigo=?", (codigo,))
            audit(c, "delete", "cpd", codigo, None, actor)
            return jsonify(after_write_state(c))
        data = request.get_json(silent=True) or {}
        data.pop("_actor", None)
        desc = str(data.get("descricao") or "").strip()
        if not desc:
            return jsonify({"error": "Descrição ausente"}), 400
        c.execute("INSERT OR REPLACE INTO manual_cpd(codigo,descricao,updated_at) VALUES(?,?,?)", (codigo, desc, agora()))
        audit(c, "upsert", "cpd", codigo, {"descricao": desc}, actor)
        return jsonify(after_write_state(c))


@app.post("/api/import-json")
def api_import_json():
    denied = require_editor()
    if denied:
        return denied
    actor = current_actor()
    data = request.get_json(silent=True) or {}
    data.pop("_actor", None)
    with lock, conn() as c:
        for sec in data.get("sections", []) if isinstance(data.get("sections"), list) else []:
            sec = normalize_section(sec)
            c.execute("INSERT OR REPLACE INTO sections(id,payload,updated_at) VALUES(?,?,?)", (sec["id"], json_dumps(sec), agora()))
        for order in data.get("orders", []) if isinstance(data.get("orders"), list) else []:
            order = normalize_order(order)
            if not order.get("secaoId"):
                order["secaoId"] = ensure_first_section(c)
            order.setdefault("criadoPor", actor)
            order.setdefault("criadoEm", agora())
            order["alteradoPor"] = actor
            order["alteradoEm"] = agora()
            order = assign_order_number(c, order)
            c.execute("INSERT OR REPLACE INTO orders(id,payload,updated_at) VALUES(?,?,?)", (order["id"], json_dumps(order), agora()))
        manual = data.get("manualCPD") if isinstance(data.get("manualCPD"), dict) else {}
        for codigo, descricao in manual.items():
            c.execute("INSERT OR REPLACE INTO manual_cpd(codigo,descricao,updated_at) VALUES(?,?,?)", (str(codigo), str(descricao), agora()))
        audit(c, "import", "json", None, {"orders": len(data.get("orders", []) or []), "sections": len(data.get("sections", []) or []), "cpds": len(manual)}, actor)
        return jsonify(after_write_state(c))


def fmt_date_cell(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    # yyyy-mm-dd já está OK.
    if len(text) >= 10 and text[4:5] == "-" and text[7:8] == "-":
        return text[:10]
    # dd/mm/yyyy, dd-mm-yyyy ou dd/mm/aa
    for sep in ("/", "-"):
        if sep in text:
            parts = text.split(sep)
            try:
                if len(parts) >= 3 and len(parts[0]) <= 2:
                    d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                    if y < 100:
                        y += 2000
                    return date(y, m, d).isoformat()
            except Exception:
                return text
    return text


def sheet_headers(ws) -> dict:
    headers = {}
    for idx, cell in enumerate(ws[1], 1):
        if cell.value is not None:
            headers[str(cell.value).strip().lower()] = idx
    return headers


def cell_by_header(row, headers: dict, *names):
    for name in names:
        col = headers.get(name.lower())
        if col:
            return row[col - 1].value
    return None


def import_cpd_sheets_into_conn(wb, c: sqlite3.Connection) -> int:
    count = 0
    for cpd_sheet_name in ("Base CPDs", "CPDs", "CPDs Manuais", "CPD Manual", "cpdManual"):
        if cpd_sheet_name in wb.sheetnames:
            ws = wb[cpd_sheet_name]
            headers = sheet_headers(ws)
            for row in ws.iter_rows(min_row=2):
                codigo = cell_by_header(row, headers, "código", "codigo", "cpd")
                desc = cell_by_header(row, headers, "descrição", "descricao", "descrição do item", "descricao do item")
                if codigo and desc:
                    c.execute(
                        "INSERT OR REPLACE INTO manual_cpd(codigo,descricao,updated_at) VALUES(?,?,?)",
                        (str(codigo).strip(), str(desc).strip(), agora()),
                    )
                    count += 1
            break
    return count


def import_cpds_from_excel_file(path: Path, actor: str | None = None) -> dict:
    """Recarrega a aba Base CPDs do Excel.

    `actor` deve ser informado explicitamente quando chamado fora de uma
    requisição HTTP (ex.: na inicialização do servidor, use
    actor="Sistema/Inicialização"), já que nesse caso não há `request` para
    `current_actor()` consultar. Quando chamado a partir de uma rota da
    tela, deixe `actor=None` para usar o usuário autenticado normalmente.
    """
    if load_workbook is None:
        raise RuntimeError("openpyxl não está instalado")
    if not path.exists():
        raise FileNotFoundError(str(path))
    wb = load_workbook(path, data_only=True)
    with lock, conn() as c:
        imported = import_cpd_sheets_into_conn(wb, c)
        audit(c, "import", "cpd_xlsx", None, {"cpds": imported, "arquivo": path.name}, actor or current_actor("Importação Excel"))
        state = after_write_state(c)
    return {**state, "imported": {"cpds": imported}}


def import_workbook(path: Path) -> dict:
    if load_workbook is None:
        raise RuntimeError("openpyxl não está instalado")
    wb = load_workbook(path, data_only=True)
    imported = {"orders": 0, "sections": 0, "cpds": 0}
    with lock, conn() as c:
        # Datas
        for sec_sheet_name in ("Datas", "Seções", "Secoes", "secoes", "sections"):
            if sec_sheet_name in wb.sheetnames:
                ws = wb[sec_sheet_name]
                headers = sheet_headers(ws)
                for row in ws.iter_rows(min_row=2):
                    sec_id = cell_by_header(row, headers, "id", "seção id", "secao id") or uid("sec")
                    name = cell_by_header(row, headers, "nome", "name", "seção", "secao") or "Data"
                    default_date = fmt_date_cell(cell_by_header(row, headers, "data padrão", "data padrao", "defaultDate", "data")) or hoje_iso()
                    active_raw = cell_by_header(row, headers, "ativa", "active")
                    active = str(active_raw).strip().lower() in ("sim", "true", "1", "x", "ativa")
                    sec = normalize_section({"id": str(sec_id), "name": str(name), "defaultDate": default_date, "active": active})
                    c.execute("INSERT OR REPLACE INTO sections(id,payload,updated_at) VALUES(?,?,?)", (sec["id"], json_dumps(sec), agora()))
                    imported["sections"] += 1
                break

        default_section_id = ensure_first_section(c)

        # Base de CPDs: a aba "Base CPDs" é a fonte central de consulta da tela.
        imported["cpds"] += import_cpd_sheets_into_conn(wb, c)

        # Histórico agrupado por pedido
        historicos = {}
        for hist_sheet_name in ("Histórico", "Historico", "historico"):
            if hist_sheet_name in wb.sheetnames:
                ws = wb[hist_sheet_name]
                headers = sheet_headers(ws)
                for row in ws.iter_rows(min_row=2):
                    pedido_id = cell_by_header(row, headers, "pedido id", "id pedido", "id")
                    if not pedido_id:
                        continue
                    historicos.setdefault(str(pedido_id), []).append({
                        "de": fmt_date_cell(cell_by_header(row, headers, "de")),
                        "para": fmt_date_cell(cell_by_header(row, headers, "para")),
                        "quando": fmt_date_cell(cell_by_header(row, headers, "quando", "data")),
                        "motivo": str(cell_by_header(row, headers, "motivo") or ""),
                    })
                break

        # Pedidos: usa a aba Pedidos se existir; caso contrário tenta a primeira aba.
        ws = wb["Pedidos"] if "Pedidos" in wb.sheetnames else wb[wb.sheetnames[0]]
        headers = sheet_headers(ws)
        if any(k in headers for k in ("cpd", "código", "codigo")):
            for row in ws.iter_rows(min_row=2):
                cpd = cell_by_header(row, headers, "cpd", "código", "codigo")
                if not cpd:
                    continue
                numero = to_int(cell_by_header(row, headers, "nº", "n°", "numero", "número", "n pedido", "número pedido", "numero pedido"))
                order_id = cell_by_header(row, headers, "id interno", "id", "pedido id")
                if not order_id and numero:
                    order_id = order_id_by_number(c, numero)
                order_id = str(order_id or uid("ped"))
                setor = cell_by_header(row, headers, "setor") or "Prensa"
                data_atual = fmt_date_cell(cell_by_header(row, headers, "data entrega atual", "entrega", "data de entrega"))
                data_original = fmt_date_cell(cell_by_header(row, headers, "data entrega original")) or data_atual
                order = normalize_order({
                    "id": order_id,
                    "numero": numero,
                    "cpd": str(cpd),
                    "descricao": str(cell_by_header(row, headers, "descrição", "descricao", "descrição do item", "descricao do item") or ""),
                    "setor": str(setor),
                    "setorOutro": str(cell_by_header(row, headers, "setor outro", "detalhe setor") or ""),
                    "status": str(cell_by_header(row, headers, "status", "situação", "situacao") or ""),
                    "qtd": str(cell_by_header(row, headers, "qtd faltante", "quantidade faltante", "qtd") or ""),
                    "dataEntregaOriginal": data_original,
                    "dataEntregaAtual": data_atual or data_original,
                    "previsao": fmt_date_cell(cell_by_header(row, headers, "previsão retorno", "previsao retorno", "previsão", "previsao")),
                    "obs": str(cell_by_header(row, headers, "observação", "observacao", "obs") or ""),
                    "cliente": str(cell_by_header(row, headers, "cliente") or ""),
                    "secaoId": str(cell_by_header(row, headers, "grupo id", "seção id", "secao id") or default_section_id),
                    "dataConclusao": fmt_date_cell(cell_by_header(row, headers, "data conclusão", "data conclusao", "concluído em", "concluido em")) or None,
                    "criadoPor": str(cell_by_header(row, headers, "criado por", "adicionado por") or current_actor("Importação Excel")),
                    "criadoEm": str(cell_by_header(row, headers, "criado em", "adicionado em") or agora()),
                    "alteradoPor": str(cell_by_header(row, headers, "alterado por", "atualizado por") or current_actor("Importação Excel")),
                    "alteradoEm": str(cell_by_header(row, headers, "alterado em", "atualizado em") or agora()),
                    "historico": historicos.get(order_id, []),
                })
                order = assign_order_number(c, order)
                c.execute("INSERT OR REPLACE INTO orders(id,payload,updated_at) VALUES(?,?,?)", (order["id"], json_dumps(order), agora()))
                imported["orders"] += 1

        audit(c, "import", "xlsx", None, imported, current_actor("Importação Excel"))
        state = after_write_state(c)
    return {**state, "imported": imported}


@app.post("/api/import.xlsx")
def api_import_xlsx():
    denied = require_editor()
    if denied:
        return denied
    if "file" not in request.files:
        return jsonify({"error": "Arquivo não enviado"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Envie um arquivo .xlsx"}), 400
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        f.save(tmp.name)
        tmp_path = Path(tmp.name)
    try:
        return jsonify(import_workbook(tmp_path))
    finally:
        try:
            tmp_path.unlink()
        except Exception:
            pass


@app.post("/api/cpds/reload-excel")
def api_reload_cpds_from_excel():
    denied = require_editor()
    if denied:
        return denied
    try:
        data = import_cpds_from_excel_file(EXCEL_FILE)
        return jsonify(data)
    except FileNotFoundError:
        return jsonify({"error": f"Arquivo {EXCEL_FILE.name} não encontrado"}), 404
    except Exception:
        log("Erro ao recarregar CPDs do Excel:\n" + traceback.format_exc())
        return jsonify({"error": "Não consegui recarregar a aba Base CPDs do Excel"}), 500


STATUS_OPTIONS = ["Pendente", "V/ TÉRMICO", "V/ SUPERFICIAL", "Aguardando retorno", "Aguardando escolha", "Concluído", "Cancelado"]


def normalize_status(value: str | None, setor: str | None = None) -> str:
    text = str(value or "").strip()
    if not text and setor == "Escolha":
        return "Concluído"
    if not text:
        return "Pendente"
    m = {
        "pendente": "Pendente",
        "v/ térmico": "V/ TÉRMICO",
        "v/ termico": "V/ TÉRMICO",
        "mtc - tt": "MTC - TT",
        "mtc-tt": "MTC - TT",
        "térmico": "V/ TÉRMICO",
        "termico": "V/ TÉRMICO",
        "v/ superficial": "V/ SUPERFICIAL",
        "mtc - ts": "MTC - TS",
        "mtc-ts": "MTC - TS",
        "superficial": "V/ SUPERFICIAL",
        "aguardando retorno": "Aguardando retorno",
        "aguardando escolha": "Aguardando escolha",
        "concluído": "Concluído",
        "concluido": "Concluído",
        "escolha": "Concluído",
        "cancelado": "Cancelado",
        "cancelada": "Cancelado",
    }
    return m.get(text.lower(), text)


def compute_status(order: dict) -> str:
    status = normalize_status(order.get("setor") or order.get("status"), None)
    if status == "Concluído":
        return "concluido"
    if status == "Cancelado":
        return "cancelado"
    if status == "V/ TÉRMICO":
        return "termico"
    if status == "V/ SUPERFICIAL":
        return "superficial"
    if status == "Aguardando retorno":
        return "retorno"
    if status == "Aguardando escolha":
        return "escolha"
    return "pendente"


def status_label(order: dict) -> str:
    return normalize_status(order.get("setor") or order.get("status"), None)


def is_finalizado(order: dict) -> bool:
    return compute_status(order) in {"concluido", "cancelado"}


def prazo_label(order: dict) -> str:
    data = parse_date_value(order.get("dataEntregaAtual"))
    if not isinstance(data, date):
        return "Sem data"
    diff = (data - date.today()).days
    if diff < 0:
        return "Vencido"
    if diff <= 7:
        return "Esta semana"
    if diff <= 14:
        return "Atenção"
    return "No prazo"


def atraso_dias(order: dict) -> int:
    data = parse_date_value(order.get("dataEntregaAtual"))
    if not isinstance(data, date):
        return 0
    return max(0, (date.today() - data).days)


def grupo_cliente(cliente: str) -> str:
    return "HONDA" if cliente in {"HONDA", "METALFINO", "YASUFUKU"} else "DIVERSOS"


def parse_date_value(value):
    """Retorna date para valores ISO/dd-mm-aaaa/dd/mm/aaaa; mantém texto se inválido."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    try:
        if len(text) >= 10 and text[4:5] == "-" and text[7:8] == "-":
            return datetime.strptime(text[:10], "%Y-%m-%d").date()
        for sep in ("-", "/"):
            if sep in text:
                parts = text.split(sep)
                if len(parts) >= 3 and len(parts[0]) <= 2:
                    d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                    if y < 100:
                        y += 2000
                    return date(y, m, d)
    except Exception:
        return text
    return text


def parse_datetime_value(value):
    """Retorna datetime sem timezone para Excel; mantém texto se inválido."""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    try:
        clean = text.replace("Z", "+00:00")
        dt = datetime.fromisoformat(clean)
        return dt.replace(tzinfo=None)
    except Exception:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%d-%m-%Y %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text[:len(datetime.now().strftime(fmt))] if fmt != "%Y-%m-%d" else text[:10], fmt)
        except Exception:
            continue
    return text


def format_date_br(value) -> str:
    d = parse_date_value(value)
    if isinstance(d, date):
        return d.strftime("%d-%m-%Y")
    return str(value or "")


def section_label(section: dict) -> str:
    if not section:
        return ""
    return format_date_br(section.get("defaultDate")) or str(section.get("name") or "")


def audit_explicacao(row: dict) -> str:
    """Transforma o registro técnico de auditoria em uma frase simples para a equipe."""
    action = str(row.get("action") or "").lower()
    entity = str(row.get("entity") or "").lower()
    payload = json_loads(row.get("payload") or "", {})

    if entity == "order":
        numero = payload.get("numero") or payload.get("nº") or payload.get("numero_pedido") or ""
        cpd = payload.get("cpd") or ""
        ident = []
        if numero:
            ident.append(f"Nº {numero}")
        if cpd:
            ident.append(f"CPD {cpd}")
        alvo = (" (" + " / ".join(ident) + ")") if ident else ""
        if action in {"delete", "trash"}:
            motivo = payload.get("deleteReason") or payload.get("motivo") or ""
            return f"Pedido{alvo} foi enviado para a lixeira" + (f". Motivo: {motivo}." if motivo else ".")
        if action == "restore":
            return f"Pedido{alvo} foi restaurado da lixeira."
        if action in {"upsert", "create"}:
            return f"Pedido{alvo} foi criado ou salvo no controle."
        if action == "patch":
            campos = [k for k in payload.keys() if not str(k).startswith("_")]
            nomes = {
                "setor": "setor/situação",
                "previsao": "previsão de retorno",
                "dataEntregaAtual": "data de entrega",
                "dataEntregaOriginal": "data de entrega original",
                "qtd": "quantidade faltante",
                "obs": "observação",
                "cliente": "cliente",
                "secaoId": "data/grupo",
                "historico": "histórico",
            }
            legiveis = [nomes.get(c, c) for c in campos if c not in {"id", "alteradoPor", "alteradoEm"}]
            if legiveis:
                return f"Pedido{alvo} foi alterado. Campos: " + ", ".join(legiveis[:6]) + ("..." if len(legiveis) > 6 else "") + "."
            return f"Pedido{alvo} foi alterado."
        return f"Movimentação realizada no pedido{alvo}."

    if entity == "section":
        sec = payload.get("section") if isinstance(payload, dict) else {}
        data_ref = format_date_br((sec or {}).get("defaultDate")) if isinstance(sec, dict) else ""
        if action in {"delete", "trash"}:
            qtd = payload.get("orders_deleted") if isinstance(payload, dict) else 0
            if qtd:
                return f"Data {data_ref or row.get('entity_id')} foi enviada para a lixeira e {qtd} pedido(s) vinculado(s) também foram enviados."
            return f"Data {data_ref or row.get('entity_id')} foi enviada para a lixeira."
        if action == "restore":
            qtd = payload.get("orders_restored") if isinstance(payload, dict) else 0
            return f"Data {data_ref or row.get('entity_id')} foi restaurada da lixeira" + (f" com {qtd} pedido(s)." if qtd else ".")
        if action in {"upsert", "create"}:
            return "Nova data/grupo foi criada ou salva."
        return "Data/grupo foi alterada."

    if entity == "cpd":
        if action == "delete":
            return f"CPD {row.get('entity_id') or ''} foi removido da base de CPDs."
        return f"CPD {row.get('entity_id') or ''} foi cadastrado ou atualizado na base de CPDs."

    if entity in {"excel", "cpd_xlsx"}:
        return "Informações foram importadas a partir da planilha Excel."
    if entity == "json":
        return "Dados antigos/locais foram importados para a base central."

    return "Registro técnico de movimentação do sistema."


def resumo_auditoria_linha(row: dict) -> list:
    return [
        row.get("id", ""),
        parse_datetime_value(row.get("created_at", "")),
        row.get("actor", ""),
        audit_explicacao(row),
        row.get("action", ""),
        row.get("entity", ""),
        row.get("entity_id", ""),
        row.get("ip", ""),
        row.get("payload", ""),
    ]


def load_audit_rows(limit: int = 5000) -> list[dict]:
    try:
        with conn() as c:
            rows = c.execute(
                """
                SELECT id, created_at, actor, action, entity, entity_id, payload, ip
                FROM audit_log
                ORDER BY id DESC
                LIMIT ?
                """,
                (limit,),
            ).fetchall()
            return [dict(r) for r in reversed(rows)]
    except Exception:
        return []



def normalize_text(v: str) -> str:
    import unicodedata
    return "".join(ch for ch in unicodedata.normalize("NFD", str(v or "")) if unicodedata.category(ch) != "Mn").lower().strip()

def setor_rank(order: dict) -> int:
    s = normalize_text(order.get("setor", ""))
    if s == "prensa":
        return 0
    if s == "rosca":
        return 1
    if s in {"usinagem", "usf", "recortador", "fresa"}:
        return 2
    if s == "forno":
        return 3
    if s in {"v/ termico", "termico", "mtc - tt", "mtc-tt", "eden"}:
        return 4
    if s in {"v/ superficial", "superficial", "mtc - ts", "mtc-ts", "sipra", "martins", "multstamp", "realtec", "indusmek", "mogi"}:
        return 5
    if s == "nylok":
        return 6
    if s == "aguardando escolha":
        return 7
    if s == "aguardando retorno":
        return 8
    return 9

def ordem_pcp(order: dict) -> tuple:
    return (-atraso_dias(order), setor_rank(order), str(order.get("dataEntregaAtual", "")), str(order.get("cpd", "")))

def export_state_to_xlsx(state: dict, path: Path) -> None:
    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_rel = wb.create_sheet("Relatório")
    ws_diario = wb.create_sheet("Resumo Diário")
    ws_notas = wb.create_sheet("Anotações")
    ws_ped = wb.create_sheet("Pedidos")
    ws_sec = wb.create_sheet("Datas")
    ws_cpd = wb.create_sheet("Base CPDs")
    ws_hist = wb.create_sheet("Histórico")
    ws_audit = wb.create_sheet("Auditoria")
    ws_kpi_resumo = wb.create_sheet("KPI Resumo")
    ws_kpi_setor = wb.create_sheet("KPI por Setor")
    ws_kpi_cliente = wb.create_sheet("KPI por Cliente")
    ws_kpi_cpd = wb.create_sheet("KPI por CPD")
    ws_kpi_prazo = wb.create_sheet("KPI por Prazo")
    ws_kpi_grupo = wb.create_sheet("KPI por Grupo")
    ws_kpi_rota = wb.create_sheet("KPI por Rota")
    ws_powerbi = wb.create_sheet("Base Power BI")

    dark = "262624"
    blue = "2E3FFF"
    light = "ECEAE6"
    header_fill = PatternFill("solid", fgColor=dark)
    header_font = Font(color=light, bold=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    orders = state.get("orders", [])
    sections = state.get("sections", [])
    manual = state.get("manualCPD", {})
    sec_map = {s.get("id"): s for s in sections}

    # Dashboard
    ativos = [o for o in orders if not is_finalizado(o)]
    concluidos = [o for o in orders if compute_status(o) == "concluido"]
    cancelados = [o for o in orders if compute_status(o) == "cancelado"]
    ws_dash["A1"] = "Controle de Pedidos PCP"
    ws_dash["A1"].font = Font(size=16, bold=True, color=dark)
    ws_dash["A2"] = "Última atualização"
    ws_dash["B2"] = parse_datetime_value(state.get("updated_at", ""))
    ws_dash["B2"].number_format = "DD-MM-YYYY HH:MM"
    metrics = [
        ("Total ativos", len(ativos)),
        ("Concluídos", len(concluidos)),
        ("Cancelados", len(cancelados)),
        ("Total geral", len(orders)),
        ("Vencidos ativos", sum(1 for o in ativos if prazo_label(o) == "Vencido")),
        ("Sem previsão de retorno", sum(1 for o in ativos if not o.get("previsao"))),
        ("HONDA", sum(1 for o in orders if grupo_cliente(o.get("cliente", "")) == "HONDA")),
        ("DIVERSOS", sum(1 for o in orders if grupo_cliente(o.get("cliente", "")) == "DIVERSOS")),
    ]
    ws_dash.append([])
    ws_dash.append(["Indicador", "Valor"])
    for item in metrics:
        ws_dash.append(list(item))
    for cell in ws_dash[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    for row in ws_dash.iter_rows(min_row=5, max_row=4+len(metrics), min_col=1, max_col=2):
        for cell in row:
            cell.border = border
    ws_dash.column_dimensions["A"].width = 26
    ws_dash.column_dimensions["B"].width = 20

    # Pedidos
    headers = [
        "Nº", "CPD", "Descrição", "Setor", "Setor Outro", "Qtd Faltante",
        "Data Entrega Original", "Data Entrega Atual", "Previsão Retorno",
        "Cliente", "Grupo ID", "Data do Grupo", "Data Conclusão",
        "Grupo Cliente", "Observação", "Criado Por", "Criado Em", "Alterado Por", "Alterado Em", "ID Interno",
    ]
    ws_ped.append(headers)
    for idx, o in enumerate(orders, start=1):
        sec = sec_map.get(o.get("secaoId"), {})
        ws_ped.append([
            o.get("numero") or idx,
            o.get("cpd", ""),
            o.get("descricao", ""),
            o.get("setor", ""),
            o.get("setorOutro", ""),
            o.get("qtd", ""),
            parse_date_value(o.get("dataEntregaOriginal", "")),
            parse_date_value(o.get("dataEntregaAtual", "")),
            parse_date_value(o.get("previsao", "")),
            o.get("cliente", ""),
            o.get("secaoId", ""),
            section_label(sec),
            parse_date_value(o.get("dataConclusao", "")),
            grupo_cliente(o.get("cliente", "")),
            o.get("obs", ""),
            o.get("criadoPor", ""),
            parse_datetime_value(o.get("criadoEm", "")),
            o.get("alteradoPor", ""),
            parse_datetime_value(o.get("alteradoEm", "")),
            o.get("id", ""),
        ])

    # Datas: a equipe enxerga somente a data.
    ws_sec.append(["ID", "Data", "Ativa", "Criado Por", "Criado Em", "Alterado Por", "Alterado Em"])
    for sct in sections:
        ws_sec.append([
            sct.get("id", ""),
            parse_date_value(sct.get("defaultDate", "")),
            "Sim" if sct.get("active") else "Não",
            sct.get("criadoPor", ""),
            parse_datetime_value(sct.get("criadoEm", "")),
            sct.get("alteradoPor", ""),
            parse_datetime_value(sct.get("alteradoEm", "")),
        ])

    # Base de CPDs
    ws_cpd.append(["Código", "Descrição"])
    for codigo, descricao in sorted(manual.items()):
        ws_cpd.append([codigo, descricao])

    # Histórico
    ws_hist.append(["Pedido ID", "Nº", "CPD", "Tipo", "De", "Para", "Quando", "Usuário", "Motivo"])
    for o in orders:
        for h in o.get("historico") or []:
            tipo = h.get("tipo", "entrega") or "entrega"
            de_val = h.get("de", "") if tipo in {"status", "setor"} else parse_date_value(h.get("de", ""))
            para_val = h.get("para", "") if tipo in {"status", "setor"} else parse_date_value(h.get("para", ""))
            ws_hist.append([
                o.get("id", ""),
                o.get("numero", ""),
                o.get("cpd", ""),
                tipo,
                de_val,
                para_val,
                parse_date_value(h.get("quando", "")),
                h.get("usuario", ""),
                h.get("motivo", ""),
            ])

    # Auditoria geral de alterações, com explicação simples para virar relatório.
    ws_audit.append(["ID", "Quando", "Usuário", "O que foi feito", "Ação técnica", "Tipo", "ID/Registro", "IP", "Detalhes técnicos"])
    audit_rows = load_audit_rows()
    for r in audit_rows:
        ws_audit.append(resumo_auditoria_linha(r))


    # Abas analíticas para Excel/Power BI
    def add_table(ws, headers, rows):
        ws.append(headers)
        for r in rows:
            ws.append(r)

    def count_by(items, key_func):
        d = {}
        for item in items:
            key = key_func(item) or "(vazio)"
            d[key] = d.get(key, 0) + 1
        return sorted(d.items(), key=lambda x: (-x[1], str(x[0])))

    ativos = [o for o in orders if not is_finalizado(o)]
    concluidos = [o for o in orders if compute_status(o) == "concluido"]
    no_prazo_concl = sum(1 for o in concluidos if o.get("dataConclusao") and str(o.get("dataConclusao")) <= str(o.get("dataEntregaAtual")))
    pct_no_prazo = (no_prazo_concl / len(concluidos)) if concluidos else 0
    add_table(ws_kpi_resumo, ["Indicador", "Valor"], [
        ["Total geral", len(orders)],
        ["Pedidos ativos", len(ativos)],
        ["Concluídos", len(concluidos)],
        ["Cancelados", sum(1 for o in orders if compute_status(o) == "cancelado")],
        ["Vencidos ativos", sum(1 for o in ativos if prazo_label(o) == "Vencido")],
        ["Esta semana", sum(1 for o in ativos if prazo_label(o) == "Esta semana")],
        ["Sem previsão de retorno", sum(1 for o in ativos if not o.get("previsao"))],
        ["% no prazo", pct_no_prazo],
        ["Pedidos postergados", sum(1 for o in orders if str(o.get("dataEntregaAtual", "")) > str(o.get("dataEntregaOriginal", "")))],
        ["Pedidos adiantados", sum(1 for o in orders if str(o.get("dataEntregaAtual", "")) < str(o.get("dataEntregaOriginal", "")))],
    ])
    add_table(ws_kpi_setor, ["Setor", "Pedidos ativos"], count_by(ativos, lambda o: o.get("setor")))
    add_table(ws_kpi_cliente, ["Cliente", "Pedidos", "Ativos", "Vencidos", "Sem previsão"], [
        [cliente,
         sum(1 for o in orders if (o.get("cliente") or "(vazio)") == cliente),
         sum(1 for o in ativos if (o.get("cliente") or "(vazio)") == cliente),
         sum(1 for o in ativos if (o.get("cliente") or "(vazio)") == cliente and prazo_label(o) == "Vencido"),
         sum(1 for o in ativos if (o.get("cliente") or "(vazio)") == cliente and not o.get("previsao"))]
        for cliente, _ in count_by(orders, lambda o: o.get("cliente"))
    ])
    add_table(ws_kpi_cpd, ["CPD", "Descrição", "Ocorrências", "Qtd faltante total", "Vencidos ativos"], [
        [cpd,
         next((o.get("descricao", "") for o in orders if (o.get("cpd") or "(vazio)") == cpd), ""),
         sum(1 for o in orders if (o.get("cpd") or "(vazio)") == cpd),
         sum(to_int(o.get("qtd"), 0) or 0 for o in orders if (o.get("cpd") or "(vazio)") == cpd),
         sum(1 for o in ativos if (o.get("cpd") or "(vazio)") == cpd and prazo_label(o) == "Vencido")]
        for cpd, _ in count_by(orders, lambda o: o.get("cpd"))
    ])
    add_table(ws_kpi_prazo, ["Status de prazo", "Pedidos ativos"], count_by(ativos, prazo_label))
    add_table(ws_kpi_grupo, ["Grupo", "Pedidos"], count_by(orders, lambda o: grupo_cliente(o.get("cliente", ""))))

    def rota_processo_excel(o):
        s = normalize_text(status_label(o))
        if s in {"prensa", "rosca", "usinagem", "usf", "recortador", "fresa"}:
            return "Interno inicial"
        if s in {"forno", "v/ termico", "mtc - tt", "mtc-tt", "eden"}:
            return "Tratamento térmico"
        if s in {"v/ superficial", "mtc - ts", "mtc-ts", "sipra", "martins", "multstamp", "realtec", "indusmek", "mogi"}:
            return "Tratamento superficial"
        if s == "nylok":
            return "Nylok"
        if s in {"aguardando retorno", "aguardando escolha"}:
            return "Aguardando"
        if s in {"concluido", "concluído"}:
            return "Concluído"
        if s == "cancelado":
            return "Cancelado"
        return "Outros"

    add_table(ws_kpi_rota, ["Rota do processo", "Pedidos ativos"], count_by(ativos, rota_processo_excel))
    add_table(ws_powerbi, ["Nº", "CPD", "Descrição", "Setor/Situação", "Cliente", "Grupo Cliente", "Data Entrega", "Previsão Retorno", "Data Conclusão", "Prazo", "Atraso Dias", "Qtd Faltante", "Data Grupo", "Criado Por", "Criado Em", "Alterado Por", "Alterado Em", "ID Interno"], [
        [o.get("numero", ""), o.get("cpd", ""), o.get("descricao", ""), o.get("setor", ""), o.get("cliente", ""), grupo_cliente(o.get("cliente", "")), parse_date_value(o.get("dataEntregaAtual", "")), parse_date_value(o.get("previsao", "")), parse_date_value(o.get("dataConclusao", "")), prazo_label(o), atraso_dias(o), to_int(o.get("qtd", 0)) or 0, section_label(sec_map.get(o.get("secaoId", ""), {})), o.get("criadoPor", ""), parse_datetime_value(o.get("criadoEm", "")), o.get("alteradoPor", ""), parse_datetime_value(o.get("alteradoEm", "")), o.get("id", "")]
        for o in orders
    ])


    # Relatório executivo: leitura simples para reunião diária.
    ws_rel["A1"] = "Relatório PCP - Controle de Pedidos"
    ws_rel["A1"].font = Font(size=18, bold=True, color="ECEAE6")
    ws_rel["A1"].fill = PatternFill("solid", fgColor="262624")
    ws_rel["A2"] = "Atualizado em"
    ws_rel["B2"] = parse_datetime_value(state.get("updated_at", ""))
    ws_rel["B2"].number_format = "DD-MM-YYYY HH:MM"
    ws_rel["A4"] = "Leitura rápida"
    ws_rel["A4"].font = Font(size=13, bold=True, color="2E3FFF")
    leitura = [
        f"Existem {len(ativos)} pedido(s) ativo(s), {len(concluidos)} concluído(s) e {len(cancelados)} cancelado(s).",
        f"Pedidos vencidos ativos: {sum(1 for o in ativos if prazo_label(o) == 'Vencido')}.",
        f"Pedidos sem previsão de retorno: {sum(1 for o in ativos if not o.get('previsao'))}.",
        f"Itens HONDA: {sum(1 for o in orders if grupo_cliente(o.get('cliente', '')) == 'HONDA')} | DIVERSOS: {sum(1 for o in orders if grupo_cliente(o.get('cliente', '')) == 'DIVERSOS')}.",
        "A aba Base Power BI foi preparada para conexão futura no Power BI.",
        "A aba Auditoria traz uma explicação simples do que foi feito em cada movimentação.",
    ]
    for i, linha in enumerate(leitura, start=5):
        ws_rel.cell(i, 1).value = "• " + linha
    ws_rel["A13"] = "Pontos de atenção"
    ws_rel["A13"].font = Font(size=13, bold=True, color="F90202")
    pontos = []
    if sum(1 for o in ativos if prazo_label(o) == 'Vencido'):
        pontos.append("Priorizar pedidos ativos vencidos, principalmente os com maior atraso em dias.")
    if sum(1 for o in ativos if not o.get('previsao')):
        pontos.append("Preencher previsão de retorno dos pedidos sem informação.")
    if not pontos:
        pontos.append("Nenhum ponto crítico automático identificado no momento.")
    for i, linha in enumerate(pontos, start=14):
        ws_rel.cell(i, 1).value = "• " + linha

    # Dashboard visual no Excel com cartões e gráficos.
    ws_dash.delete_rows(1, ws_dash.max_row)
    ws_dash.sheet_view.showGridLines = False
    ws_dash["A1"] = "Dashboard PCP"
    ws_dash["A1"].font = Font(size=20, bold=True, color="ECEAE6")
    ws_dash["A1"].fill = PatternFill("solid", fgColor="262624")
    ws_dash.merge_cells("A1:H1")
    ws_dash["A2"] = "Atualização"
    ws_dash["B2"] = parse_datetime_value(state.get("updated_at", ""))
    ws_dash["B2"].number_format = "DD-MM-YYYY HH:MM"
    cards = [
        ("Ativos", len(ativos), "2E3FFF"),
        ("Vencidos", sum(1 for o in ativos if prazo_label(o) == "Vencido"), "F90202"),
        ("Sem previsão", sum(1 for o in ativos if not o.get("previsao")), "FBBF24"),
        ("Concluídos", len(concluidos), "16A34A"),
    ]
    card_positions = [("A4", "B5"), ("C4", "D5"), ("E4", "F5"), ("G4", "H5")]
    for (titulo, valor, cor), (tl, br) in zip(cards, card_positions):
        ws_dash[tl] = titulo
        ws_dash[tl].font = Font(bold=True, color="ECEAE6", size=11)
        ws_dash[tl].fill = PatternFill("solid", fgColor=cor)
        ws_dash[tl].alignment = Alignment(horizontal="center")
        val_cell = tl[0] + "5"
        ws_dash[val_cell] = valor
        ws_dash[val_cell].font = Font(bold=True, color="262624", size=18)
        ws_dash[val_cell].alignment = Alignment(horizontal="center")
        ws_dash[val_cell].fill = PatternFill("solid", fgColor="ECEAE6")
        ws_dash.merge_cells(f"{tl}:{chr(ord(tl[0])+1)}4")
        ws_dash.merge_cells(f"{val_cell}:{chr(ord(val_cell[0])+1)}5")

    # Segunda linha de cartões com visão executiva adicional.
    extra_cards = [
        ("HONDA", sum(1 for o in orders if grupo_cliente(o.get("cliente", "")) == "HONDA"), "3A1111"),
        ("DIVERSOS", sum(1 for o in orders if grupo_cliente(o.get("cliente", "")) == "DIVERSOS"), "3A3936"),
        ("Trat. térmico", sum(1 for o in ativos if rota_processo_excel(o) == "Tratamento térmico"), "B45309"),
        ("Trat. superficial", sum(1 for o in ativos if rota_processo_excel(o) == "Tratamento superficial"), "7C3AED"),
    ]
    card_positions2 = [("A7", "B8"), ("C7", "D8"), ("E7", "F8"), ("G7", "H8")]
    for (titulo, valor, cor), (tl, br) in zip(extra_cards, card_positions2):
        ws_dash[tl] = titulo
        ws_dash[tl].font = Font(bold=True, color="ECEAE6", size=11)
        ws_dash[tl].fill = PatternFill("solid", fgColor=cor)
        ws_dash[tl].alignment = Alignment(horizontal="center")
        val_cell = tl[0] + "8"
        ws_dash[val_cell] = valor
        ws_dash[val_cell].font = Font(bold=True, color="262624", size=18)
        ws_dash[val_cell].alignment = Alignment(horizontal="center")
        ws_dash[val_cell].fill = PatternFill("solid", fgColor="ECEAE6")
        ws_dash.merge_cells(f"{tl}:{chr(ord(tl[0])+1)}7")
        ws_dash.merge_cells(f"{val_cell}:{chr(ord(val_cell[0])+1)}8")

    ws_dash["A11"] = "Resumo para análise"
    ws_dash["A11"].font = Font(size=13, bold=True, color="2E3FFF")
    ws_dash["A12"] = "Esta aba foi refeita para Office 2021: não usa Tabelas estruturadas nem AutoFiltro automático, evitando reparos do Excel. Para análises mais interativas, conecte a aba Base Power BI no Power BI."
    ws_dash["A12"].alignment = Alignment(wrap_text=True)
    ws_dash.merge_cells("A12:H13")

    def safe_add_chart(chart, ws, anchor):
        try:
            ws.add_chart(chart, anchor)
        except Exception:
            pass

    # Chart: Prazo
    if ws_kpi_prazo.max_row >= 3:
        chart = PieChart()
        chart.title = "Pedidos ativos por prazo"
        labels = Reference(ws_kpi_prazo, min_col=1, min_row=2, max_row=ws_kpi_prazo.max_row)
        data = Reference(ws_kpi_prazo, min_col=2, min_row=1, max_row=ws_kpi_prazo.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.height = 7
        chart.width = 10
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        safe_add_chart(chart, ws_dash, "A15")

    # Chart: Setor/Situação
    if ws_kpi_setor.max_row >= 3:
        chart = BarChart()
        chart.title = "Ativos por Setor/Situação"
        chart.y_axis.title = "Pedidos"
        chart.x_axis.title = "Setor/Situação"
        data = Reference(ws_kpi_setor, min_col=2, min_row=1, max_row=ws_kpi_setor.max_row)
        cats = Reference(ws_kpi_setor, min_col=1, min_row=2, max_row=ws_kpi_setor.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 12
        safe_add_chart(chart, ws_dash, "D15")

    # Chart: cliente
    if ws_kpi_cliente.max_row >= 3:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Pedidos por cliente"
        chart.y_axis.title = "Cliente"
        chart.x_axis.title = "Pedidos"
        data = Reference(ws_kpi_cliente, min_col=2, min_row=1, max_row=min(ws_kpi_cliente.max_row, 12))
        cats = Reference(ws_kpi_cliente, min_col=1, min_row=2, max_row=min(ws_kpi_cliente.max_row, 12))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 14
        safe_add_chart(chart, ws_dash, "A31")

    # Chart: CPD top 10
    if ws_kpi_cpd.max_row >= 3:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Top CPDs por ocorrência"
        chart.x_axis.title = "Ocorrências"
        chart.y_axis.title = "CPD"
        maxr = min(ws_kpi_cpd.max_row, 11)
        data = Reference(ws_kpi_cpd, min_col=3, min_row=1, max_row=maxr)
        cats = Reference(ws_kpi_cpd, min_col=1, min_row=2, max_row=maxr)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 14
        safe_add_chart(chart, ws_dash, "D31")

    # Chart: HONDA x DIVERSOS
    if ws_kpi_grupo.max_row >= 3:
        chart = PieChart()
        chart.title = "HONDA x DIVERSOS"
        labels = Reference(ws_kpi_grupo, min_col=1, min_row=2, max_row=ws_kpi_grupo.max_row)
        data = Reference(ws_kpi_grupo, min_col=2, min_row=1, max_row=ws_kpi_grupo.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.height = 7
        chart.width = 10
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        safe_add_chart(chart, ws_dash, "A50")

    # Chart: rota do processo
    if ws_kpi_rota.max_row >= 3:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Pedidos ativos por rota do processo"
        chart.x_axis.title = "Pedidos"
        chart.y_axis.title = "Rota"
        data = Reference(ws_kpi_rota, min_col=2, min_row=1, max_row=ws_kpi_rota.max_row)
        cats = Reference(ws_kpi_rota, min_col=1, min_row=2, max_row=ws_kpi_rota.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 14
        safe_add_chart(chart, ws_dash, "D50")

    # Estética geral do dashboard/relatório.
    for col in range(1, 9):
        ws_dash.column_dimensions[get_column_letter(col)].width = 16
    for row in range(1, 50):
        ws_dash.row_dimensions[row].height = 20
    ws_rel.sheet_view.showGridLines = False
    ws_rel.column_dimensions["A"].width = 90
    ws_rel.column_dimensions["B"].width = 24
    for row in ws_rel.iter_rows(min_row=1, max_row=max(ws_rel.max_row, 16), min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


    # Resumo diário do PCP: visão pronta para reunião diária.
    hoje_txt = hoje_iso()
    audit_rows = load_audit_rows(5000)
    audit_hoje = [r for r in audit_rows if str(r.get("created_at", ""))[:10] == hoje_txt]
    novos_hoje = [o for o in orders if str(o.get("criadoEm", ""))[:10] == hoje_txt]
    alterados_hoje = [o for o in orders if str(o.get("alteradoEm", ""))[:10] == hoje_txt]
    concluidos_hoje = [o for o in orders if str(o.get("dataConclusao", ""))[:10] == hoje_txt]
    ws_diario.sheet_view.showGridLines = False
    ws_diario["A1"] = "Resumo diário do PCP"
    ws_diario["A1"].font = Font(size=18, bold=True, color="ECEAE6")
    ws_diario["A1"].fill = PatternFill("solid", fgColor="262624")
    ws_diario.merge_cells("A1:F1")
    ws_diario["A2"] = "Data do relatório"
    ws_diario["B2"] = date.today()
    ws_diario["B2"].number_format = "DD-MM-YYYY"
    diario_metrics = [
        ("Pedidos novos hoje", len(novos_hoje)),
        ("Pedidos alterados hoje", len(alterados_hoje)),
        ("Pedidos concluídos hoje", len(concluidos_hoje)),
        ("Movimentações registradas hoje", len(audit_hoje)),
        ("Pedidos vencidos ativos", sum(1 for o in ativos if prazo_label(o) == "Vencido")),
        ("Pedidos sem previsão", sum(1 for o in ativos if not o.get("previsao"))),
        ("Aguardando retorno", sum(1 for o in ativos if status_label(o) == "Aguardando retorno")),
        ("Aguardando escolha", sum(1 for o in ativos if status_label(o) == "Aguardando escolha")),
        ("V/ TÉRMICO / MTC - TT / ÉDEN", sum(1 for o in ativos if normalize_text(status_label(o)) in {"v/ termico", "mtc - tt", "mtc-tt", "eden"})),
        ("V/ SUPERFICIAL / MTC - TS", sum(1 for o in ativos if normalize_text(status_label(o)) in {"v/ superficial", "mtc - ts", "mtc-ts", "sipra", "martins", "multstamp", "realtec", "indusmek"})),
        ("Nylok", sum(1 for o in ativos if normalize_text(status_label(o)) == "nylok")),
    ]
    ws_diario.append([])
    ws_diario.append(["Indicador", "Valor", "Leitura prática"])
    leitura = {
        "Pedidos novos hoje": "Entradas adicionadas no controle durante o dia.",
        "Pedidos alterados hoje": "Itens com alguma atualização de data, previsão, setor/situação ou dados principais.",
        "Pedidos concluídos hoje": "Itens finalizados hoje.",
        "Movimentações registradas hoje": "Total de ações capturadas na auditoria.",
        "Pedidos vencidos ativos": "Atenção imediata: pedidos com entrega anterior à data atual.",
        "Pedidos sem previsão": "Itens que precisam de previsão de retorno preenchida.",
        "Aguardando retorno": "Itens aguardando retorno de área/fornecedor/processo.",
        "Aguardando escolha": "Itens aguardando etapa de escolha/seleção.",
        "V/ TÉRMICO / MTC - TT / ÉDEN": "Itens em rota de tratamento térmico, interno ou externo.",
        "V/ SUPERFICIAL / MTC - TS": "Itens em rota de tratamento superficial, interno ou externo.",
        "Nylok": "Itens em processo de aplicação de trava nos parafusos auto-tarrachantes.",
    }
    for k, v in diario_metrics:
        ws_diario.append([k, v, leitura.get(k, "")])
    ws_diario.append([])
    ws_diario.append(["Ações recomendadas"])
    linha_ini = ws_diario.max_row + 1
    if sum(1 for o in ativos if prazo_label(o) == "Vencido"):
        ws_diario.append(["Priorizar tratativa dos pedidos vencidos."])
    if sum(1 for o in ativos if not o.get("previsao")):
        ws_diario.append(["Preencher previsão de retorno dos pedidos sem informação."])
    if not audit_hoje:
        ws_diario.append(["Sem movimentações registradas hoje até o momento."])
    ws_diario.append([])
    ws_diario.append(["Últimas movimentações de hoje", "Usuário", "O que foi feito"])
    for r in audit_hoje[-20:]:
        ws_diario.append([parse_datetime_value(r.get("created_at", "")), r.get("actor", ""), audit_explicacao(r)])
    for cell in ws_diario[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    for cell in ws_diario[ws_diario.max_row-len(audit_hoje[-20:]) if audit_hoje else ws_diario.max_row]:
        pass



    # Anotações: relatório operacional simples com fila de atenção.
    ws_notas.sheet_view.showGridLines = False
    ws_notas["A1"] = "Anotações PCP - Fila de atenção"
    ws_notas["A1"].font = Font(size=18, bold=True, color="ECEAE6")
    ws_notas["A1"].fill = PatternFill("solid", fgColor="262624")
    ws_notas.merge_cells("A1:H1")
    ws_notas["A2"] = "Leitura rápida"
    ws_notas["A2"].font = Font(size=13, bold=True, color="2E3FFF")
    vencidos_ord = sorted([o for o in ativos if atraso_dias(o) > 0], key=ordem_pcp)
    texto_notas = [
        f"Pedidos vencidos ativos: {len(vencidos_ord)}.",
        "A ordem PCP considera: maior atraso em dias e, em empate, Prensa, Rosca, Usinagem/USF/Recortador/Fresa, Forno, V/ TÉRMICO/MTC - TT/Éden, V/ SUPERFICIAL/MTC - TS/Sipra/Martins/Multstamp/Realtec/Indusmek/Mogi, Nylok e Aguardando escolha.",
        "Use esta aba como pauta rápida para reunião diária e tratativas urgentes."
    ]
    r = 3
    for linha in texto_notas:
        ws_notas.cell(r, 1).value = "• " + linha
        r += 1
    r += 1
    ws_notas.cell(r, 1).value = "Mais atrasados - ordem PCP"
    ws_notas.cell(r, 1).font = Font(size=13, bold=True, color="F90202")
    r += 1
    cab = ["CPD", "Descrição", "Setor/Situação", "Cliente", "Entrega", "Atraso dias", "Previsão", "Observação"]
    for c, h in enumerate(cab, start=1):
        ws_notas.cell(r, c).value = h
    for o in vencidos_ord[:25]:
        r += 1
        ws_notas.cell(r, 1).value = o.get("cpd", "")
        ws_notas.cell(r, 2).value = o.get("descricao", "")
        ws_notas.cell(r, 3).value = o.get("setor", "")
        ws_notas.cell(r, 4).value = o.get("cliente", "")
        ws_notas.cell(r, 5).value = parse_date_value(o.get("dataEntregaAtual", ""))
        ws_notas.cell(r, 6).value = atraso_dias(o)
        ws_notas.cell(r, 7).value = parse_date_value(o.get("previsao", ""))
        ws_notas.cell(r, 8).value = o.get("obs", "")


    # Compatibilidade Office 2021:
    # Não usamos Tabelas estruturadas nem AutoFiltro automático aqui.
    # Algumas instalações do Excel 2021 reparam/removem table*.xml quando o arquivo
    # é gerado por biblioteca externa. Mantemos tudo como intervalos formatados
    # e gráficos nativos, preservando abertura sem reparos.

    date_cols = {
        "Pedidos": {7: "DD-MM-YYYY", 8: "DD-MM-YYYY", 9: "DD-MM-YYYY", 13: "DD-MM-YYYY", 17: "DD-MM-YYYY HH:MM", 19: "DD-MM-YYYY HH:MM"},
        "Datas": {2: "DD-MM-YYYY", 5: "DD-MM-YYYY HH:MM", 7: "DD-MM-YYYY HH:MM"},
        "Histórico": {7: "DD-MM-YYYY"},
        "Auditoria": {2: "DD-MM-YYYY HH:MM"},
        "Relatório": {2: "DD-MM-YYYY HH:MM"},
        "Resumo Diário": {2: "DD-MM-YYYY", 18: "DD-MM-YYYY HH:MM"},
        "Anotações": {5: "DD-MM-YYYY", 7: "DD-MM-YYYY"},
        "KPI Resumo": {},
        "Base Power BI": {7: "DD-MM-YYYY", 8: "DD-MM-YYYY", 9: "DD-MM-YYYY", 15: "DD-MM-YYYY HH:MM", 17: "DD-MM-YYYY HH:MM"},
    }

    widths = {
        "Pedidos": [8, 12, 42, 20, 18, 14, 18, 18, 18, 22, 18, 18, 18, 16, 38, 18, 20, 18, 20, 22],
        "Datas": [18, 18, 10, 18, 20, 18, 20],
        "Base CPDs": [14, 60],
        "Histórico": [18, 8, 12, 14, 18, 18, 14, 18, 50],
        "Auditoria": [8, 20, 18, 60, 16, 14, 22, 16, 70],
        "Relatório": [26, 70, 24, 24],
        "Resumo Diário": [34, 14, 64, 20, 18, 70],
        "Anotações": [14, 50, 22, 24, 14, 14, 14, 45],
        "KPI Resumo": [32, 18],
        "KPI por Setor": [28, 16],
        "KPI por Cliente": [32, 16, 16, 16, 18],
        "KPI por CPD": [14, 42, 16, 18, 16],
        "KPI por Prazo": [22, 16],
        "KPI por Grupo": [22, 16],
        "KPI por Rota": [30, 16],
        "Base Power BI": [8, 12, 42, 22, 24, 16, 16, 18, 18, 16, 14, 14, 18, 18, 20, 18, 20, 22],
    }

    for ws in [ws_rel, ws_diario, ws_notas, ws_ped, ws_sec, ws_cpd, ws_hist, ws_audit, ws_kpi_resumo, ws_kpi_setor, ws_kpi_cliente, ws_kpi_cpd, ws_kpi_prazo, ws_kpi_grupo, ws_kpi_rota, ws_powerbi]:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, fmt in date_cols.get(ws.title, {}).items():
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = fmt
        for i, width in enumerate(widths[ws.title], start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        # Sem AutoFiltro automático para evitar reparos no Excel 2021.

    # Mantém IDs técnicos escondidos: a equipe vê o Nº sequencial e a data do grupo.
    if "ID Interno" in headers:
        ws_ped.column_dimensions[get_column_letter(headers.index("ID Interno") + 1)].hidden = True
    if "Grupo ID" in headers:
        ws_ped.column_dimensions[get_column_letter(headers.index("Grupo ID") + 1)].hidden = True
    ws_sec.column_dimensions["A"].hidden = True

    wb.save(path)


@app.get("/api/export.xlsx")
def api_export_xlsx():
    if Workbook is None:
        return jsonify({"error": "openpyxl não está instalado"}), 500
    state = load_state()
    tmp = Path(tempfile.gettempdir()) / f"dados_pcp_{uuid.uuid4().hex}.xlsx"
    export_state_to_xlsx(state, tmp)
    return send_file(tmp, as_attachment=True, download_name="dados_pcp.xlsx")


def export_pendentes_to_xlsx(state: dict, path: Path) -> None:
    """Exporta os pedidos ainda não finalizados, na formatação da planilha follow (cabeçalho azul claro, agrupado por data)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos Pendentes"

    header_fill = PatternFill("solid", fgColor="C6D9F1")
    header_font = Font(name="Arial", size=11, bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    medium = Side(style="medium")
    header_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    date_font = Font(bold=True)

    orders = state.get("orders", [])
    sections = state.get("sections", [])
    sec_map = {s.get("id"): s for s in sections}

    pendentes = [o for o in orders if not is_finalizado(o)]

    def sort_key(o):
        sec = sec_map.get(o.get("secaoId"), {})
        d = parse_date_value(sec.get("defaultDate"))
        d_ord = d if isinstance(d, date) else date.max
        try:
            numero = int(o.get("numero") or 0)
        except (TypeError, ValueError):
            numero = 0
        return (d_ord, numero)

    pendentes.sort(key=sort_key)

    headers = [
        "ITEM", "DATA DE ENTREGA", "CPD", "STATUS",
        "QUANTIDADE FALTANTE", "PREVISÃO DE RETORNO", "CLIENTE", "PRAZO", "OBS.",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = header_border

    widths = [8, 14, 10, 16, 14, 15, 18, 12, 32]
    for col, w in zip("ABCDEFGHI", widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    row_idx = 2
    last_sec = object()
    for o in pendentes:
        sec_key = o.get("secaoId")
        if sec_key != last_sec:
            sec = sec_map.get(sec_key)
            label = section_label(sec) if sec else "Sem data"
            banner = ws.cell(row=row_idx, column=1, value=label)
            banner.font = date_font
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
            row_idx += 1
            last_sec = sec_key
        values = [
            o.get("numero", ""),
            parse_date_value(o.get("dataEntregaAtual", "")),
            o.get("cpd", ""),
            status_label(o),
            o.get("qtd", ""),
            parse_date_value(o.get("previsao", "")),
            o.get("cliente", ""),
            prazo_label(o),
            o.get("obs", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if col_idx in (2, 6) and isinstance(val, date):
                cell.number_format = "DD-MM-YYYY"
        row_idx += 1

    wb.save(path)


@app.get("/api/export_pendentes.xlsx")
def api_export_pendentes_xlsx():
    if Workbook is None:
        return jsonify({"error": "openpyxl não está instalado"}), 500
    state = load_state()
    tmp = Path(tempfile.gettempdir()) / f"pedidos_pendentes_{uuid.uuid4().hex}.xlsx"
    export_pendentes_to_xlsx(state, tmp)
    return send_file(tmp, as_attachment=True, download_name="pedidos_pendentes.xlsx")


@app.get("/")
@app.get("/pcp_prototype.html")
@app.get("/pcp_prototype_sqlite.html")
def pagina_principal():
    if not HTML_FILE.exists():
        return Response("Arquivo pcp_prototype.html não encontrado.", status=404, mimetype="text/plain")
    return Response(HTML_FILE.read_text(encoding="utf-8"), mimetype="text/html; charset=utf-8")


def ensure_existing_order_metadata() -> None:
    """Preenche campos de auditoria em pedidos antigos sem apagar dados existentes."""
    with lock, conn() as c:
        changed = False
        rows = c.execute("SELECT id,payload,updated_at FROM orders").fetchall()
        for row in rows:
            payload = normalize_order(json_loads(row["payload"], {}))
            touched = False
            if not payload.get("criadoPor"):
                payload["criadoPor"] = "Sistema/legado"
                touched = True
            if not payload.get("criadoEm"):
                payload["criadoEm"] = row["updated_at"] or agora()
                touched = True
            if not payload.get("alteradoPor"):
                payload["alteradoPor"] = payload.get("criadoPor") or "Sistema/legado"
                touched = True
            if not payload.get("alteradoEm"):
                payload["alteradoEm"] = row["updated_at"] or payload.get("criadoEm") or agora()
                touched = True
            if touched:
                c.execute("UPDATE orders SET payload=? WHERE id=?", (json_dumps(payload), row["id"]))
                changed = True
        if changed:
            bump_version(c)
        c.commit()


def verificar_e_avisar_migracao_pendente() -> dict:
    """Detecta migração de schema pendente e só registra aviso no log.

    Nunca altera o banco. Usada no início do servidor quando
    deve_aplicar_migracoes_automaticamente() é False, para que quem opera o
    sistema saiba que há uma migração esperando uma decisão explícita
    (PCP_APLICAR_MIGRACOES=1), em vez de simplesmente não dizer nada.
    """
    with lock, conn() as c:
        versao_atual = db_manutencao.get_schema_version(c)
    pendente = versao_atual in db_manutencao.MIGRATIONS
    if pendente:
        log(
            f"[migracao] Há migração pendente (schema_version={versao_atual} de "
            f"{db_manutencao.LATEST_SCHEMA_VERSION}), mas PCP_APLICAR_MIGRACOES não "
            "está definida como '1'; nada foi alterado. Defina PCP_APLICAR_MIGRACOES=1 "
            "e reinicie para aplicar deliberadamente."
        )
    return {"pendente": pendente, "versao_atual": versao_atual}


def deve_importar_cpds_na_inicializacao() -> bool:
    """Só reimporta CPDs do Excel automaticamente ao iniciar se
    PCP_IMPORTAR_CPDS_INICIALIZACAO=1.

    Por padrão (variável ausente ou com qualquer outro valor, incluindo
    "0") o servidor NÃO abre o Excel nem chama import_cpds_from_excel_file()
    sozinho — evita criar uma auditoria de import a cada reinício mesmo
    quando nenhum CPD mudou. Importação controlada continua disponível via
    scripts/importar_cpds_excel.py.
    """
    return os.environ.get("PCP_IMPORTAR_CPDS_INICIALIZACAO") == "1"


def executar_importacao_cpds_na_inicializacao() -> dict:
    """Chamada pelo __main__ na inicialização: só abre o Excel e chama
    import_cpds_from_excel_file() se deve_importar_cpds_na_inicializacao()
    for True. Caso contrário, só registra no log que está desativada — não
    toca no Excel, não importa nada, não cria auditoria. Para importar de
    forma controlada, use scripts/importar_cpds_excel.py.
    """
    if not deve_importar_cpds_na_inicializacao():
        log("Importação automática de CPDs desativada.")
        return {"importou": False}
    if EXCEL_FILE.exists() and load_workbook is not None:
        try:
            resultado = import_cpds_from_excel_file(EXCEL_FILE, actor="Sistema/Inicialização")
            return {"importou": True, "cpds": resultado["imported"]["cpds"]}
        except Exception:
            log("Não consegui recarregar CPDs do Excel na inicialização:\n" + traceback.format_exc())
            return {"importou": False, "erro": True}
    return {"importou": False, "motivo": "excel_ausente_ou_openpyxl_indisponivel"}


if __name__ == "__main__":
    init_db()
    if deve_aplicar_migracoes_automaticamente():
        run_pending_migrations()
    else:
        verificar_e_avisar_migracao_pendente()
    ensure_existing_order_numbers()
    ensure_existing_order_metadata()
    executar_importacao_cpds_na_inicializacao()
    # Gera um Excel inicial caso ainda não exista.
    if not EXCEL_FILE.exists():
        save_excel_snapshot_safely()
    run_daily_backup(False)
    print("")
    print("Servidor PCP iniciado.")
    print(f"Banco SQLite: {DB_FILE}")
    print(f"Excel sincronizado: {EXCEL_FILE}")
    print(f"Acesse neste computador: http://localhost:{PORT}")
    print(f"Na equipe, use: http://IP-DA-MAQUINA:{PORT}")
    print("")
    app.run(host="0.0.0.0", port=PORT, debug=False, threaded=True)
