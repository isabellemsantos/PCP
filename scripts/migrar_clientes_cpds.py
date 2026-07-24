"""Migração definitiva de clientes e CPDs para as tabelas novas (clientes,
cliente_grupos, cpds, cpd_variacoes, cliente_cpds, grupo_cpds,
cpd_descricoes_fontes, cpd_pendencias_revisao), com rastreabilidade total de
onde cada descrição veio e pendências para tudo que não pôde ser decidido
automaticamente.

Reaproveita os módulos de análise já existentes (scripts/analisar_migracao_clientes_cpds.py
e scripts/analisar_lista_oficial_cpds.py) como fonte de verdade dos fatos;
este módulo só decide o que fazer com esses fatos e (no modo --confirmar)
grava.

Dois modos:
  --simular    (padrão): só calcula e reporta o que SERIA feito. Nada é
               gravado -- a conexão de leitura é aberta em modo somente
               leitura (file:...?mode=ro).
  --confirmar: aplica de verdade, dentro de uma transação, com backup
               obrigatório antes e validação (integrity_check +
               foreign_key_check) depois. Exige uma confirmação textual
               adicional (a string exata "CONFIRMAR MIGRACAO"), além da
               flag --confirmar.

Regras já decididas (não deste script escolher de novo):
- Todos os clientes e todos os CPDs entram ativo=1, desativado_em=NULL.
  Nunca inativado por ausência de pedidos ou por só aparecer na lixeira/
  planilha (mesmo um cliente que só apareça em pedido deletado continua
  ativo).
- Quais clientes formam o grupo Honda e quais os demais (grupo Diversos)
  nunca fica hardcoded aqui -- vem sempre da configuração local carregada
  via --config-clientes (ver scripts/config_clientes_migracao.py).
- Regra oficial de 5 dígitos (substitui o antigo sistema de confiança
  alta/média/baixa): todo CPD-pai tem exatamente 5 dígitos numéricos.
  Uma base de 1 a 5 dígitos é sempre preenchida com zero à esquerda,
  automática e silenciosamente (nunca cria "442" e "00442" como CPDs
  separados -- o canônico é sempre "00442"). Só vira pendência (nunca
  descarta, sempre preserva o valor original): base com mais de 5 dígitos
  ou com letras (CODIGO_INVALIDO), ou célula Excel numérica com valor
  genuinamente não-inteiro, cuja fração pode ser artefato do Excel em vez
  de extensão de verdade (FORMATO_AMBIGUO). Ver `normalizar_cpd_cinco_digitos`.
- Descrição canônica: categorias A/B usam a descrição oficial direto;
  C/D também usam a oficial **provisoriamente**, mas geram pendência
  CONFLITO_DESCRICAO. Nunca sobrescreve manual_cpd nem a descrição
  armazenada nos pedidos -- essas continuam exatamente como estão,
  intactas, nas tabelas legadas.
- As duplicidades da planilha (mesmo código repetido em duas linhas com
  descrições diferentes): as duas linhas viram duas fontes distintas em
  cpd_descricoes_fontes; a canônica só é marcada se já existir uma
  descrição usada no sistema (manual_cpd ou pedido) -- senão fica None;
  sempre gera pendência DUPLICIDADE_LISTA.
"""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys
import traceback
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import db_manutencao as dbm
import scripts.analisar_migracao_clientes_cpds as base_analise
import scripts.analisar_lista_oficial_cpds as lista_oficial
import scripts.config_clientes_migracao as config_clientes
from scripts.config_clientes_migracao import ConfigClientes

DEFAULT_PLANILHA = ROOT / "importacao" / "lista_oficial_cpds.xlsx"
DEFAULT_BANCO = ROOT / "pcp.sqlite3"
DEFAULT_SAIDA = ROOT / "relatorios_migracao"

TEXTO_CONFIRMACAO_ESPERADO = "CONFIRMAR MIGRACAO"

# Nenhum nome real de cliente, alias ou grupo fica hardcoded aqui -- essas
# regras de negócio vêm sempre de um arquivo de configuração local (nunca
# versionado), carregado via --config-clientes / carregar_config_clientes().
# Ver scripts/config_clientes_migracao.py e
# config_local/regras_clientes_migracao.example.json.


def eh_cliente_valido(nome_bruto, placeholders_invalidos: set[str]) -> bool:
    normalizado = str(nome_bruto or "").strip().upper()
    return normalizado not in placeholders_invalidos


def _construir_indice_aliases(aliases: dict[str, list[str]]) -> dict[str, str]:
    """Mapa CHAVE_MAIUSCULA (canônico ou alias) -> nome canônico exato."""
    indice = {}
    for canonico, variantes in aliases.items():
        indice[canonico.strip().upper()] = canonico
        for alias in variantes:
            indice[alias.strip().upper()] = canonico
    return indice


def canonicalizar_cliente(nome_bruto: str, indice_aliases: dict[str, str]) -> str:
    """Nome canônico do cliente, aplicando só os aliases confirmados na
    configuração (comparação case-insensitive). Se não houver alias
    conhecido, devolve o nome original só com espaços nas pontas removidos --
    nunca inventa nem funde outros clientes sozinho.
    """
    nome = str(nome_bruto or "").strip()
    return indice_aliases.get(nome.upper(), nome)


# ---------------------------------------------------------------------------
# Regra oficial: CPD-pai sempre com exatamente 5 dígitos numéricos
# ---------------------------------------------------------------------------
#
# Substitui o antigo sistema de confiança (alta/média/baixa) para decidir SE
# preenche com zero à esquerda: agora o preenchimento é sempre obrigatório e
# automático para qualquer código-base de 1 a 5 dígitos. A confiança deixou
# de servir para essa decisão -- "442" e "00442" são sempre o mesmo código
# ("00442"), sem exceção. O que pode gerar pendência agora é só: formato
# inesperado, código-base com mais de 5 dígitos, extensão não numérica, ou
# ambiguidade genuína de origem (célula Excel numérica com valor não-inteiro).
#
# DECISÃO OFICIAL (barra = extensão): "/" é tratado exatamente como "." --
# "12345/1" e "12345.1" são o MESMO código completo ("12345.1", canônico
# sempre com ponto). Formato aceito: base numérica (1-5 dígitos) + extensão
# numérica opcional separada por "." OU "/". Qualquer outra coisa (letras,
# múltiplos separadores, extensão não numérica, etc.) é inválida.
_CODIGO_5_DIGITOS_RE = re.compile(r"^(\d+)(?:[./](\d+))?$")

# DECISÃO OFICIAL (arruelas): códigos alfanuméricos que NÃO são CPDs.
# Padrão verificado nos códigos reais da lista oficial (nenhum foi assumido
# sem checar): sempre 2 letras + hífen + 2-3 dígitos (ex.: "XY-090"). A
# comparação de letras é case-insensitive na DETECÇÃO (pra nunca perder um
# código só por causa de caixa), mas o valor armazenado (codigo_original/
# codigo_canonico) nunca tem maiúsculas/minúsculas
# ou hífen alterados -- só trim de espaços externos.
_CODIGO_ARRUELA_RE = re.compile(r"^[A-Za-z]{2}-[0-9]{2,3}$")


def _resultado_invalido(texto: str, motivo: str, extensao: str | None = None) -> dict:
    tipo_item = "FORMATO_AMBIGUO" if motivo == "FORMATO_AMBIGUO" else "CODIGO_INVALIDO"
    return {
        "tipo_item": tipo_item, "codigo_original": texto, "codigo_pai": None, "extensao": extensao,
        "codigo_completo": None, "codigo_canonico": None,
        "valido": False, "motivo": motivo, "regra_aplicada": None, "usou_barra": "/" in texto,
    }


def normalizar_cpd_cinco_digitos(codigo) -> dict:
    """Aplica a regra oficial de 5 dígitos a um código JÁ COMO TEXTO (sem
    ambiguidade de origem Excel -- essa checagem é feita antes, por
    `_normalizar_codigo_de_fonte`, que sabe se o código veio de célula
    numérica). Só reconhece o FORMATO de CPD (numérico, com "." ou "/" como
    separador de extensão) -- não sabe nada sobre arruelas.

    Retorna sempre um dict com as chaves:
      tipo_item ("CPD", ou "FORMATO_AMBIGUO"/"CODIGO_INVALIDO" se inválido),
      codigo_original, codigo_pai, codigo_completo, codigo_canonico (igual a
      codigo_completo pra CPD; todos None se inválido), extensao, valido
      (bool), motivo (None se válido, senão uma das constantes de motivo),
      regra_aplicada ("ja_5_digitos" ou "zero_padding", None se inválido),
      usou_barra (bool -- True se o código original usava "/" em vez de ".").

    Nunca converte a extensão para número: zeros à esquerda nela (ex.: em
    "04772.01" ou "04772/01") são sempre preservados como texto. O código
    canônico armazenado usa sempre ponto, independente do separador original.
    """
    texto = str(codigo if codigo is not None else "").strip()

    if texto == "":
        return _resultado_invalido(texto, "CODIGO_VAZIO")

    match = _CODIGO_5_DIGITOS_RE.match(texto)
    if not match:
        return _resultado_invalido(texto, "FORMATO_INESPERADO")

    base, extensao = match.group(1), match.group(2)
    if len(base) > 5:
        return _resultado_invalido(texto, "CODIGO_BASE_LONGO", extensao=extensao)

    pai = base.zfill(5)
    completo = pai if extensao is None else f"{pai}.{extensao}"
    regra = "ja_5_digitos" if len(base) == 5 else "zero_padding"

    return {
        "tipo_item": "CPD", "codigo_original": texto, "codigo_pai": pai, "extensao": extensao,
        "codigo_completo": completo, "codigo_canonico": completo,
        "valido": True, "motivo": None, "regra_aplicada": regra, "usou_barra": "/" in texto,
    }


def classificar_arruela(codigo) -> dict | None:
    """Reconhece o padrão de código de arruela verificado nos dados reais da
    lista oficial (2 letras + hífen + 2-3 dígitos). Devolve None se o texto
    não bater com o padrão -- quem chama decide o que fazer (nunca inventa um
    código de arruela por exclusão/eliminação).

    Nunca transforma maiúsculas/minúsculas nem mexe no hífen -- só trim de
    espaços externos, preservando o código oficial exatamente como veio.
    """
    texto = str(codigo if codigo is not None else "").strip()
    if not _CODIGO_ARRUELA_RE.match(texto):
        return None
    return {
        "tipo_item": "ARRUELA", "codigo_original": texto, "codigo_pai": None, "extensao": None,
        "codigo_completo": texto, "codigo_canonico": texto,
        "valido": True, "motivo": None, "regra_aplicada": None, "usou_barra": False,
    }


def _normalizar_codigo_de_fonte(codigo_bruto, numerico: bool = False) -> dict:
    """Classificador único de um código bruto vindo de qualquer fonte
    (lista oficial, manual_cpd, pedidos): decide se é CPD, ARRUELA,
    FORMATO_AMBIGUO ou CODIGO_INVALIDO -- nunca descarta nada.

    Uma célula NUMÉRICA do Excel com valor NÃO-inteiro (ex.: 442.5) é
    intrinsecamente ambígua -- não dá pra saber se ".5" era uma extensão de
    verdade ou um artefato de formatação/arredondamento do Excel (esse caso
    nunca ocorre com "/", que só aparece em células de texto). Por isso,
    nesse caso específico, NÃO tentamos interpretar como base+extensão:
    marcamos FORMATO_AMBIGUO e preservamos o valor bruto, sem criar CPD
    automaticamente. Células de TEXTO (numerico=False) sempre podem ser
    interpretadas normalmente, mesmo com ponto ou barra.
    """
    texto = str(codigo_bruto if codigo_bruto is not None else "").strip()
    if numerico and "." in texto:
        return _resultado_invalido(texto, "FORMATO_AMBIGUO")

    resultado_cpd = normalizar_cpd_cinco_digitos(texto)
    if resultado_cpd["valido"]:
        return resultado_cpd

    resultado_arruela = classificar_arruela(texto)
    if resultado_arruela is not None:
        return resultado_arruela

    return resultado_cpd


# ---------------------------------------------------------------------------
# Construção do plano (só leitura, nenhuma escrita)
# ---------------------------------------------------------------------------

def construir_plano(conn_leitura: sqlite3.Connection, caminho_planilha: Path, config: ConfigClientes) -> dict:
    orders = base_analise.carregar_orders(conn_leitura)
    manual_cpd_rows = base_analise.carregar_manual_cpd(conn_leitura)
    manual_cpd_desc = {r["codigo"]: r["descricao"] for r in manual_cpd_rows if r["descricao"]}

    resultado_oficial = lista_oficial.executar_analise_lista_oficial(conn_leitura, caminho_planilha, config.grupo_honda)
    consolidacao = resultado_oficial["consolidacao"]
    clientes_oficial = resultado_oficial["clientes"]
    conflitos_tecnicos = resultado_oficial["conflitos_tecnicos"]
    duplicidades = resultado_oficial["duplicidades"]
    leitura = resultado_oficial["leitura"]

    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Regras de negócio de cliente vêm todas da configuração local (nunca
    # hardcoded) -- ver scripts/config_clientes_migracao.py.
    indice_aliases = _construir_indice_aliases(config.aliases)

    def cliente_valido(nome_bruto) -> bool:
        return eh_cliente_valido(nome_bruto, config.placeholders_invalidos)

    def canonico_de(nome_bruto) -> str:
        return canonicalizar_cliente(nome_bruto, indice_aliases)

    # --- Regra oficial: todo CPD-pai tem exatamente 5 dígitos numéricos.
    #     Aplicada, sem exceção, a TODO código bruto encontrado nas 3 fontes
    #     (lista oficial, manual_cpd, pedidos) -- substitui por completo o
    #     antigo sistema de confiança (alta/média/baixa) que decidia SE
    #     preenchia com zero à esquerda. Agora o preenchimento é sempre
    #     automático e silencioso para 1-5 dígitos; só formato inesperado
    #     (letras, mais de 5 dígitos na base, célula Excel numérica
    #     genuinamente não-inteira) vira pendência. ---
    numerico_por_codigo: dict[str, bool] = {}
    for linha in leitura["linhas"]:
        if linha["zero_esquerda_desconhecido"]:
            numerico_por_codigo[linha["codigo_original"]] = True
        else:
            numerico_por_codigo.setdefault(linha["codigo_original"], False)

    todos_codigos_brutos = set(consolidacao["consolidados"])  # união das 3 fontes
    normalizacao_por_codigo: dict[str, dict] = {
        codigo: _normalizar_codigo_de_fonte(codigo, numerico=numerico_por_codigo.get(codigo, False))
        for codigo in todos_codigos_brutos
    }
    # Três classificações possíveis (nunca duas ao mesmo tempo): CPD, ARRUELA
    # (item completamente separado, nunca entra em cpds/cpd_variacoes) ou
    # inválido/ambíguo (nunca descartado, sempre vira pendência de formato).
    cpds_validos = {c: n for c, n in normalizacao_por_codigo.items() if n["tipo_item"] == "CPD"}
    arruelas_validas = {c: n for c, n in normalizacao_por_codigo.items() if n["tipo_item"] == "ARRUELA"}
    invalidos = {c: n for c, n in normalizacao_por_codigo.items() if not n["valido"]}

    def remapear(codigo: str) -> str | None:
        n = cpds_validos.get(codigo)
        return n["codigo_completo"] if n else None

    codigos_com_barra_convertidos = sorted(c for c, n in cpds_validos.items() if n["usou_barra"])
    codigos_preenchidos_com_zero = sum(1 for n in cpds_validos.values() if n["regra_aplicada"] == "zero_padding")
    codigos_ja_5_digitos = sum(1 for n in cpds_validos.values() if n["regra_aplicada"] == "ja_5_digitos")

    finais_por_bruto: dict[str, list[str]] = defaultdict(list)
    for bruto, n in cpds_validos.items():
        finais_por_bruto[n["codigo_completo"]].append(bruto)
    colisoes_normalizacao = {final: brutos for final, brutos in finais_por_bruto.items() if len(brutos) > 1}
    colisoes_barra_ponto = {
        final: brutos for final, brutos in colisoes_normalizacao.items()
        if any("/" in b for b in brutos) and any("/" not in b for b in brutos)
    }

    # --- Pendências de formato: nunca descarta um código inválido/ambíguo --
    #     preserva o valor bruto original + todas as ocorrências (fonte,
    #     linha/pedido, descrição, cliente) em detalhes_json. ---
    pendencias_formato = []
    for codigo_bruto, n in invalidos.items():
        ocorrencias = []
        for linha in leitura["linhas"]:
            if linha["codigo_original"] == codigo_bruto:
                ocorrencias.append({
                    "fonte": "LISTA_OFICIAL", "linha_planilha": linha["linha_planilha"],
                    "descricao": linha["descricao"] or None, "cliente": linha["cliente"] or None,
                })
        if any(r["codigo"] == codigo_bruto for r in manual_cpd_rows):
            ocorrencias.append({"fonte": "MANUAL_CPD", "descricao": manual_cpd_desc.get(codigo_bruto)})
        for o in orders:
            if str(o.get("cpd") or "").strip() == codigo_bruto:
                ocorrencias.append({
                    "fonte": "PEDIDO", "pedido_id": o.get("_id"),
                    "descricao": str(o.get("descricao") or "").strip() or None,
                    "cliente": str(o.get("cliente") or "").strip() or None,
                })

        tipo_pendencia = "FORMATO_AMBIGUO" if n["motivo"] == "FORMATO_AMBIGUO" else "CODIGO_INVALIDO"
        pendencias_formato.append({
            "codigo_completo": codigo_bruto,
            "tipo": tipo_pendencia, "nivel_confianca": None,
            "detalhes_json": json.dumps(
                {"codigo_original": codigo_bruto, "motivo": n["motivo"], "regra": "cpd_pai_5_digitos", "ocorrencias": ocorrencias},
                ensure_ascii=False, default=str,
            ),
            "status": "PENDENTE", "resolvido_em": None,
        })

    # --- Clientes consolidados (já aplicando os 6 aliases confirmados e
    #     excluindo placeholders como "." -- esses viram pendência à parte) ---
    nomes_canonicos = set()
    for o in orders:
        nome = str(o.get("cliente") or "").strip()
        if cliente_valido(nome):
            nomes_canonicos.add(canonico_de(nome))
    for nome in clientes_oficial["clientes_planilha_com_contagem_cpds"]:
        if cliente_valido(nome):
            nomes_canonicos.add(canonico_de(nome))

    def grupo_de(nome: str) -> str:
        return "Honda" if nome in config.grupo_honda else "Diversos"

    clientes_plano = [{"nome": nome, "ativo": 1, "grupo": grupo_de(nome)} for nome in sorted(nomes_canonicos)]

    # Aliases só são gravados para o cliente canônico que realmente existe no
    # plano. A comparação "idêntico ao canônico" é EXATA (preserva caixa) --
    # um alias com caixa diferente do canônico (ex.: "Fornecedor X" vs
    # "FORNECEDOR X") precisa virar alias; só descartaríamos um alias
    # byte-a-byte igual ao canônico (ex.: se a configuração listasse o
    # próprio nome canônico como alias de si mesmo).
    cliente_aliases_plano = [
        {"cliente_canonico": canonico, "alias": alias}
        for canonico, aliases in config.aliases.items()
        if canonico in nomes_canonicos
        for alias in aliases
        if alias != canonico
    ]

    # --- CPDs pai + variações, já considerando o remapeamento de zero à esquerda ---
    codigos_finais: dict[str, dict] = {}
    for codigo_original, n in cpds_validos.items():
        codigo_final = n["codigo_completo"]
        if codigo_final in codigos_finais:
            continue
        codigos_finais[codigo_final] = {
            "codigo_completo": codigo_final,
            "codigo_pai": n["codigo_pai"],
            "extensao": n["extensao"],
            "ativo": 1,
        }

    pais_agrupados: dict[str, list[str]] = defaultdict(list)
    for codigo_final, info in codigos_finais.items():
        pais_agrupados[info["codigo_pai"]].append(codigo_final)

    pais_criados_automaticamente = []
    for pai in list(pais_agrupados):
        if pai not in codigos_finais:
            codigos_finais[pai] = {"codigo_completo": pai, "codigo_pai": pai, "extensao": None, "ativo": 1}
            pais_agrupados[pai].append(pai)
            pais_criados_automaticamente.append(pai)

    # --- Descrições-fonte + pendências de descrição ---
    duplicados_codigos = {d["codigo"] for d in duplicidades if d["codigo"] in cpds_validos}
    conflitos_por_codigo = {c["codigo"]: c for c in conflitos_tecnicos["classificados"]}

    pedidos_desc_por_codigo: dict[str, list[dict]] = defaultdict(list)
    for o in orders:
        cpd = str(o.get("cpd") or "").strip()
        desc = str(o.get("descricao") or "").strip()
        if cpd and desc:
            pedidos_desc_por_codigo[cpd].append({
                "descricao": desc, "cliente": str(o.get("cliente") or "").strip() or None,
                "id": o.get("_id"), "data": o.get("criadoEm") or None,
            })

    linhas_planilha_por_codigo: dict[str, list[dict]] = defaultdict(list)
    for linha in leitura["linhas"]:
        linhas_planilha_por_codigo[linha["codigo_original"]].append(linha)

    descricoes_fontes: list[dict] = []
    pendencias_descricao: list[dict] = []

    # Agrupa candidatas por CÓDIGO FINAL (não por codigo_original): quando a
    # normalização de 5 dígitos funde "442" -> "00442", as descrições de
    # ambos os códigos originais precisam ser decididas JUNTAS (senão cada
    # um vira "canônica" isoladamente e o índice único parcial do banco
    # rejeita a segunda ao aplicar de verdade).
    candidatas_por_final: dict[str, list[dict]] = defaultdict(list)
    conflito_por_final: dict[str, dict] = {}
    _GRAVIDADE = {"A": 0, "B": 1, "C": 2, "D": 3}

    for codigo_original in cpds_validos:
        if codigo_original in duplicados_codigos:
            continue  # tratado à parte, abaixo
        codigo_final = remapear(codigo_original)

        linhas_oficiais = linhas_planilha_por_codigo.get(codigo_original, [])
        if linhas_oficiais and linhas_oficiais[0]["descricao"]:
            linha = linhas_oficiais[0]
            candidatas_por_final[codigo_final].append({
                "fonte": "LISTA_OFICIAL", "descricao": linha["descricao"],
                "referencia_origem": f"linha {linha['linha_planilha']}",
                "cliente_origem": linha["cliente"] or None, "data_origem": None,
            })
        if manual_cpd_desc.get(codigo_original):
            candidatas_por_final[codigo_final].append({
                "fonte": "MANUAL_CPD", "descricao": manual_cpd_desc[codigo_original],
                "referencia_origem": "manual_cpd", "cliente_origem": None, "data_origem": None,
            })
        for p in pedidos_desc_por_codigo.get(codigo_original, []):
            candidatas_por_final[codigo_final].append({
                "fonte": "PEDIDO", "descricao": p["descricao"],
                "referencia_origem": str(p["id"]) if p["id"] else None,
                "cliente_origem": p["cliente"], "data_origem": p["data"],
            })

        conflito = conflitos_por_codigo.get(codigo_original)
        if conflito:
            atual = conflito_por_final.get(codigo_final)
            if atual is None or _GRAVIDADE[conflito["categoria"]] > _GRAVIDADE[atual["categoria"]]:
                conflito_por_final[codigo_final] = conflito

    for codigo_final, candidatas in candidatas_por_final.items():
        conflito = conflito_por_final.get(codigo_final)
        canonica_idx = None

        if conflito:
            for i, cand in enumerate(candidatas):
                if cand["fonte"] == "LISTA_OFICIAL":
                    canonica_idx = i
                    break
            if conflito["categoria"] in ("C", "D"):
                pendencias_descricao.append({
                    "codigo_completo": codigo_final, "tipo": "CONFLITO_DESCRICAO", "nivel_confianca": None,
                    "detalhes_json": json.dumps(conflito, ensure_ascii=False, default=str),
                    "status": "PENDENTE", "resolvido_em": None,
                })
        elif candidatas:
            descricoes_distintas = {c["descricao"] for c in candidatas if c["descricao"]}
            if len(descricoes_distintas) <= 1:
                canonica_idx = 0
            else:
                # Descrições diferentes só se encontraram aqui por causa da
                # fusão de 5 dígitos (nunca foram comparadas pela
                # classificação de conflitos, que roda por codigo_original).
                # Mesma prioridade: lista oficial primeiro.
                for i, cand in enumerate(candidatas):
                    if cand["fonte"] == "LISTA_OFICIAL":
                        canonica_idx = i
                        break
                if canonica_idx is None:
                    canonica_idx = 0

        for i, cand in enumerate(candidatas):
            descricoes_fontes.append({
                "codigo_completo": codigo_final, "fonte": cand["fonte"], "descricao": cand["descricao"],
                "referencia_origem": cand["referencia_origem"], "cliente_origem": cand["cliente_origem"],
                "data_origem": cand["data_origem"], "descricao_canonica": 1 if i == canonica_idx else 0,
            })

    # --- Duplicidades na planilha: preserva as duas linhas, nunca escolhe entre elas ---
    for d in duplicidades:
        codigo_original = d["codigo"]
        if codigo_original not in cpds_validos:
            continue  # já virou pendência CODIGO_INVALIDO/FORMATO_AMBIGUO acima
        codigo_final = remapear(codigo_original)
        linhas_completas = d.get("linhas_completas") or [{}] * len(d["linhas_de_origem"])
        for ln, linha_completa in zip(d["linhas_de_origem"], linhas_completas):
            desc = lista_oficial.buscar_valor_coluna(linha_completa, "descrição", "descricao")
            cliente_valor = lista_oficial.buscar_valor_coluna(linha_completa, "cliente")
            cliente_origem = str(cliente_valor).strip() or None if cliente_valor not in (None, "") else None
            descricoes_fontes.append({
                "codigo_completo": codigo_final, "fonte": "LISTA_OFICIAL", "descricao": desc,
                "referencia_origem": f"linha {ln}", "cliente_origem": cliente_origem,
                "data_origem": None, "descricao_canonica": 0,
            })

        desc_existente, fonte_existente = None, None
        if manual_cpd_desc.get(codigo_original):
            desc_existente, fonte_existente = manual_cpd_desc[codigo_original], "MANUAL_CPD"
        elif pedidos_desc_por_codigo.get(codigo_original):
            desc_existente = pedidos_desc_por_codigo[codigo_original][0]["descricao"]
            fonte_existente = "PEDIDO"
        if desc_existente:
            descricoes_fontes.append({
                "codigo_completo": codigo_final, "fonte": fonte_existente, "descricao": desc_existente,
                "referencia_origem": "cadastro atual", "cliente_origem": None, "data_origem": None,
                "descricao_canonica": 1,
            })

        pendencias_descricao.append({
            "codigo_completo": codigo_final, "tipo": "DUPLICIDADE_LISTA", "nivel_confianca": None,
            "detalhes_json": json.dumps(
                {"linhas": d["linhas_de_origem"], "descricoes_encontradas": d["descricoes_encontradas"]},
                ensure_ascii=False, default=str,
            ),
            "status": "PENDENTE", "resolvido_em": None,
        })

    # --- Pai ausente ---
    pendencias_pai_ausente = [
        {
            "codigo_completo": pai, "tipo": "PAI_AUSENTE", "nivel_confianca": None,
            "detalhes_json": json.dumps({"motivo": "Só existiam variações deste código; o código-pai foi criado automaticamente."}),
            "status": "PENDENTE", "resolvido_em": None,
        }
        for pai in sorted(set(pais_criados_automaticamente))
    ]

    # --- Cliente indefinido: CPD (de pedido OU da lista oficial), com
    #     código VÁLIDO, sem cliente válido informado (vazio, ".", "-",
    #     "N/A" ou equivalente). O CPD é preservado normalmente; só o
    #     vínculo de cliente não é criado. Códigos inválidos/ambíguos já
    #     geraram sua própria pendência de formato acima -- não duplica. ---
    pendencias_cliente_indefinido = []
    for o in orders:
        cliente = str(o.get("cliente") or "").strip()
        cpd = str(o.get("cpd") or "").strip()
        if cpd and cpd in cpds_validos and not cliente_valido(cliente):
            pendencias_cliente_indefinido.append({
                "codigo_completo": remapear(cpd), "tipo": "CLIENTE_INDEFINIDO", "nivel_confianca": None,
                "detalhes_json": json.dumps(
                    {"origem": "PEDIDO", "pedido_id": o.get("_id"), "cliente_bruto": cliente or None},
                    ensure_ascii=False, default=str,
                ),
                "status": "PENDENTE", "resolvido_em": None,
            })
    for linha in leitura["linhas"]:
        cliente = linha["cliente"]
        if linha["codigo_original"] in cpds_validos and not cliente_valido(cliente):
            pendencias_cliente_indefinido.append({
                "codigo_completo": remapear(linha["codigo_original"]), "tipo": "CLIENTE_INDEFINIDO", "nivel_confianca": None,
                "detalhes_json": json.dumps(
                    {
                        "origem": "LISTA_OFICIAL", "linha_planilha": linha["linha_planilha"],
                        "cpd": linha["codigo_original"], "descricao": linha["descricao"],
                        "cliente_bruto": cliente or None,
                    },
                    ensure_ascii=False, default=str,
                ),
                "status": "PENDENTE", "resolvido_em": None,
            })

    # --- Vínculos cliente x CPD-pai: união de lista oficial + pedidos.
    #     Só cria vínculo quando o código normaliza pra um CPD-pai válido --
    #     códigos inválidos/ambíguos nunca geram vínculo (já viraram
    #     pendência de formato acima). ---
    vinculos_lista_oficial = []
    for linha in leitura["linhas"]:
        cliente_bruto = linha["cliente"]
        if not cliente_valido(cliente_bruto):
            continue  # tratado como pendência CLIENTE_INDEFINIDO acima, nunca vira vínculo
        codigo_completo_origem = linha["codigo_original"]
        n = cpds_validos.get(codigo_completo_origem)
        if n is None:
            continue
        vinculos_lista_oficial.append({
            "cliente_canonico": canonico_de(cliente_bruto),
            "cliente_original": cliente_bruto,
            "cpd_completo_origem": codigo_completo_origem,
            "cpd_pai": n["codigo_pai"],
            "referencia_origem": f"linha {linha['linha_planilha']}",
        })

    vinculos_pedidos = []
    for o in orders:
        cliente_bruto = str(o.get("cliente") or "").strip()
        cpd = str(o.get("cpd") or "").strip()
        if not cpd or not cliente_valido(cliente_bruto):
            continue
        n = cpds_validos.get(cpd)
        if n is None:
            continue
        vinculos_pedidos.append({
            "cliente_canonico": canonico_de(cliente_bruto),
            "cliente_original": cliente_bruto,
            "cpd_completo_origem": cpd,
            "cpd_pai": n["codigo_pai"],
            "referencia_origem": str(o.get("_id")) if o.get("_id") else None,
        })

    origens_por_par: dict[tuple[str, str], set[str]] = defaultdict(set)
    exemplos_por_par: dict[tuple[str, str], dict[str, list[dict]]] = defaultdict(lambda: {"LISTA_OFICIAL": [], "PEDIDO": []})
    for v in vinculos_lista_oficial:
        chave = (v["cliente_canonico"], v["cpd_pai"])
        origens_por_par[chave].add("LISTA_OFICIAL")
        exemplos_por_par[chave]["LISTA_OFICIAL"].append(v)
    for v in vinculos_pedidos:
        chave = (v["cliente_canonico"], v["cpd_pai"])
        origens_por_par[chave].add("PEDIDO")
        exemplos_por_par[chave]["PEDIDO"].append(v)

    cliente_cpds_plano = []
    for (cliente, pai), origens in sorted(origens_por_par.items()):
        origem_final = "AMBAS" if len(origens) == 2 else next(iter(origens))
        cliente_cpds_plano.append({
            "cliente": cliente, "cpd_pai": pai, "origem": origem_final,
            "exemplos_lista_oficial": exemplos_por_par[(cliente, pai)]["LISTA_OFICIAL"],
            "exemplos_pedido": exemplos_por_par[(cliente, pai)]["PEDIDO"],
        })

    grupo_cpd_pairs = {(grupo_de(item["cliente"]), item["cpd_pai"]) for item in cliente_cpds_plano}

    todas_pendencias = pendencias_formato + pendencias_descricao + pendencias_pai_ausente + pendencias_cliente_indefinido

    # =========================================================================
    # ARRUELAS: cadastro completamente separado. Nunca entram em codigos_finais/
    # pais_agrupados/cliente_cpds/grupo_cpds -- cada código de arruela É o
    # próprio item (sem pai/variação), preservado exatamente como veio
    # (codigo_canonico == codigo_original, só trim). Reaproveita os MESMOS
    # clientes canônicos e aliases já definidos acima pra CPDs.
    # =========================================================================
    arruelas_finais: dict[str, dict] = {
        n["codigo_canonico"]: {"codigo": n["codigo_canonico"], "ativo": 1}
        for n in arruelas_validas.values()
    }

    arruela_descricoes_fontes: list[dict] = []
    pendencias_arruela_descricao: list[dict] = []
    total_duplicidades_arruela = 0
    total_conflitos_descricao_arruela = 0

    candidatas_arruela_por_codigo: dict[str, list[dict]] = defaultdict(list)
    for codigo_original in arruelas_validas:
        for linha in linhas_planilha_por_codigo.get(codigo_original, []):
            if linha["descricao"]:
                candidatas_arruela_por_codigo[codigo_original].append({
                    "fonte": "LISTA_OFICIAL", "descricao": linha["descricao"],
                    "referencia_origem": f"linha {linha['linha_planilha']}",
                    "cliente_origem": linha["cliente"] or None,
                })
        if manual_cpd_desc.get(codigo_original):
            candidatas_arruela_por_codigo[codigo_original].append({
                "fonte": "MANUAL_CPD", "descricao": manual_cpd_desc[codigo_original],
                "referencia_origem": "manual_cpd", "cliente_origem": None,
            })
        for p in pedidos_desc_por_codigo.get(codigo_original, []):
            candidatas_arruela_por_codigo[codigo_original].append({
                "fonte": "PEDIDO", "descricao": p["descricao"],
                "referencia_origem": str(p["id"]) if p["id"] else None, "cliente_origem": p["cliente"],
            })

    for codigo_arruela, candidatas in candidatas_arruela_por_codigo.items():
        if len(linhas_planilha_por_codigo.get(codigo_arruela, [])) > 1:
            total_duplicidades_arruela += 1

        descricoes_distintas = {c["descricao"] for c in candidatas if c["descricao"]}
        canonica_idx = None
        if len(descricoes_distintas) <= 1:
            canonica_idx = 0 if candidatas else None
        else:
            total_conflitos_descricao_arruela += 1
            for i, cand in enumerate(candidatas):
                if cand["fonte"] == "LISTA_OFICIAL":
                    canonica_idx = i
                    break
            if canonica_idx is None:
                canonica_idx = 0

        for i, cand in enumerate(candidatas):
            arruela_descricoes_fontes.append({
                "codigo_original": codigo_arruela, "fonte": cand["fonte"], "descricao": cand["descricao"],
                "referencia_origem": cand["referencia_origem"], "cliente_origem": cand["cliente_origem"],
                "descricao_canonica": 1 if i == canonica_idx else 0,
            })

    # --- Vínculos cliente x arruela: mesma união lista oficial + pedidos,
    #     mesmos clientes canônicos/aliases, mesma regra de grupo. ---
    vinculos_arruela_lista_oficial = []
    for linha in leitura["linhas"]:
        cliente_bruto = linha["cliente"]
        if not cliente_valido(cliente_bruto):
            continue
        codigo_original = linha["codigo_original"]
        if codigo_original not in arruelas_validas:
            continue
        vinculos_arruela_lista_oficial.append({
            "cliente_canonico": canonico_de(cliente_bruto),
            "cliente_original": cliente_bruto,
            "arruela": codigo_original,
            "referencia_origem": f"linha {linha['linha_planilha']}",
        })

    vinculos_arruela_pedidos = []
    for o in orders:
        cliente_bruto = str(o.get("cliente") or "").strip()
        cpd = str(o.get("cpd") or "").strip()
        if not cpd or not cliente_valido(cliente_bruto) or cpd not in arruelas_validas:
            continue
        vinculos_arruela_pedidos.append({
            "cliente_canonico": canonico_de(cliente_bruto),
            "cliente_original": cliente_bruto,
            "arruela": cpd,
            "referencia_origem": str(o.get("_id")) if o.get("_id") else None,
        })

    origens_arruela_por_par: dict[tuple[str, str], set[str]] = defaultdict(set)
    for v in vinculos_arruela_lista_oficial:
        origens_arruela_por_par[(v["cliente_canonico"], v["arruela"])].add("LISTA_OFICIAL")
    for v in vinculos_arruela_pedidos:
        origens_arruela_por_par[(v["cliente_canonico"], v["arruela"])].add("PEDIDO")

    cliente_arruelas_plano = []
    for (cliente, arruela), origens in sorted(origens_arruela_por_par.items()):
        origem_final = "AMBAS" if len(origens) == 2 else next(iter(origens))
        cliente_arruelas_plano.append({"cliente": cliente, "arruela": arruela, "origem": origem_final})

    grupo_arruela_pairs = {(grupo_de(item["cliente"]), item["arruela"]) for item in cliente_arruelas_plano}

    return {
        "gerado_em": agora,
        "clientes": clientes_plano,
        "cliente_aliases": cliente_aliases_plano,
        "codigos_finais": codigos_finais,
        "pais_agrupados": dict(pais_agrupados),
        "cliente_cpds": cliente_cpds_plano,
        "grupo_cpds": sorted(grupo_cpd_pairs),
        "descricoes_fontes": descricoes_fontes,
        "pendencias": todas_pendencias,
        "normalizacao_por_codigo": normalizacao_por_codigo,
        "codigos_invalidos": invalidos,
        "colisoes_normalizacao": colisoes_normalizacao,
        "colisoes_barra_ponto": colisoes_barra_ponto,
        "codigos_com_barra_convertidos": codigos_com_barra_convertidos,
        "codigos_preenchidos_com_zero": codigos_preenchidos_com_zero,
        "codigos_ja_5_digitos": codigos_ja_5_digitos,
        "arruelas": arruelas_finais,
        "arruela_descricoes_fontes": arruela_descricoes_fontes,
        "cliente_arruelas": cliente_arruelas_plano,
        "grupo_arruelas": sorted(grupo_arruela_pairs),
        "total_duplicidades_arruela": total_duplicidades_arruela,
        "total_conflitos_descricao_arruela": total_conflitos_descricao_arruela,
        "resultado_analise_oficial": resultado_oficial,
    }


# ---------------------------------------------------------------------------
# Simulação (só relatório, nada gravado)
# ---------------------------------------------------------------------------

def simular(plano: dict) -> dict:
    codigos_finais = plano["codigos_finais"]
    variacoes = [c for c, info in codigos_finais.items() if info["extensao"] is not None]
    pendencias_por_tipo = dict(Counter(p["tipo"] for p in plano["pendencias"]))
    pendencias_por_status = dict(Counter(p["status"] for p in plano["pendencias"]))

    todos_os_pais = set(plano["pais_agrupados"])
    origem_por_pai: dict[str, set[str]] = defaultdict(set)
    grupos_por_pai: dict[str, set[str]] = defaultdict(set)
    for item in plano["cliente_cpds"]:
        pai = item["cpd_pai"]
        if item["origem"] == "AMBAS":
            origem_por_pai[pai] |= {"LISTA_OFICIAL", "PEDIDO"}
        else:
            origem_por_pai[pai].add(item["origem"])
    for gr, pai in plano["grupo_cpds"]:
        grupos_por_pai[pai].add(gr)

    pais_sem_cliente = todos_os_pais - set(origem_por_pai)
    pais_so_lista_oficial = {p for p, o in origem_por_pai.items() if o == {"LISTA_OFICIAL"}}
    pais_so_pedidos = {p for p, o in origem_por_pai.items() if o == {"PEDIDO"}}
    pais_ambas_fontes = {p for p, o in origem_por_pai.items() if o == {"LISTA_OFICIAL", "PEDIDO"}}

    pais_so_honda = {p for p, g in grupos_por_pai.items() if g == {"Honda"}}
    pais_so_diversos = {p for p, g in grupos_por_pai.items() if g == {"Diversos"}}
    pais_ambos_grupos = {p for p, g in grupos_por_pai.items() if g == {"Honda", "Diversos"}}
    pais_sem_grupo = todos_os_pais - set(grupos_por_pai)

    leitura = plano["resultado_analise_oficial"]["leitura"]
    normalizacao_por_codigo = plano["normalizacao_por_codigo"]
    codigos_invalidos = plano["codigos_invalidos"]
    total_ambiguos = sum(1 for n in codigos_invalidos.values() if n["motivo"] == "FORMATO_AMBIGUO")
    total_invalidos = len(codigos_invalidos) - total_ambiguos

    arruelas = plano["arruelas"]
    clientes_arruela = {item["cliente"] for item in plano["cliente_arruelas"]}

    return {
        "total_clientes": len(plano["clientes"]),
        "total_aliases": len(plano["cliente_aliases"]),
        "total_cpds_pai": len(plano["pais_agrupados"]),
        "total_codigos_completos": len(codigos_finais),
        "total_variacoes": len(variacoes),
        "total_vinculos_cliente_cpd": len(plano["cliente_cpds"]),
        "total_vinculos_grupo_cpd": len(plano["grupo_cpds"]),
        "total_descricoes_fontes": len(plano["descricoes_fontes"]),
        "total_pendencias": len(plano["pendencias"]),
        "pendencias_por_tipo": pendencias_por_tipo,
        "pendencias_por_status": pendencias_por_status,
        "total_linhas_planilha": len(leitura["linhas"]),
        "total_codigos_originais_unicos": len(normalizacao_por_codigo),
        "total_codigos_canonicos_unicos": len(codigos_finais),
        "total_codigos_preenchidos_com_zero": plano["codigos_preenchidos_com_zero"],
        "total_codigos_ja_5_digitos": plano["codigos_ja_5_digitos"],
        "total_colisoes_normalizacao": len(plano["colisoes_normalizacao"]),
        "total_codigos_com_barra_convertidos": len(plano["codigos_com_barra_convertidos"]),
        "total_colisoes_barra_ponto": len(plano["colisoes_barra_ponto"]),
        "total_codigos_invalidos": total_invalidos,
        "total_codigos_ambiguos": total_ambiguos,
        "cpds_sem_cliente": len(pais_sem_cliente),
        "cpds_somente_lista_oficial": len(pais_so_lista_oficial),
        "cpds_somente_pedidos": len(pais_so_pedidos),
        "cpds_ambas_fontes": len(pais_ambas_fontes),
        "cpds_so_grupo_honda": len(pais_so_honda),
        "cpds_so_grupo_diversos": len(pais_so_diversos),
        "cpds_ambos_grupos": len(pais_ambos_grupos),
        "cpds_sem_grupo": len(pais_sem_grupo),
        "todos_clientes_ativos": all(c["ativo"] == 1 for c in plano["clientes"]),
        "todos_cpds_ativos": all(info["ativo"] == 1 for info in codigos_finais.values()),
        "ponto_final_nao_virou_cliente": not any(c["nome"] == "." for c in plano["clientes"]),
        "nenhuma_arruela_em_cpds": not (set(arruelas) & set(codigos_finais)),
        "total_arruelas_linhas": sum(1 for linha in leitura["linhas"] if linha["codigo_original"] in arruelas),
        "total_arruelas_codigos_unicos": len(arruelas),
        "total_arruelas_descricoes": len(plano["arruela_descricoes_fontes"]),
        "total_arruelas_clientes_relacionados": len(clientes_arruela),
        "total_vinculos_cliente_arruela": len(plano["cliente_arruelas"]),
        "total_vinculos_grupo_arruela": len(plano["grupo_arruelas"]),
        "total_duplicidades_arruela": plano["total_duplicidades_arruela"],
        "total_conflitos_descricao_arruela": plano["total_conflitos_descricao_arruela"],
        "todas_arruelas_ativas": all(a["ativo"] == 1 for a in arruelas.values()),
    }


# ---------------------------------------------------------------------------
# Aplicação real (só chamada em --confirmar)
# ---------------------------------------------------------------------------

def aplicar_migracao(conn: sqlite3.Connection, plano: dict) -> dict:
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    grupo_id_por_nome = {row[1]: row[0] for row in conn.execute("SELECT id, nome FROM grupos_clientes")}

    cliente_id_por_nome: dict[str, int] = {}
    for cliente in plano["clientes"]:
        conn.execute(
            "INSERT INTO clientes(nome, ativo, criado_em, atualizado_em) VALUES (?,1,?,?)",
            (cliente["nome"], agora, agora),
        )
        cliente_id = conn.execute("SELECT id FROM clientes WHERE nome=?", (cliente["nome"],)).fetchone()[0]
        cliente_id_por_nome[cliente["nome"]] = cliente_id
        grupo_id = grupo_id_por_nome.get(cliente["grupo"])
        if grupo_id is not None:
            conn.execute(
                "INSERT OR IGNORE INTO cliente_grupos(cliente_id, grupo_id, criado_em) VALUES (?,?,?)",
                (cliente_id, grupo_id, agora),
            )

    total_aliases = 0
    for item in plano["cliente_aliases"]:
        cliente_id = cliente_id_por_nome.get(item["cliente_canonico"])
        if cliente_id is not None:
            conn.execute(
                "INSERT INTO cliente_aliases(cliente_id, alias, ativo, criado_em) VALUES (?,?,1,?)",
                (cliente_id, item["alias"], agora),
            )
            total_aliases += 1

    canonicas_por_codigo = {
        d["codigo_completo"]: d["descricao"] for d in plano["descricoes_fontes"] if d["descricao_canonica"] == 1
    }

    cpd_id_por_pai: dict[str, int] = {}
    for pai in sorted(plano["pais_agrupados"]):
        desc_padrao = canonicas_por_codigo.get(pai)
        conn.execute(
            "INSERT INTO cpds(codigo_pai, descricao_padrao, ativo, criado_em, atualizado_em) VALUES (?,?,1,?,?)",
            (pai, desc_padrao, agora, agora),
        )
        cpd_id_por_pai[pai] = conn.execute("SELECT id FROM cpds WHERE codigo_pai=?", (pai,)).fetchone()[0]

    cpd_variacao_id_por_codigo: dict[str, int] = {}
    for codigo_completo, info in sorted(plano["codigos_finais"].items()):
        if info["extensao"] is None:
            continue
        cpd_id = cpd_id_por_pai[info["codigo_pai"]]
        desc_especifica = canonicas_por_codigo.get(codigo_completo)
        conn.execute(
            "INSERT INTO cpd_variacoes(cpd_id, codigo_completo, extensao, descricao_especifica, ativo, criado_em, atualizado_em) "
            "VALUES (?,?,?,?,1,?,?)",
            (cpd_id, codigo_completo, info["extensao"], desc_especifica, agora, agora),
        )
        cpd_variacao_id_por_codigo[codigo_completo] = conn.execute(
            "SELECT id FROM cpd_variacoes WHERE codigo_completo=?", (codigo_completo,)
        ).fetchone()[0]

    total_cliente_cpds = 0
    for item in plano["cliente_cpds"]:
        cliente_id = cliente_id_por_nome.get(item["cliente"])
        cpd_id = cpd_id_por_pai.get(item["cpd_pai"])
        if cliente_id is not None and cpd_id is not None:
            conn.execute(
                "INSERT OR IGNORE INTO cliente_cpds(cliente_id, cpd_id, criado_em) VALUES (?,?,?)",
                (cliente_id, cpd_id, agora),
            )
            total_cliente_cpds += 1

    total_grupo_cpds = 0
    for grupo_nome, pai in plano["grupo_cpds"]:
        grupo_id = grupo_id_por_nome.get(grupo_nome)
        cpd_id = cpd_id_por_pai.get(pai)
        if grupo_id is not None and cpd_id is not None:
            conn.execute(
                "INSERT OR IGNORE INTO grupo_cpds(grupo_id, cpd_id, criado_em) VALUES (?,?,?)",
                (grupo_id, cpd_id, agora),
            )
            total_grupo_cpds += 1

    for d in plano["descricoes_fontes"]:
        info = plano["codigos_finais"][d["codigo_completo"]]
        cpd_id = cpd_id_por_pai[info["codigo_pai"]]
        variacao_id = cpd_variacao_id_por_codigo.get(d["codigo_completo"]) if info["extensao"] is not None else None
        conn.execute(
            "INSERT INTO cpd_descricoes_fontes"
            "(cpd_id, cpd_variacao_id, codigo_completo, descricao, fonte, referencia_origem, cliente_origem, data_origem, descricao_canonica, criado_em) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            (cpd_id, variacao_id, d["codigo_completo"], d["descricao"], d["fonte"], d["referencia_origem"],
             d["cliente_origem"], d["data_origem"], d["descricao_canonica"], agora),
        )

    for p in plano["pendencias"]:
        info = plano["codigos_finais"].get(p["codigo_completo"])
        cpd_id = cpd_id_por_pai.get(info["codigo_pai"]) if info else None
        variacao_id = cpd_variacao_id_por_codigo.get(p["codigo_completo"]) if info and info["extensao"] is not None else None
        conn.execute(
            "INSERT INTO cpd_pendencias_revisao"
            "(cpd_id, cpd_variacao_id, codigo_completo, tipo, nivel_confianca, detalhes_json, status, criado_em, resolvido_em) "
            "VALUES (?,?,?,?,?,?,?,?,?)",
            (cpd_id, variacao_id, p["codigo_completo"], p["tipo"], p.get("nivel_confianca"), p["detalhes_json"],
             p["status"], agora, p.get("resolvido_em")),
        )

    # --- Arruelas: só grava se o schema de destino já tem as tabelas de
    #     arruela (migração v4->v5, aplicada separadamente e ainda NÃO no
    #     banco real). Se não existirem, não falha -- só não grava nada de
    #     arruela nesta rodada (uso esperado ao aplicar --confirmar num banco
    #     ainda em schema v4, sem quebrar o fluxo de CPDs já existente). ---
    tem_tabelas_arruela = conn.execute(
        "SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='arruelas'"
    ).fetchone()[0] > 0

    total_arruelas_inseridas = 0
    total_cliente_arruelas = 0
    total_grupo_arruelas = 0
    total_arruela_descricoes_fontes = 0

    if tem_tabelas_arruela:
        arruela_id_por_codigo: dict[str, int] = {}
        canonicas_arruela_por_codigo = {
            d["codigo_original"]: d["descricao"] for d in plano["arruela_descricoes_fontes"] if d["descricao_canonica"] == 1
        }
        for codigo, info in sorted(plano["arruelas"].items()):
            desc_padrao = canonicas_arruela_por_codigo.get(codigo)
            conn.execute(
                "INSERT INTO arruelas(codigo, descricao_padrao, ativo, criado_em, atualizado_em) VALUES (?,?,1,?,?)",
                (codigo, desc_padrao, agora, agora),
            )
            arruela_id_por_codigo[codigo] = conn.execute("SELECT id FROM arruelas WHERE codigo=?", (codigo,)).fetchone()[0]
        total_arruelas_inseridas = len(arruela_id_por_codigo)

        for item in plano["cliente_arruelas"]:
            cliente_id = cliente_id_por_nome.get(item["cliente"])
            arruela_id = arruela_id_por_codigo.get(item["arruela"])
            if cliente_id is not None and arruela_id is not None:
                conn.execute(
                    "INSERT OR IGNORE INTO cliente_arruelas(cliente_id, arruela_id, origem, criado_em) VALUES (?,?,?,?)",
                    (cliente_id, arruela_id, item["origem"], agora),
                )
                total_cliente_arruelas += 1

        for grupo_nome, codigo in plano["grupo_arruelas"]:
            grupo_id = grupo_id_por_nome.get(grupo_nome)
            arruela_id = arruela_id_por_codigo.get(codigo)
            if grupo_id is not None and arruela_id is not None:
                conn.execute(
                    "INSERT OR IGNORE INTO grupo_arruelas(grupo_id, arruela_id, criado_em) VALUES (?,?,?)",
                    (grupo_id, arruela_id, agora),
                )
                total_grupo_arruelas += 1

        for d in plano["arruela_descricoes_fontes"]:
            arruela_id = arruela_id_por_codigo[d["codigo_original"]]
            conn.execute(
                "INSERT INTO arruela_descricoes_fontes"
                "(arruela_id, codigo_original, descricao, fonte, referencia_origem, cliente_origem, descricao_canonica, criado_em) "
                "VALUES (?,?,?,?,?,?,?,?)",
                (arruela_id, d["codigo_original"], d["descricao"], d["fonte"], d["referencia_origem"],
                 d["cliente_origem"], d["descricao_canonica"], agora),
            )
        total_arruela_descricoes_fontes = len(plano["arruela_descricoes_fontes"])

    violacoes = conn.execute("PRAGMA foreign_key_check").fetchall()
    if violacoes:
        raise RuntimeError(f"foreign_key_check encontrou violações após a migração: {violacoes}")

    return {
        "total_clientes_inseridos": len(cliente_id_por_nome),
        "total_aliases_inseridos": total_aliases,
        "total_cpds_pai_inseridos": len(cpd_id_por_pai),
        "total_variacoes_inseridas": len(cpd_variacao_id_por_codigo),
        "total_cliente_cpds_inseridos": total_cliente_cpds,
        "total_grupo_cpds_inseridos": total_grupo_cpds,
        "total_descricoes_fontes_inseridas": len(plano["descricoes_fontes"]),
        "tabelas_arruela_disponiveis": tem_tabelas_arruela,
        "total_arruelas_inseridas": total_arruelas_inseridas,
        "total_cliente_arruelas_inseridos": total_cliente_arruelas,
        "total_grupo_arruelas_inseridos": total_grupo_arruelas,
        "total_arruela_descricoes_fontes_inseridas": total_arruela_descricoes_fontes,
        "total_pendencias_inseridas": len(plano["pendencias"]),
    }


# ---------------------------------------------------------------------------
# Relatório de simulação
# ---------------------------------------------------------------------------

def gravar_relatorio_simulacao(resumo: dict, metadados: dict, pasta_saida: Path) -> Path:
    pasta_saida.mkdir(parents=True, exist_ok=True)
    caminho = pasta_saida / "simulacao_migracao_clientes_cpds.json"
    with caminho.open("w", encoding="utf-8") as f:
        json.dump({"metadados": metadados, "resumo": resumo}, f, ensure_ascii=False, indent=2, default=str)
    return caminho


def gravar_relatorio_vinculos(plano: dict, resumo: dict, metadados: dict, pasta_saida: Path) -> Path:
    from openpyxl import Workbook

    pasta_saida.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws_clientes = wb.active
    ws_clientes.title = "Clientes canonicos"
    ws_clientes.append([f"Gerado em: {metadados['gerado_em']}"])
    ws_clientes.append([])
    ws_clientes.append(["Cliente canônico", "Grupo", "Ativo"])
    for c in plano["clientes"]:
        ws_clientes.append([c["nome"], c["grupo"], c["ativo"]])

    ws_aliases = wb.create_sheet("Aliases")
    ws_aliases.append(["Cliente canônico", "Alias (grafia original preservada)"])
    for a in plano["cliente_aliases"]:
        ws_aliases.append([a["cliente_canonico"], a["alias"]])

    ws_lo = wb.create_sheet("Vinculos lista oficial")
    ws_lo.append(["Cliente canônico", "Nome original", "CPD completo (origem)", "CPD pai", "Referência (linha)"])
    for item in plano["cliente_cpds"]:
        for v in item["exemplos_lista_oficial"]:
            ws_lo.append([v["cliente_canonico"], v["cliente_original"], v["cpd_completo_origem"], v["cpd_pai"], v["referencia_origem"]])

    ws_ped = wb.create_sheet("Vinculos pedidos")
    ws_ped.append(["Cliente canônico", "Nome original", "CPD completo (origem)", "CPD pai", "Referência (pedido)"])
    for item in plano["cliente_cpds"]:
        for v in item["exemplos_pedido"]:
            ws_ped.append([v["cliente_canonico"], v["cliente_original"], v["cpd_completo_origem"], v["cpd_pai"], v["referencia_origem"]])

    grupo_por_cliente = {c["nome"]: c["grupo"] for c in plano["clientes"]}
    ws_cons = wb.create_sheet("Vinculos consolidados")
    ws_cons.append(["Cliente canônico", "CPD pai", "Origem", "Grupo"])
    for item in plano["cliente_cpds"]:
        ws_cons.append([item["cliente"], item["cpd_pai"], item["origem"], grupo_por_cliente.get(item["cliente"])])

    pais_com_vinculo = {item["cpd_pai"] for item in plano["cliente_cpds"]}
    ws_sem_cliente = wb.create_sheet("CPDs sem cliente")
    ws_sem_cliente.append(["CPD pai"])
    for pai in sorted(set(plano["pais_agrupados"]) - pais_com_vinculo):
        ws_sem_cliente.append([pai])

    ws_indef = wb.create_sheet("Clientes indefinidos")
    ws_indef.append(["Código completo", "Detalhes (JSON)"])
    for p in plano["pendencias"]:
        if p["tipo"] == "CLIENTE_INDEFINIDO":
            ws_indef.append([p["codigo_completo"], p["detalhes_json"]])

    ws_resumo = wb.create_sheet("Resumo")
    ws_resumo.append(["Item", "Valor"])
    for k, v in resumo.items():
        if isinstance(v, (int, float, bool, str)) or v is None:
            ws_resumo.append([k, v])

    caminho = pasta_saida / "vinculos_clientes_cpds_final.xlsx"
    wb.save(caminho)
    return caminho


def gravar_relatorio_normalizacao(plano: dict, metadados: dict, pasta_saida: Path) -> Path:
    """normalizacao_cpds_5_digitos.xlsx: toda a decisão de normalização de 5
    dígitos, código a código, mais as colisões (dois ou mais códigos brutos
    que se fundiram no mesmo código final)."""
    from openpyxl import Workbook

    pasta_saida.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "Normalizacao"
    ws.append([f"Gerado em: {metadados['gerado_em']}"])
    ws.append([])
    ws.append(["Código original", "Código pai", "Código completo", "Extensão", "Válido", "Motivo", "Regra aplicada"])
    for codigo, n in sorted(plano["normalizacao_por_codigo"].items()):
        ws.append([codigo, n["codigo_pai"], n["codigo_completo"], n["extensao"], n["valido"], n["motivo"], n["regra_aplicada"]])

    ws_col = wb.create_sheet("Colisoes")
    ws_col.append(["Código final (canônico)", "Códigos originais fundidos", "Quantidade"])
    for final, brutos in sorted(plano["colisoes_normalizacao"].items()):
        ws_col.append([final, ", ".join(sorted(brutos)), len(brutos)])

    ws_resumo = wb.create_sheet("Resumo")
    ws_resumo.append(["Item", "Valor"])
    ws_resumo.append(["Total de códigos originais únicos", len(plano["normalizacao_por_codigo"])])
    ws_resumo.append(["Códigos preenchidos com zero à esquerda", plano["codigos_preenchidos_com_zero"]])
    ws_resumo.append(["Códigos que já tinham 5 dígitos", plano["codigos_ja_5_digitos"]])
    ws_resumo.append(["Colisões após normalização (códigos finais com >1 origem)", len(plano["colisoes_normalizacao"])])
    ws_resumo.append(["Códigos inválidos ou ambíguos", len(plano["codigos_invalidos"])])

    caminho = pasta_saida / "normalizacao_cpds_5_digitos.xlsx"
    wb.save(caminho)
    return caminho


def gravar_relatorio_simulacao_5_digitos(resumo: dict, metadados: dict, pasta_saida: Path) -> Path:
    """simulacao_final_cpds_5_digitos.xlsx: mesmas contagens de simular(),
    em planilha (além do simulacao_migracao_clientes_cpds.json já existente),
    pra facilitar a conferência humana lado a lado com a simulação anterior."""
    from openpyxl import Workbook

    pasta_saida.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.append([f"Gerado em: {metadados['gerado_em']}"])
    ws.append([f"Planilha: {metadados['planilha_origem']}"])
    ws.append([f"Modo: {metadados['modo']}"])
    ws.append([f"Config. clientes (SHA-256, nunca o conteúdo): {metadados['config_clientes_hash_sha256']}"])
    ws.append([])
    ws.append(["Item", "Valor"])
    for k, v in resumo.items():
        if isinstance(v, dict):
            for sub_k, sub_v in v.items():
                ws.append([f"{k}.{sub_k}", sub_v])
        else:
            ws.append([k, v])

    caminho = pasta_saida / "simulacao_final_cpds_5_digitos.xlsx"
    wb.save(caminho)
    return caminho


def gravar_relatorio_codigos_invalidos(plano: dict, metadados: dict, pasta_saida: Path) -> Path:
    """codigos_invalidos_ou_ambiguos.xlsx: todo código que não pôde virar
    CPD automaticamente -- nunca descartado, sempre com o valor original e
    todas as ocorrências (fonte, linha/pedido, descrição, cliente)."""
    from openpyxl import Workbook

    pasta_saida.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Codigos invalidos"
    ws.append([f"Gerado em: {metadados['gerado_em']}"])
    ws.append([])
    ws.append(["Código original", "Motivo", "Tipo de pendência", "Ocorrências (JSON)"])

    pendencias_por_codigo = {
        p["codigo_completo"]: p for p in plano["pendencias"] if p["tipo"] in ("CODIGO_INVALIDO", "FORMATO_AMBIGUO")
    }
    for codigo, n in sorted(plano["codigos_invalidos"].items()):
        pendencia = pendencias_por_codigo.get(codigo)
        ws.append([codigo, n["motivo"], pendencia["tipo"] if pendencia else None, pendencia["detalhes_json"] if pendencia else None])

    caminho = pasta_saida / "codigos_invalidos_ou_ambiguos.xlsx"
    wb.save(caminho)
    return caminho


def _padrao_arruela(codigo: str) -> str:
    """Máscara legível do formato ('L'=letra, 'D'=dígito, resto preservado),
    ex.: 'XY-090' -> 'LL-DDD'. Só pra exibição no relatório -- nunca usada
    pra decidir classificação (isso é sempre _CODIGO_ARRUELA_RE)."""
    return "".join("L" if ch.isalpha() else "D" if ch.isdigit() else ch for ch in codigo)


def gravar_relatorio_barra(plano: dict, metadados: dict, pasta_saida: Path) -> Path:
    """cpds_com_barra_normalizados.xlsx: todo código que usava '/' como
    separador de extensão, convertido pra canônico com '.', preservando o
    original."""
    from openpyxl import Workbook

    pasta_saida.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Barra para ponto"
    ws.append([f"Gerado em: {metadados['gerado_em']}"])
    ws.append([])
    ws.append(["Código original (barra)", "Código pai", "Extensão", "Código canônico (ponto)", "Colidiu com outro código?"])

    normalizacao = plano["normalizacao_por_codigo"]
    for codigo in plano["codigos_com_barra_convertidos"]:
        n = normalizacao[codigo]
        colidiu = n["codigo_completo"] in plano["colisoes_normalizacao"] and len(plano["colisoes_normalizacao"][n["codigo_completo"]]) > 1
        ws.append([codigo, n["codigo_pai"], n["extensao"], n["codigo_completo"], colidiu])

    ws_resumo = wb.create_sheet("Resumo")
    ws_resumo.append(["Item", "Valor"])
    ws_resumo.append(["Total de códigos com barra convertidos", len(plano["codigos_com_barra_convertidos"])])
    ws_resumo.append(["Colisões causadas por barra/ponto (mesmo código, duas grafias)", len(plano["colisoes_barra_ponto"])])

    caminho = pasta_saida / "cpds_com_barra_normalizados.xlsx"
    wb.save(caminho)
    return caminho


def gravar_relatorio_arruelas(plano: dict, metadados: dict, pasta_saida: Path) -> Path:
    """arruelas_identificadas.xlsx: código original; descrição; cliente;
    padrão identificado; linha da planilha; status inicial ativo -- uma
    linha por ocorrência na lista oficial (preserva duplicidades, se houver)."""
    from openpyxl import Workbook

    pasta_saida.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Arruelas"
    ws.append([f"Gerado em: {metadados['gerado_em']}"])
    ws.append([])
    ws.append(["Código original", "Descrição", "Cliente", "Padrão identificado", "Linha da planilha", "Ativo"])

    leitura = plano["resultado_analise_oficial"]["leitura"]
    linhas_da_arruela = [linha for linha in leitura["linhas"] if linha["codigo_original"] in plano["arruelas"]]
    for linha in sorted(linhas_da_arruela, key=lambda l: l["linha_planilha"]):
        ws.append([
            linha["codigo_original"], linha["descricao"] or None, linha["cliente"] or None,
            _padrao_arruela(linha["codigo_original"]), linha["linha_planilha"], 1,
        ])

    codigos_so_de_outras_fontes = sorted(set(plano["arruelas"]) - {l["codigo_original"] for l in linhas_da_arruela})
    for codigo in codigos_so_de_outras_fontes:
        ws.append([codigo, None, None, _padrao_arruela(codigo), None, 1])

    ws_vinc = wb.create_sheet("Vinculos cliente-arruela")
    ws_vinc.append(["Cliente canônico", "Arruela", "Origem"])
    for item in plano["cliente_arruelas"]:
        ws_vinc.append([item["cliente"], item["arruela"], item["origem"]])

    ws_resumo = wb.create_sheet("Resumo")
    ws_resumo.append(["Item", "Valor"])
    ws_resumo.append(["Total de linhas na planilha classificadas como arruela", len(linhas_da_arruela)])
    ws_resumo.append(["Total de códigos de arruela únicos", len(plano["arruelas"])])
    ws_resumo.append(["Total de descrições coletadas", len(plano["arruela_descricoes_fontes"])])
    ws_resumo.append(["Total de vínculos cliente-arruela", len(plano["cliente_arruelas"])])
    ws_resumo.append(["Total de vínculos grupo-arruela", len(plano["grupo_arruelas"])])
    ws_resumo.append(["Duplicidades (mesmo código, 2+ linhas na planilha)", plano["total_duplicidades_arruela"]])
    ws_resumo.append(["Conflitos de descrição entre fontes", plano["total_conflitos_descricao_arruela"]])

    caminho = pasta_saida / "arruelas_identificadas.xlsx"
    wb.save(caminho)
    return caminho


def gravar_relatorio_simulacao_cpds_e_arruelas(resumo: dict, metadados: dict, pasta_saida: Path) -> Path:
    """simulacao_cpds_e_arruelas.xlsx: visão consolidada -- uma aba pra
    métricas de CPD, uma pra arruelas, uma pra pendências, uma resumo geral."""
    from openpyxl import Workbook

    pasta_saida.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    def preencher(ws, chaves):
        ws.append(["Item", "Valor"])
        for k in chaves:
            if k in resumo:
                ws.append([k, resumo[k]])

    ws_cpds = wb.active
    ws_cpds.title = "CPDs"
    ws_cpds.append([f"Gerado em: {metadados['gerado_em']}"])
    ws_cpds.append([])
    preencher(ws_cpds, [
        "total_linhas_planilha", "total_codigos_originais_unicos", "total_codigos_canonicos_unicos",
        "total_cpds_pai", "total_codigos_completos", "total_variacoes",
        "total_codigos_preenchidos_com_zero", "total_codigos_ja_5_digitos",
        "total_colisoes_normalizacao", "total_codigos_com_barra_convertidos", "total_colisoes_barra_ponto",
        "total_codigos_invalidos", "total_codigos_ambiguos",
        "total_vinculos_cliente_cpd", "total_vinculos_grupo_cpd", "total_descricoes_fontes",
        "cpds_sem_cliente", "cpds_somente_lista_oficial", "cpds_somente_pedidos", "cpds_ambas_fontes",
        "cpds_so_grupo_honda", "cpds_so_grupo_diversos", "cpds_ambos_grupos", "cpds_sem_grupo",
        "todos_cpds_ativos",
    ])

    ws_arruelas = wb.create_sheet("Arruelas")
    ws_arruelas.append([])
    preencher(ws_arruelas, [
        "total_arruelas_linhas", "total_arruelas_codigos_unicos", "total_arruelas_descricoes",
        "total_arruelas_clientes_relacionados", "total_vinculos_cliente_arruela", "total_vinculos_grupo_arruela",
        "total_duplicidades_arruela", "total_conflitos_descricao_arruela", "todas_arruelas_ativas",
        "nenhuma_arruela_em_cpds",
    ])

    ws_pend = wb.create_sheet("Pendencias")
    ws_pend.append(["Tipo", "Quantidade"])
    for tipo, qtd in sorted(resumo.get("pendencias_por_tipo", {}).items()):
        ws_pend.append([tipo, qtd])
    ws_pend.append(["TOTAL", resumo.get("total_pendencias")])

    ws_geral = wb.create_sheet("Resumo geral")
    ws_geral.append(["Item", "Valor"])
    for k, v in resumo.items():
        if isinstance(v, (int, float, bool, str)) or v is None:
            ws_geral.append([k, v])

    caminho = pasta_saida / "simulacao_cpds_e_arruelas.xlsx"
    wb.save(caminho)
    return caminho


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Migração definitiva de clientes e CPDs (simulação por padrão).")
    parser.add_argument("--planilha", type=Path, default=DEFAULT_PLANILHA)
    parser.add_argument("--banco", type=Path, default=DEFAULT_BANCO)
    parser.add_argument("--saida", type=Path, default=DEFAULT_SAIDA)
    parser.add_argument(
        "--config-clientes", type=Path, required=True,
        help="Caminho do JSON local com as regras de clientes (grupo_honda/aliases/placeholders_invalidos). "
             "Obrigatório -- nunca há regra hardcoded nem configuração vazia implícita. "
             "Ver config_local/regras_clientes_migracao.example.json.",
    )
    parser.add_argument("--confirmar", action="store_true", help="Aplica de verdade. Sem isso, é sempre um dry-run (--simular).")
    parser.add_argument("--texto-confirmacao", default=None, help=f"Evita prompt interativo; precisa ser exatamente '{TEXTO_CONFIRMACAO_ESPERADO}'.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    if not args.planilha.exists():
        print(f"ERRO: planilha não encontrada: {args.planilha}")
        return 1
    if not args.banco.exists():
        print(f"ERRO: banco não encontrado: {args.banco}")
        return 1

    try:
        config = config_clientes.carregar_config_clientes(args.config_clientes)
    except (FileNotFoundError, config_clientes.ConfigClientesInvalida) as exc:
        print(f"ERRO: {exc}")
        return 1

    conn_leitura = base_analise.abrir_somente_leitura(args.banco)
    try:
        plano = construir_plano(conn_leitura, args.planilha, config)
    finally:
        conn_leitura.close()

    resumo = simular(plano)
    metadados = {
        "planilha_origem": str(args.planilha),
        "banco_origem": str(args.banco),
        "gerado_em": plano["gerado_em"],
        "modo": "confirmar" if args.confirmar else "simular",
        # Só o hash -- nunca o conteúdo -- da configuração de clientes usada,
        # pra rastreabilidade sem expor nomes reais de clientes no relatório.
        "config_clientes_caminho": config.caminho,
        "config_clientes_hash_sha256": config.hash_sha256,
    }

    caminho_vinculos = gravar_relatorio_vinculos(plano, resumo, metadados, args.saida)
    caminho_normalizacao = gravar_relatorio_normalizacao(plano, metadados, args.saida)
    caminho_simulacao_5d = gravar_relatorio_simulacao_5_digitos(resumo, metadados, args.saida)
    caminho_invalidos = gravar_relatorio_codigos_invalidos(plano, metadados, args.saida)
    caminho_barra = gravar_relatorio_barra(plano, metadados, args.saida)
    caminho_arruelas = gravar_relatorio_arruelas(plano, metadados, args.saida)
    caminho_simulacao_completa = gravar_relatorio_simulacao_cpds_e_arruelas(resumo, metadados, args.saida)

    if not args.confirmar:
        caminho_relatorio = gravar_relatorio_simulacao(resumo, metadados, args.saida)
        print("Modo simulação (--simular): nada foi gravado no banco.")
        print(json.dumps(resumo, ensure_ascii=False, indent=2))
        print(f"Relatório de simulação: {caminho_relatorio}")
        print(f"Relatório de vínculos: {caminho_vinculos}")
        print(f"Relatório de normalização (5 dígitos): {caminho_normalizacao}")
        print(f"Relatório de simulação (5 dígitos): {caminho_simulacao_5d}")
        print(f"Relatório de códigos inválidos/ambíguos: {caminho_invalidos}")
        print(f"Relatório de barra->ponto: {caminho_barra}")
        print(f"Relatório de arruelas: {caminho_arruelas}")
        print(f"Relatório consolidado CPDs+Arruelas: {caminho_simulacao_completa}")
        return 0

    texto = args.texto_confirmacao
    if texto is None:
        texto = input(f"Digite exatamente '{TEXTO_CONFIRMACAO_ESPERADO}' para prosseguir: ")
    if texto.strip() != TEXTO_CONFIRMACAO_ESPERADO:
        print("Confirmação não corresponde ao texto exigido. Nada foi alterado.")
        return 1

    backup_dir = args.saida / "_backups_migracao"
    backup = dbm.criar_backup_sqlite(args.banco, backup_dir, prefixo="pcp_pre_migracao_clientes_cpds")
    if backup is None:
        print("Backup pré-migração falhou ou é inválido. Migração cancelada, nada foi alterado.")
        return 1
    print(f"Backup pré-migração validado: {backup}")

    conn = sqlite3.connect(str(args.banco))
    conn.execute("PRAGMA foreign_keys=ON")
    try:
        resultado_aplicacao = aplicar_migracao(conn, plano)
        conn.commit()
    except Exception:
        conn.rollback()
        print("ERRO ao aplicar a migração; rollback total executado. Nada foi alterado.")
        print(traceback.format_exc())
        return 1
    finally:
        conn.close()

    print("Migração aplicada com sucesso.")
    print(json.dumps(resultado_aplicacao, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
