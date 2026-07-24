"""Análise somente-leitura da lista oficial de CPDs (importacao/lista_oficial_cpds.xlsx)
combinada com manual_cpd e os CPDs usados em pedidos, para preparar (sem
aplicar) a consolidação de CPDs com ativo=1.

NÃO grava nada no banco analisado, NÃO altera pedidos, NÃO apaga nenhum dos
698 CPDs existentes, NÃO importa Excel automaticamente no servidor, NÃO
migra nada. Gera relatórios em `relatorios_migracao/` para decisão humana.

Suporta tanto .xlsx (openpyxl) quanto o formato antigo .xls (via xlrd) --
o arquivo real entregue está em .xls (Excel 97-2003), então este script lê
os dois formatos e escolhe o leitor pela extensão do arquivo.

DECISÃO OFICIAL já registrada (não deste script decidir de novo):
- Todos os CPDs consolidados entram com ativo=1 (cpds.ativo e
  cpd_variacoes.ativo), desativado_em NULL. Não se infere inatividade por
  ausência em pedidos nem por ausência na planilha oficial -- isso fica
  para a futura tela de configurações.
- Um cliente que só apareça em pedido na lixeira permanece ativo; aparecer
  só na lixeira não implica cliente inativo.
"""

from __future__ import annotations

import argparse
import difflib
import json
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import db_manutencao as dbm  # normalizar_cpd()
import scripts.analisar_migracao_clientes_cpds as base_analise  # reaproveita carregar_orders/manual_cpd, cópia segura
import scripts.config_clientes_migracao as config_clientes

DEFAULT_PLANILHA = ROOT / "importacao" / "lista_oficial_cpds.xlsx"
DEFAULT_BANCO = ROOT / "pcp.sqlite3"
DEFAULT_SAIDA = ROOT / "relatorios_migracao"

# Nomes de coluna reconhecidos (comparação sempre em minúsculas, sem acento).
_ALIASES_CPD = ("cpd", "código", "codigo")
_ALIASES_DESCRICAO = ("descrição", "descricao", "descrição do item", "descricao do item")
_ALIASES_CLIENTE = ("cliente",)
_COLUNAS_UTEIS_CONHECIDAS = ("peso", "material", "op", "cliente", "padrao embalagem", "peso padrao embalagem")


def _normalizar_texto_cabecalho(valor) -> str:
    import re
    import unicodedata
    txt = str(valor or "").strip().lower()
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode("ascii")
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


# ---------------------------------------------------------------------------
# Leitura da planilha (.xls legado via xlrd, .xlsx via openpyxl)
# ---------------------------------------------------------------------------

def _abrir_planilha(caminho: Path):
    sufixo = caminho.suffix.lower()
    if sufixo == ".xls":
        import xlrd
        return ("xlrd", xlrd.open_workbook(str(caminho), on_demand=True))
    if sufixo == ".xlsx":
        from openpyxl import load_workbook
        return ("openpyxl", load_workbook(caminho, data_only=True, read_only=True))
    raise ValueError(f"Formato de planilha não suportado: {sufixo}")


def _linhas_da_aba(motor: str, wb, nome_aba: str):
    """Gera (indice_linha_0based, [valores_da_linha]) para uma aba, não importa o motor."""
    if motor == "xlrd":
        ws = wb.sheet_by_name(nome_aba)
        for r in range(ws.nrows):
            yield r, [ws.cell_value(r, c) for c in range(ws.ncols)], ws
    else:
        ws = wb[nome_aba]
        for r, row in enumerate(ws.iter_rows(values_only=True)):
            yield r, list(row), ws


def _tipo_celula_numerico(motor: str, ws, r: int, c: int) -> bool:
    if motor == "xlrd":
        import xlrd
        return ws.cell_type(r, c) == xlrd.XL_CELL_NUMBER
    # openpyxl (read_only): valor já vem como int/float ou str; se for número
    # (int/float) tratamos como "numérico" para fins de zero à esquerda.
    valor = ws.cell(row=r + 1, column=c + 1).value
    return isinstance(valor, (int, float))


def detectar_aba_principal(caminho: Path) -> dict:
    """Identifica a aba com a lista principal, sem assumir nome de arquivo/aba.

    Critério: entre as abas do arquivo, procura uma linha (nas primeiras 5)
    que tenha, na mesma linha, uma coluna reconhecida como CPD e uma
    reconhecida como descrição. Entre as abas candidatas, escolhe a que tem
    mais linhas de dados e mais colunas úteis reconhecidas (ex.: contém
    também uma coluna de cliente).
    """
    motor, wb = _abrir_planilha(caminho)
    nomes_abas = wb.sheet_names() if motor == "xlrd" else wb.sheetnames

    candidatos = []
    for nome_aba in nomes_abas:
        linhas_amostra = []
        for r, valores, ws in _linhas_da_aba(motor, wb, nome_aba):
            linhas_amostra.append(valores)
            if r >= 4:
                break
        if not linhas_amostra:
            continue

        melhor_linha = None
        for idx, valores in enumerate(linhas_amostra):
            cabecalhos = {_normalizar_texto_cabecalho(v): i for i, v in enumerate(valores) if v not in (None, "")}
            col_cpd = next((i for chave, i in cabecalhos.items() if chave in _ALIASES_CPD), None)
            col_desc = next((i for chave, i in cabecalhos.items() if chave in _ALIASES_DESCRICAO), None)
            if col_cpd is not None and col_desc is not None:
                melhor_linha = idx
                break
        if melhor_linha is None:
            continue

        cabecalhos_finais = {_normalizar_texto_cabecalho(v): i for i, v in enumerate(linhas_amostra[melhor_linha]) if v not in (None, "")}
        col_cpd = next(i for chave, i in cabecalhos_finais.items() if chave in _ALIASES_CPD)
        col_desc = next(i for chave, i in cabecalhos_finais.items() if chave in _ALIASES_DESCRICAO)
        col_cliente = next((i for chave, i in cabecalhos_finais.items() if chave in _ALIASES_CLIENTE), None)
        colunas_uteis = {chave: i for chave, i in cabecalhos_finais.items() if chave in _COLUNAS_UTEIS_CONHECIDAS}

        total_linhas = sum(1 for _ in _linhas_da_aba(motor, wb, nome_aba))
        score = total_linhas + len(cabecalhos_finais) * 2 + (10_000 if col_cliente is not None else 0)

        candidatos.append({
            "aba": nome_aba,
            "linha_cabecalho_0based": melhor_linha,
            "coluna_cpd": col_cpd,
            "coluna_descricao": col_desc,
            "coluna_cliente": col_cliente,
            "colunas_uteis": colunas_uteis,
            "total_linhas": total_linhas,
            "score": score,
        })

    if motor == "xlrd":
        wb.release_resources()

    if not candidatos:
        raise RuntimeError(f"Nenhuma aba com colunas de CPD + descrição reconhecidas em {caminho}")

    candidatos.sort(key=lambda c: c["score"], reverse=True)
    return {"escolhida": candidatos[0], "todas_candidatas": candidatos}


def ler_lista_oficial(caminho: Path) -> dict:
    """Lê a aba principal detectada, preservando código como texto e zeros
    à esquerda quando a célula já é texto. Quando a célula é numérica (o
    Excel guarda o CPD como número), não há como recuperar um zero à
    esquerda que porventura tenha existido -- isso é sinalizado
    explicitamente em cada linha (`zero_esquerda_desconhecido`), nunca
    inventado.
    """
    deteccao = detectar_aba_principal(caminho)
    info = deteccao["escolhida"]
    motor, wb = _abrir_planilha(caminho)

    linhas = []
    for r, valores, ws in _linhas_da_aba(motor, wb, info["aba"]):
        if r <= info["linha_cabecalho_0based"]:
            continue
        if info["coluna_cpd"] >= len(valores):
            continue
        cpd_bruto = valores[info["coluna_cpd"]]
        if cpd_bruto in (None, ""):
            continue

        numerico = _tipo_celula_numerico(motor, ws, r, info["coluna_cpd"])
        if numerico:
            valor_float = float(cpd_bruto)
            cpd_texto = str(int(valor_float)) if valor_float == int(valor_float) else str(valor_float)
        else:
            cpd_texto = str(cpd_bruto).strip()

        descricao = valores[info["coluna_descricao"]] if info["coluna_descricao"] < len(valores) else ""
        cliente = ""
        if info["coluna_cliente"] is not None and info["coluna_cliente"] < len(valores):
            cliente = valores[info["coluna_cliente"]]

        linhas.append({
            "linha_planilha": r + 1,  # 1-based, como o usuário veria no Excel
            "codigo_original": cpd_texto,
            "descricao": str(descricao or "").strip(),
            "cliente": str(cliente or "").strip(),
            "zero_esquerda_desconhecido": numerico,
        })

    if motor == "xlrd":
        wb.release_resources()

    return {"deteccao": deteccao, "linhas": linhas}


# ---------------------------------------------------------------------------
# 3) Duplicidades na planilha (código completo exatamente repetido)
# ---------------------------------------------------------------------------

def identificar_duplicidades_planilha(linhas: list[dict]) -> list[dict]:
    por_codigo: dict[str, list[dict]] = defaultdict(list)
    for linha in linhas:
        por_codigo[linha["codigo_original"]].append(linha)

    duplicidades = []
    for codigo, ocorrencias in por_codigo.items():
        if len(ocorrencias) <= 1:
            continue
        descricoes = [o["descricao"] for o in ocorrencias]
        descricoes_distintas = sorted(set(d for d in descricoes if d))
        if len(descricoes_distintas) <= 1:
            escolhida = descricoes_distintas[0] if descricoes_distintas else ""
            justificativa = "Todas as ocorrências têm a mesma descrição (ou só uma está preenchida); sem conflito."
        else:
            escolhida = None
            justificativa = "Descrições diferentes para o mesmo código -- não escolhida automaticamente, decisão humana necessária."
        duplicidades.append({
            "codigo": codigo,
            "linhas_de_origem": [o["linha_planilha"] for o in ocorrencias],
            "descricoes_encontradas": descricoes,
            "descricao_que_seria_escolhida": escolhida,
            "justificativa": justificativa,
        })
    return duplicidades


# ---------------------------------------------------------------------------
# 1) Investigação de zeros à esquerda (código numérico na planilha oficial
#    x código existente em manual_cpd/pedidos, potencialmente com zero à
#    esquerda que a planilha perdeu)
# ---------------------------------------------------------------------------

def _chave_numerica(codigo: str) -> str | None:
    """Valor numérico do código (zeros à esquerda removidos), ou None se o
    código tiver qualquer caractere não-dígito (ex.: 'XY-002', '04772.1')."""
    codigo = str(codigo or "").strip()
    if codigo.isdigit():
        return str(int(codigo))
    return None


def _normalizar_descricao_basica(desc: str) -> str:
    txt = str(desc or "").strip().upper()
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode("ascii")
    txt = re.sub(r"[^\w\s]", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def detectar_candidatos_zero_esquerda(leitura: dict, manual_cpd_rows: list[dict], orders: list[dict]) -> dict:
    """Para cada código da planilha que veio de célula numérica (portanto
    pode ter perdido um zero à esquerda), procura candidatos entre
    manual_cpd e os CPDs usados em pedidos cujo valor numérico (zeros à
    esquerda removidos) seja igual, mas cujo texto seja diferente. Nunca
    une nada -- só gera candidatos com nível de confiança, para decisão
    humana.
    """
    cliente_por_codigo: dict[str, set[str]] = defaultdict(set)
    descricao_pedido_por_codigo: dict[str, set[str]] = defaultdict(set)
    for o in orders:
        cpd = str(o.get("cpd") or "").strip()
        if not cpd:
            continue
        cliente = str(o.get("cliente") or "").strip()
        if cliente:
            cliente_por_codigo[cpd].add(cliente)
        desc = str(o.get("descricao") or "").strip()
        if desc:
            descricao_pedido_por_codigo[cpd].add(desc)

    descricao_manual_por_codigo = {row["codigo"]: str(row["descricao"] or "").strip() for row in manual_cpd_rows}

    existentes_por_chave: dict[str, list[str]] = defaultdict(list)
    for codigo in set(descricao_manual_por_codigo) | set(cliente_por_codigo) | set(descricao_pedido_por_codigo):
        chave = _chave_numerica(codigo)
        if chave is not None:
            existentes_por_chave[chave].append(codigo)

    candidatos = []
    colisoes = []
    for linha in leitura["linhas"]:
        if not linha["zero_esquerda_desconhecido"]:
            continue
        codigo_oficial = linha["codigo_original"]
        chave = _chave_numerica(codigo_oficial)
        if chave is None:
            continue
        existentes = [c for c in existentes_por_chave.get(chave, []) if c != codigo_oficial]
        if not existentes:
            continue
        if len(existentes) > 1:
            colisoes.append({"chave_numerica": chave, "codigo_oficial": codigo_oficial, "codigos_existentes_conflitantes": sorted(existentes)})

        for codigo_existente in existentes:
            desc_existente = descricao_manual_por_codigo.get(codigo_existente) or next(iter(descricao_pedido_por_codigo.get(codigo_existente, [])), "")
            clientes_existentes = sorted(cliente_por_codigo.get(codigo_existente, set()))
            cliente_oficial = linha["cliente"]

            similaridade_desc = difflib.SequenceMatcher(
                None, _normalizar_descricao_basica(linha["descricao"]), _normalizar_descricao_basica(desc_existente)
            ).ratio() if (linha["descricao"] and desc_existente) else 0.0

            # Três estados possíveis: compatível (evidência de cliente aponta pro
            # mesmo cliente), incompatível (evidência aponta pra clientes
            # diferentes) ou sem evidência (nenhum dos dois lados tem
            # informação suficiente pra comparar). "Sem evidência" é neutro --
            # não deve contar como suporte a favor da confiança, só
            # "incompatível" derruba a confiança.
            tem_evidencia_cliente = bool(cliente_oficial) and bool(clientes_existentes)
            cliente_compativel = tem_evidencia_cliente and _normalizar_descricao_basica(cliente_oficial) in {
                _normalizar_descricao_basica(c) for c in clientes_existentes
            }
            cliente_incompativel = tem_evidencia_cliente and not cliente_compativel

            if similaridade_desc >= 0.85 and not cliente_incompativel:
                confianca = "alta"
            elif (similaridade_desc >= 0.6 and not cliente_incompativel) or cliente_compativel:
                confianca = "media"
            else:
                confianca = "baixa"

            candidatos.append({
                "cpd_lista_oficial": codigo_oficial,
                "cpd_existente": codigo_existente,
                "descricao_oficial": linha["descricao"],
                "descricao_existente": desc_existente,
                "cliente_oficial": cliente_oficial,
                "clientes_existentes": clientes_existentes,
                "similaridade_descricao": round(similaridade_desc, 3),
                "nivel_confianca": confianca,
                "justificativa": (
                    f"Mesmo valor numérico ({chave}); similaridade de descrição {similaridade_desc:.0%}; "
                    f"cliente {'compatível' if cliente_compativel else ('incompatível' if cliente_incompativel else 'sem evidência')}."
                ),
                "colisao": len(existentes) > 1,
            })

    return {
        "candidatos": candidatos,
        "total_candidatos": len(candidatos),
        "total_alta_confianca": sum(1 for c in candidatos if c["nivel_confianca"] == "alta"),
        "total_media_confianca": sum(1 for c in candidatos if c["nivel_confianca"] == "media"),
        "total_baixa_confianca": sum(1 for c in candidatos if c["nivel_confianca"] == "baixa"),
        "colisoes": colisoes,
    }


# ---------------------------------------------------------------------------
# 2) Clientes da lista oficial x clientes dos pedidos
# ---------------------------------------------------------------------------

def analisar_clientes_lista_oficial(leitura: dict, orders: list[dict], grupo_honda: set[str]) -> dict:
    clientes_planilha_raw: dict[str, int] = defaultdict(int)
    celulas_vazias = 0
    for linha in leitura["linhas"]:
        cliente = linha["cliente"]
        if not cliente:
            celulas_vazias += 1
            continue
        clientes_planilha_raw[cliente] += 1

    clientes_pedidos_raw: dict[str, int] = defaultdict(int)
    for o in orders:
        cliente = str(o.get("cliente") or "").strip()
        if cliente:
            clientes_pedidos_raw[cliente] += 1

    norm = base_analise.normalizar_cliente_para_comparacao
    normalizados_planilha = {norm(c): c for c in clientes_planilha_raw}
    normalizados_pedidos = {norm(c): c for c in clientes_pedidos_raw}

    em_ambas = sorted(set(normalizados_planilha) & set(normalizados_pedidos))
    somente_planilha = sorted(set(normalizados_planilha) - set(normalizados_pedidos))
    somente_pedidos = sorted(set(normalizados_pedidos) - set(normalizados_planilha))

    possiveis_aliases = []
    for chave_p in somente_planilha:
        nome_p = normalizados_planilha[chave_p]
        for chave_o in somente_pedidos:
            nome_o = normalizados_pedidos[chave_o]
            score = base_analise.similaridade(chave_p, chave_o)
            abreviacao = base_analise.eh_possivel_abreviacao(chave_p, chave_o) or base_analise.eh_possivel_abreviacao(chave_o, chave_p)
            if abreviacao or score >= 0.72:
                possiveis_aliases.append({
                    "cliente_planilha": nome_p,
                    "cliente_pedidos": nome_o,
                    "similaridade": round(score, 3),
                    "possivel_abreviacao": abreviacao,
                })

    def grupo_sugerido(nome: str) -> str:
        return "Honda" if nome in grupo_honda else "Diversos"

    todos_clientes_planilha = sorted(clientes_planilha_raw)
    sugestao_grupos = {c: {"grupo_sugerido": grupo_sugerido(c), "ativo": 1} for c in todos_clientes_planilha}

    return {
        "clientes_planilha_com_contagem_cpds": dict(sorted(clientes_planilha_raw.items())),
        "clientes_presentes_em_ambas_fontes": [normalizados_planilha[c] for c in em_ambas],
        "clientes_somente_na_lista_oficial": [normalizados_planilha[c] for c in somente_planilha],
        "clientes_somente_em_pedidos": [normalizados_pedidos[c] for c in somente_pedidos],
        "possiveis_aliases_planilha_x_pedidos": possiveis_aliases,
        "total_celulas_cliente_vazias_na_planilha": celulas_vazias,
        "sugestao_grupos": sugestao_grupos,
    }


# ---------------------------------------------------------------------------
# 3) Classificação conservadora dos conflitos de descrição (A/B/C/D)
# ---------------------------------------------------------------------------

_PREFIXO_PARAF_RE = re.compile(r"^(PARAFUSO|PARAF)\.?\s+")


def _normalizar_cosmetico(desc: str) -> str:
    """Só remove diferenças de formatação: maiúsculas/minúsculas, espaços
    repetidos, acentos, pontuação e separador decimal (vírgula vs ponto).
    Nunca remove números, letras de código técnico, material, classe,
    tratamento, acabamento, norma, cliente ou desenho -- essas diferenças
    devem continuar aparecendo depois da normalização.
    """
    txt = str(desc or "").strip().upper()
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode("ascii")
    # Marca o separador decimal (vírgula OU ponto entre dois dígitos) com "_"
    # (caractere de palavra, sobrevive à remoção de pontuação abaixo) antes de
    # tirar qualquer outra pontuação -- inclusive pontos soltos que não são
    # separador decimal (ex.: final de frase), que devem mesmo ser removidos.
    txt = re.sub(r"(?<=\d)[.,](?=\d)", "_", txt)
    txt = re.sub(r"[^\w\s]", " ", txt)  # remove toda pontuação restante
    txt = txt.replace("_", ".")  # restaura separador decimal como ponto
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def _remover_prefixo_paraf(desc_cosmetica: str) -> str:
    return _PREFIXO_PARAF_RE.sub("", desc_cosmetica).strip()


def classificar_par_descricoes(desc_a: str, desc_b: str) -> tuple[str, str]:
    """Classifica o par em A/B/C/D. Retorna (categoria, detalhe)."""
    norm_a = _normalizar_cosmetico(desc_a)
    norm_b = _normalizar_cosmetico(desc_b)
    if norm_a == norm_b:
        return "A", "Mesma informação técnica, só difere em formatação (caixa/espaços/acentos/pontuação/separador decimal)."

    sem_paraf_a = _remover_prefixo_paraf(norm_a)
    sem_paraf_b = _remover_prefixo_paraf(norm_b)
    if sem_paraf_a == sem_paraf_b:
        return "B", "Diferem só pelo prefixo 'PARAF'/'PARAFUSO' -- convenção de descrição, não informação técnica."

    similaridade = difflib.SequenceMatcher(None, sem_paraf_a, sem_paraf_b).ratio()
    if similaridade >= 0.5:
        return "C", f"Descrições parecidas (similaridade {similaridade:.0%}) mas com diferença técnica (dimensão/material/acabamento/etc.) -- não decidido automaticamente."
    return "D", f"Descrições pouco parecidas (similaridade {similaridade:.0%}) -- provável erro, duplicidade ou uso incorreto do código."


def classificar_conflitos_tecnicos(conflitos_para_decisao: list[dict]) -> dict:
    _ORDEM_GRAVIDADE = {"A": 0, "B": 1, "C": 2, "D": 3}
    classificados = []
    for conflito in conflitos_para_decisao:
        descricoes = []
        for fonte in ("lista_oficial", "manual_cpd", "pedidos"):
            descricoes.extend(conflito["descricoes_por_fonte"].get(fonte, []))
        descricoes = list(dict.fromkeys(descricoes))  # remove duplicatas, preserva ordem

        referencia = descricoes[0]
        pares = []
        pior_categoria = "A"
        for outra in descricoes[1:]:
            categoria, detalhe = classificar_par_descricoes(referencia, outra)
            pares.append({"descricao_a": referencia, "descricao_b": outra, "categoria": categoria, "detalhe": detalhe})
            if _ORDEM_GRAVIDADE[categoria] > _ORDEM_GRAVIDADE[pior_categoria]:
                pior_categoria = categoria

        oficial = conflito["descricoes_por_fonte"].get("lista_oficial", [])
        descricao_canonica_proposta = oficial[0] if (pior_categoria in ("A", "B") and len(oficial) == 1) else None

        classificados.append({
            "codigo": conflito["codigo"],
            "categoria": pior_categoria,
            "pares_comparados": pares,
            "descricoes_por_fonte": conflito["descricoes_por_fonte"],
            "descricao_canonica_proposta": descricao_canonica_proposta,
        })

    contagem = {cat: sum(1 for c in classificados if c["categoria"] == cat) for cat in ("A", "B", "C", "D")}
    return {"classificados": classificados, "contagem_por_categoria": contagem}


# ---------------------------------------------------------------------------
# 4) Linhas completas da planilha para um conjunto de números de linha
#    (usado para mostrar as duplicidades por inteiro, todas as colunas)
# ---------------------------------------------------------------------------

def obter_linhas_completas_planilha(caminho: Path, numeros_linha: set[int]) -> dict[int, dict]:
    deteccao = detectar_aba_principal(caminho)
    info = deteccao["escolhida"]
    motor, wb = _abrir_planilha(caminho)

    cabecalho_valores = None
    resultado: dict[int, dict] = {}
    for r, valores, ws in _linhas_da_aba(motor, wb, info["aba"]):
        if r == info["linha_cabecalho_0based"]:
            cabecalho_valores = valores
            continue
        linha_1based = r + 1
        if linha_1based not in numeros_linha:
            continue
        nomes_colunas = [str(v).strip() if v not in (None, "") else f"coluna_{i}" for i, v in enumerate(cabecalho_valores or [])]
        resultado[linha_1based] = {nomes_colunas[i]: valores[i] for i in range(min(len(nomes_colunas), len(valores)))}

    if motor == "xlrd":
        wb.release_resources()
    return resultado


def comparar_linhas_duplicidade(linha_a: dict, linha_b: dict) -> list[str]:
    diferencas = []
    todas_colunas = sorted(set(linha_a) | set(linha_b))
    for coluna in todas_colunas:
        va, vb = linha_a.get(coluna), linha_b.get(coluna)
        if va != vb:
            diferencas.append(f"{coluna}: {va!r} != {vb!r}")
    return diferencas


def buscar_valor_coluna(linha_completa: dict, *aliases: str):
    """Busca um valor em `linha_completa` (dict {nome_da_coluna: valor}, com
    o nome da coluna exatamente como veio do cabeçalho da planilha, preservando
    o casing original) por um dos `aliases`, comparando de forma normalizada
    (minúsculas, sem acento) -- já que o casing do cabeçalho varia entre
    planilhas (ex.: "DESCRIÇÃO" maiúsculo na lista oficial real, mas poderia
    vir como "Descrição" em outra fonte).
    """
    normalizados = {_normalizar_texto_cabecalho(k): v for k, v in linha_completa.items()}
    for alias in aliases:
        chave = _normalizar_texto_cabecalho(alias)
        if chave in normalizados and normalizados[chave] not in (None, ""):
            return normalizados[chave]
    return None


# ---------------------------------------------------------------------------
# 2/4/5) Consolidação das 3 fontes
# ---------------------------------------------------------------------------

def consolidar_fontes(manual_cpd_rows: list[dict], orders: list[dict], linhas_planilha: list[dict]) -> dict:
    """União de manual_cpd + CPDs de pedidos + lista oficial. Nunca exclui um
    código por não aparecer em uma das fontes. Remove só duplicidade exata
    do mesmo código completo (mantendo a 1ª ocorrência de cada fonte, na
    ordem de prioridade oficial > manual_cpd > pedidos, só quando não há
    conflito de descrição)."""

    # codigo_completo -> {"descricoes": {fonte: [descricoes]}, "linhas_planilha": [...]}
    por_codigo: dict[str, dict] = {}

    def registrar(codigo: str, descricao: str, fonte: str, linha_origem=None):
        item = por_codigo.setdefault(codigo, {"descricoes": defaultdict(set), "linhas_planilha": [], "fontes": set()})
        if descricao:
            item["descricoes"][fonte].add(descricao)
        item["fontes"].add(fonte)
        if linha_origem is not None:
            item["linhas_planilha"].append(linha_origem)

    for linha in linhas_planilha:
        registrar(linha["codigo_original"], linha["descricao"], "lista_oficial", linha["linha_planilha"])
    for row in manual_cpd_rows:
        registrar(row["codigo"], str(row["descricao"] or "").strip(), "manual_cpd")
    for o in orders:
        cpd = str(o.get("cpd") or "").strip()
        if cpd:
            registrar(cpd, str(o.get("descricao") or "").strip(), "pedidos")

    novos_da_lista = {c for c, i in por_codigo.items() if i["fontes"] == {"lista_oficial"}}
    existentes_ausentes_da_lista = {
        row["codigo"] for row in manual_cpd_rows if "lista_oficial" not in por_codigo.get(row["codigo"], {}).get("fontes", set())
    }

    consolidados = {}
    conflitos_para_decisao = []
    _PRIORIDADE_FONTES = ("lista_oficial", "manual_cpd", "pedidos")

    for codigo, info in por_codigo.items():
        cpd_info = dbm.normalizar_cpd(codigo)
        todas_descricoes = set()
        for descs in info["descricoes"].values():
            todas_descricoes |= descs

        if len(todas_descricoes) <= 1:
            # Sem conflito: no máximo uma descrição não-vazia entre todas as fontes.
            descricao_final = next(iter(todas_descricoes), "")
            conflito = False
        else:
            # Só decide sozinho se a fonte de maior prioridade disponível tiver,
            # ela própria, uma única descrição -- e mesmo assim isso é reportado
            # como sugestão (não aplicado), porque as fontes de menor
            # prioridade divergem. Item 4: "colocar em relatório para decisão
            # humana" sempre que houver descrições divergentes relevantes.
            sugestao_por_prioridade = None
            for fonte in _PRIORIDADE_FONTES:
                candidatas = info["descricoes"].get(fonte)
                if candidatas:
                    sugestao_por_prioridade = next(iter(candidatas)) if len(candidatas) == 1 else None
                    break
            descricao_final = None
            conflito = True
            conflitos_para_decisao.append({
                "codigo": codigo,
                "descricoes_por_fonte": {f: sorted(d) for f, d in info["descricoes"].items()},
                "descricao_sugerida_por_prioridade": sugestao_por_prioridade,
                "motivo": "Descrições divergentes entre fontes para o mesmo código.",
            })

        consolidados[codigo] = {
            "codigo_original": codigo,
            "codigo_pai": cpd_info["codigo_pai"],
            "extensao": cpd_info["extensao"],
            "descricao": descricao_final,
            "conflito_descricao": conflito,
            "fontes": sorted(info["fontes"]),
            "ativo": 1,  # decisão oficial: todos ativos
            "desativado_em": None,
        }

    pais: dict[str, list[dict]] = defaultdict(list)
    for item in consolidados.values():
        pais[item["codigo_pai"]].append(item)

    pais_criados_sem_variacao_original = []
    for pai, itens in pais.items():
        tem_pai_sem_extensao = any(i["extensao"] is None for i in itens)
        if not tem_pai_sem_extensao:
            # Só existem variações (ex.: 31494.1) -- cria o pai 31494, mantém a extensão, ambos ativos.
            pais_criados_sem_variacao_original.append(pai)
            consolidados[pai] = {
                "codigo_original": pai,
                "codigo_pai": pai,
                "extensao": None,
                "descricao": None,  # descrição do pai fica pra revisão humana (sugestão: repetir a da variação)
                "conflito_descricao": False,
                "fontes": ["gerado_automaticamente_pai_ausente"],
                "ativo": 1,
                "desativado_em": None,
                "pai_criado_automaticamente": True,
            }
            pais[pai].append(consolidados[pai])

    return {
        "consolidados": consolidados,
        "pais": {p: v for p, v in pais.items()},
        "novos_da_lista_oficial": sorted(novos_da_lista),
        "existentes_em_manual_cpd_ausentes_da_lista": sorted(existentes_ausentes_da_lista),
        "conflitos_para_decisao": conflitos_para_decisao,
        "pais_criados_automaticamente_por_variacao_orfa": sorted(pais_criados_sem_variacao_original),
    }


# ---------------------------------------------------------------------------
# Orquestração + relatórios
# ---------------------------------------------------------------------------

def executar_analise_lista_oficial(conn, caminho_planilha: Path, grupo_honda: set[str]) -> dict:
    orders = base_analise.carregar_orders(conn)
    manual_cpd_rows = base_analise.carregar_manual_cpd(conn)
    leitura = ler_lista_oficial(caminho_planilha)
    duplicidades = identificar_duplicidades_planilha(leitura["linhas"])
    consolidacao = consolidar_fontes(manual_cpd_rows, orders, leitura["linhas"])
    zeros_esquerda = detectar_candidatos_zero_esquerda(leitura, manual_cpd_rows, orders)
    clientes = analisar_clientes_lista_oficial(leitura, orders, grupo_honda)
    conflitos_tecnicos = classificar_conflitos_tecnicos(consolidacao["conflitos_para_decisao"])

    numeros_linha_duplicidades = {ln for d in duplicidades for ln in d["linhas_de_origem"]}
    linhas_completas_duplicidades = obter_linhas_completas_planilha(caminho_planilha, numeros_linha_duplicidades)
    for d in duplicidades:
        linhas = [linhas_completas_duplicidades.get(ln, {}) for ln in d["linhas_de_origem"]]
        d["linhas_completas"] = linhas
        d["diferencas"] = comparar_linhas_duplicidade(linhas[0], linhas[1]) if len(linhas) == 2 else []

    total_variacoes = sum(1 for i in consolidacao["consolidados"].values() if i["extensao"] is not None)
    total_pais = len(consolidacao["pais"])

    total_decisoes_humanas = (
        zeros_esquerda["total_candidatos"]
        + len(clientes["clientes_somente_na_lista_oficial"])
        + conflitos_tecnicos["contagem_por_categoria"]["C"]
        + conflitos_tecnicos["contagem_por_categoria"]["D"]
        + len(duplicidades)
    )

    resumo = {
        "total_linhas_planilha": len(leitura["linhas"]),
        "total_codigos_unicos_planilha": len({l["codigo_original"] for l in leitura["linhas"]}),
        "total_atual_manual_cpd": len(manual_cpd_rows),
        "total_consolidado_codigos_completos": len(consolidacao["consolidados"]),
        "total_consolidado_cpds_pai": total_pais,
        "total_variacoes": total_variacoes,
        "total_novos_codigos_da_lista": len(consolidacao["novos_da_lista_oficial"]),
        "total_conflitos_descricao": len(consolidacao["conflitos_para_decisao"]),
        "total_duplicidades_na_planilha": len(duplicidades),
        "total_pais_criados_automaticamente": len(consolidacao["pais_criados_automaticamente_por_variacao_orfa"]),
        "todos_ficariam_ativos": all(i["ativo"] == 1 for i in consolidacao["consolidados"].values()),
        "todos_com_desativado_em_nulo": all(i["desativado_em"] is None for i in consolidacao["consolidados"].values()),
        "total_candidatos_zero_esquerda": zeros_esquerda["total_candidatos"],
        "total_zero_esquerda_alta_confianca": zeros_esquerda["total_alta_confianca"],
        "total_zero_esquerda_media_confianca": zeros_esquerda["total_media_confianca"],
        "total_zero_esquerda_baixa_confianca": zeros_esquerda["total_baixa_confianca"],
        "total_colisoes_zero_esquerda": len(zeros_esquerda["colisoes"]),
        "total_clientes_somente_lista_oficial": len(clientes["clientes_somente_na_lista_oficial"]),
        "total_clientes_somente_pedidos": len(clientes["clientes_somente_em_pedidos"]),
        "total_clientes_em_ambas_fontes": len(clientes["clientes_presentes_em_ambas_fontes"]),
        "conflitos_categoria_A": conflitos_tecnicos["contagem_por_categoria"]["A"],
        "conflitos_categoria_B": conflitos_tecnicos["contagem_por_categoria"]["B"],
        "conflitos_categoria_C": conflitos_tecnicos["contagem_por_categoria"]["C"],
        "conflitos_categoria_D": conflitos_tecnicos["contagem_por_categoria"]["D"],
        "total_decisoes_humanas_necessarias": total_decisoes_humanas,
    }

    return {
        "leitura": leitura,
        "duplicidades": duplicidades,
        "consolidacao": consolidacao,
        "zeros_esquerda": zeros_esquerda,
        "clientes": clientes,
        "conflitos_tecnicos": conflitos_tecnicos,
        "resumo": resumo,
    }


def gravar_relatorios(resultado: dict, metadados: dict, pasta_saida: Path) -> list[Path]:
    from openpyxl import Workbook

    pasta_saida.mkdir(parents=True, exist_ok=True)
    arquivos = []

    caminho_json = pasta_saida / "resumo_lista_oficial_cpds.json"
    with caminho_json.open("w", encoding="utf-8") as f:
        json.dump(
            {"metadados": metadados, "resumo": resultado["resumo"], "deteccao_aba": resultado["leitura"]["deteccao"]},
            f, ensure_ascii=False, indent=2, default=str,
        )
    arquivos.append(caminho_json)

    def nova_planilha(nome_arquivo: str, cabecalho: list[str], linhas: list[list]) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados"
        ws.append([f"Planilha oficial analisada em: {metadados['gerado_em']}"])
        ws.append([f"Arquivo: {metadados['planilha_origem']}"])
        ws.append([])
        ws.append(cabecalho)
        for linha in linhas:
            ws.append(linha)
        caminho = pasta_saida / nome_arquivo
        wb.save(caminho)
        arquivos.append(caminho)
        return caminho

    nova_planilha(
        "comparacao_lista_oficial_manual.xlsx",
        ["Código", "Nas fontes", "Descrição final", "Conflito?"],
        [
            [c, ", ".join(i["fontes"]), i["descricao"], i["conflito_descricao"]]
            for c, i in sorted(resultado["consolidacao"]["consolidados"].items())
        ],
    )

    nova_planilha(
        "cpds_consolidados_para_migracao.xlsx",
        ["Código original", "CPD pai", "Extensão", "Descrição", "Ativo", "Desativado em", "Fontes"],
        [
            [i["codigo_original"], i["codigo_pai"], i["extensao"], i["descricao"], i["ativo"], i["desativado_em"], ", ".join(i["fontes"])]
            for i in sorted(resultado["consolidacao"]["consolidados"].values(), key=lambda x: x["codigo_original"])
        ],
    )

    nova_planilha(
        "duplicidades_lista_oficial.xlsx",
        ["Código", "Linhas de origem", "Descrições encontradas", "Descrição que seria escolhida", "Justificativa"],
        [
            [d["codigo"], ", ".join(map(str, d["linhas_de_origem"])), " | ".join(d["descricoes_encontradas"]), d["descricao_que_seria_escolhida"], d["justificativa"]]
            for d in resultado["duplicidades"]
        ],
    )

    nova_planilha(
        "conflitos_descricao_para_decisao.xlsx",
        ["Código", "Descrições por fonte (JSON)", "Sugestão por prioridade", "Motivo"],
        [
            [c["codigo"], json.dumps(c["descricoes_por_fonte"], ensure_ascii=False), c["descricao_sugerida_por_prioridade"], c["motivo"]]
            for c in resultado["consolidacao"]["conflitos_para_decisao"]
        ],
    )

    nova_planilha(
        "cpds_novos_da_lista.xlsx",
        ["Código novo (só na lista oficial)"],
        [[c] for c in resultado["consolidacao"]["novos_da_lista_oficial"]],
    )

    nova_planilha(
        "cpds_existentes_ausentes_da_lista.xlsx",
        ["Código em manual_cpd, ausente da lista oficial"],
        [[c] for c in resultado["consolidacao"]["existentes_em_manual_cpd_ausentes_da_lista"]],
    )

    nova_planilha(
        "candidatos_zeros_esquerda.xlsx",
        ["CPD lista oficial", "CPD existente", "Descrição oficial", "Descrição existente", "Cliente oficial",
         "Clientes existentes", "Nível de confiança", "Justificativa", "Colisão?"],
        [
            [c["cpd_lista_oficial"], c["cpd_existente"], c["descricao_oficial"], c["descricao_existente"],
             c["cliente_oficial"], ", ".join(c["clientes_existentes"]), c["nivel_confianca"], c["justificativa"], c["colisao"]]
            for c in resultado["zeros_esquerda"]["candidatos"]
        ],
    )

    clientes = resultado["clientes"]
    linhas_clientes = []
    for cliente, contagem in clientes["clientes_planilha_com_contagem_cpds"].items():
        sugestao = clientes["sugestao_grupos"].get(cliente, {})
        origem = (
            "ambas" if cliente in clientes["clientes_presentes_em_ambas_fontes"]
            else "somente_lista_oficial" if cliente in clientes["clientes_somente_na_lista_oficial"]
            else "outro"
        )
        linhas_clientes.append([cliente, contagem, origem, sugestao.get("grupo_sugerido"), sugestao.get("ativo")])
    for cliente in clientes["clientes_somente_em_pedidos"]:
        linhas_clientes.append([cliente, 0, "somente_pedidos", None, 1])
    nova_planilha(
        "clientes_lista_oficial.xlsx",
        ["Cliente", "Qtd CPDs na planilha", "Origem", "Grupo sugerido", "Ativo"],
        linhas_clientes,
    )

    ct = resultado["conflitos_tecnicos"]
    nova_planilha(
        "conflitos_descricoes_classificados.xlsx",
        ["Código", "Categoria", "Descrição canônica proposta", "Descrições por fonte (JSON)", "Detalhes das comparações"],
        [
            [c["codigo"], c["categoria"], c["descricao_canonica_proposta"],
             json.dumps(c["descricoes_por_fonte"], ensure_ascii=False),
             " | ".join(p["detalhe"] for p in c["pares_comparados"])]
            for c in ct["classificados"]
        ],
    )

    nova_planilha(
        "duplicidades_para_decisao.xlsx",
        ["Código", "Linha Excel", "Descrição completa", "Cliente", "Peso", "Material", "Padrão embalagem", "Outras colunas (JSON)"],
        [
            [d["codigo"], ln, buscar_valor_coluna(linha, "descrição", "descricao"), buscar_valor_coluna(linha, "cliente"),
             buscar_valor_coluna(linha, "peso"), buscar_valor_coluna(linha, "material"),
             buscar_valor_coluna(linha, "padrao embalagem", "padrão embalagem"),
             json.dumps(linha, ensure_ascii=False, default=str)]
            for d in resultado["duplicidades"]
            for ln, linha in zip(d["linhas_de_origem"], d["linhas_completas"])
        ],
    )

    # Consolidado multi-abas com as 5 frentes de decisão.
    wb_final = Workbook()
    ws_resumo = wb_final.active
    ws_resumo.title = "Resumo"
    ws_resumo.append([f"Gerado em: {metadados['gerado_em']}"])
    ws_resumo.append([f"Planilha: {metadados['planilha_origem']}"])
    ws_resumo.append([])
    for chave, valor in resultado["resumo"].items():
        ws_resumo.append([chave, valor])

    ws_zeros = wb_final.create_sheet("Zeros a esquerda")
    ws_zeros.append(["CPD lista oficial", "CPD existente", "Descrição oficial", "Descrição existente", "Confiança", "Colisão?"])
    for c in resultado["zeros_esquerda"]["candidatos"]:
        ws_zeros.append([c["cpd_lista_oficial"], c["cpd_existente"], c["descricao_oficial"], c["descricao_existente"], c["nivel_confianca"], c["colisao"]])

    ws_clientes = wb_final.create_sheet("Clientes novos")
    ws_clientes.append(["Cliente", "Qtd CPDs", "Grupo sugerido"])
    for cliente in clientes["clientes_somente_na_lista_oficial"]:
        sugestao = clientes["sugestao_grupos"].get(cliente, {})
        ws_clientes.append([cliente, clientes["clientes_planilha_com_contagem_cpds"].get(cliente, 0), sugestao.get("grupo_sugerido")])

    ws_conf = wb_final.create_sheet("Conflitos tecnicos")
    ws_conf.append(["Código", "Categoria", "Descrição canônica proposta"])
    for c in ct["classificados"]:
        if c["categoria"] in ("C", "D"):
            ws_conf.append([c["codigo"], c["categoria"], c["descricao_canonica_proposta"]])

    ws_dup = wb_final.create_sheet("Duplicidades")
    ws_dup.append(["Código", "Linhas de origem", "Diferenças"])
    for d in resultado["duplicidades"]:
        ws_dup.append([d["codigo"], ", ".join(map(str, d["linhas_de_origem"])), " | ".join(d["diferencas"])])

    caminho_final = pasta_saida / "decisoes_pendentes_antes_migracao.xlsx"
    wb_final.save(caminho_final)
    arquivos.append(caminho_final)

    return arquivos


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args(argv=None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Análise somente-leitura da lista oficial de CPDs.")
    parser.add_argument("--planilha", type=Path, default=DEFAULT_PLANILHA)
    parser.add_argument("--banco", type=Path, default=DEFAULT_BANCO)
    parser.add_argument("--saida", type=Path, default=DEFAULT_SAIDA)
    parser.add_argument("--usar-copia", action="store_true", help="Trata --banco como já sendo uma cópia segura.")
    parser.add_argument(
        "--config-clientes", type=Path, required=True,
        help="Caminho do JSON local com as regras de clientes (grupo_honda/aliases/placeholders_invalidos). "
             "Ver config_local/regras_clientes_migracao.example.json.",
    )
    return parser.parse_args(argv)


def main(argv=None) -> int:
    args = parse_args(argv)
    if not args.planilha.exists():
        print(f"ERRO: planilha não encontrada: {args.planilha}")
        return 1
    if not args.banco.exists():
        print(f"ERRO: banco não encontrado: {args.banco}")
        return 1

    try:
        config = config_clientes.carregar_config_clientes(args.config_clientes)
    except (FileNotFoundError, config_clientes.ConfigClientesInvalida) as exc:
        print(f"ERRO: {exc}")
        return 1

    args.saida.mkdir(parents=True, exist_ok=True)
    if args.usar_copia:
        copia = args.banco
    else:
        copia = base_analise.criar_copia_segura(args.banco, args.saida / "_copias_temporarias")

    conn = base_analise.abrir_somente_leitura(copia)
    try:
        resultado = executar_analise_lista_oficial(conn, args.planilha, config.grupo_honda)
    finally:
        conn.close()

    metadados = {
        "planilha_origem": str(args.planilha),
        "banco_origem": str(args.banco),
        "copia_analisada": str(copia),
        "gerado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "config_clientes_caminho": config.caminho,
        "config_clientes_hash_sha256": config.hash_sha256,
    }

    arquivos = gravar_relatorios(resultado, metadados, args.saida)
    resumo = resultado["resumo"]

    print(f"Aba principal detectada: {resultado['leitura']['deteccao']['escolhida']['aba']}")
    print()
    print("--- Resumo objetivo ---")
    print(f"Candidatos a zero à esquerda: {resumo['total_candidatos_zero_esquerda']} "
          f"(alta={resumo['total_zero_esquerda_alta_confianca']}, "
          f"média={resumo['total_zero_esquerda_media_confianca']}, "
          f"baixa={resumo['total_zero_esquerda_baixa_confianca']}, "
          f"colisões={resumo['total_colisoes_zero_esquerda']})")
    print(f"Clientes somente na lista oficial: {resumo['total_clientes_somente_lista_oficial']}")
    print(f"Conflitos categoria A (formatação): {resumo['conflitos_categoria_A']}")
    print(f"Conflitos categoria B (prefixo PARAF): {resumo['conflitos_categoria_B']}")
    print(f"Conflitos categoria C (diferença técnica provável): {resumo['conflitos_categoria_C']}")
    print(f"Conflitos categoria D (descrições claramente diferentes): {resumo['conflitos_categoria_D']}")
    print()
    print("--- Duplicidades (detalhe completo) ---")
    for d in resultado["duplicidades"]:
        print(f"CPD {d['codigo']} (linhas {d['linhas_de_origem']}):")
        for ln, linha in zip(d["linhas_de_origem"], d["linhas_completas"]):
            print(f"  Linha {ln}: {json.dumps(linha, ensure_ascii=False, default=str)}")
        print(f"  Diferenças: {d['diferencas']}")
        print(f"  Descrição escolhida automaticamente: {d['descricao_que_seria_escolhida']!r} ({d['justificativa']})")
    print()
    print(f"Total de decisões humanas realmente necessárias: {resumo['total_decisoes_humanas_necessarias']}")
    print()
    print("Arquivos gerados:")
    for arq in arquivos:
        print(f"  {arq}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
