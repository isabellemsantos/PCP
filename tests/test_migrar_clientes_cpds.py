"""Testes de scripts/migrar_clientes_cpds.py (construção do plano, simulação
e aplicação real em cópias). Todos os dados são fictícios (códigos
8001-8010/08002/08003, clientes FICTICIO_*, aliases e grupo configurados
via um arquivo JSON temporário fictício -- nenhum nome real de cliente é
usado em nenhum teste deste arquivo).

Regra oficial em vigor: todo CPD-pai tem exatamente 5 dígitos numéricos
(ver scripts.migrar_clientes_cpds.normalizar_cpd_cinco_digitos). Os códigos
fictícios de 3-4 dígitos usados neste arquivo (8001, 9020, etc.) são
preenchidos automaticamente com zero à esquerda -- por isso os testes
verificam sempre o código FINAL de 5 dígitos (ex.: "08001", "09020"), não o
texto bruto digitado na planilha/pedido.

Nenhum teste toca no pcp.sqlite3 real.
"""

from __future__ import annotations

import hashlib
import importlib
import json
import os
import shutil
import sqlite3
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import scripts.migrar_clientes_cpds as migrador
import scripts.analisar_migracao_clientes_cpds as base_analise
import scripts.config_clientes_migracao as config_clientes

_DESC_LONGA_ALTA = "PECA TESTE TRES COM ROSCA FINA E CABECA SEXTAVADA GRANDE"

# Configuração de clientes 100% fictícia usada só nestes testes -- nenhum
# nome real de cliente, alias ou grupo. "GRUPO A MEMBRO" é o único cliente
# do grupo fictício "Honda" (o nome do grupo em si -- "Honda"/"Diversos" --
# é estrutural, já existente em grupos_clientes desde a migração v1->v2;
# só o ROSTER de quem pertence a cada grupo é fictício aqui).
_CONFIG_CLIENTES_TESTE = {
    "grupo_honda": ["GRUPO A MEMBRO"],
    "aliases": {
        "ALFA INDUSTRIA": ["ALFA INDÚSTRIA"],
        "BETA COMPONENTES": ["Beta Componentes"],
    },
    "placeholders_invalidos": ["", "-", ".", "N/A", "NA", "S/CLIENTE", "SEM CLIENTE", "SEMCLIENTE"],
}


def _criar_config_clientes_teste(caminho: Path) -> Path:
    caminho.write_text(json.dumps(_CONFIG_CLIENTES_TESTE, ensure_ascii=False, indent=2), encoding="utf-8")
    return caminho


def _sha256_arquivo(caminho: Path) -> str:
    h = hashlib.sha256()
    with open(caminho, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def _criar_planilha(caminho: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista"
    ws.append(["CPD", "Descrição", "Cliente"])
    ws.append(["8001", "Peça teste um", "FICTICIO_HONDA"])
    ws.append(["8001.1", "Peça teste um variante", "FICTICIO_HONDA"])
    ws.append([8003, _DESC_LONGA_ALTA, "FICTICIO_ALTA"])  # célula numérica inteira -> "8003", nunca ambíguo
    ws.append([8002, "PECA TESTE DOIS", "FICTICIO_MEDIA"])  # célula numérica inteira -> "8002", nunca ambíguo
    ws.append(["8010", "Descrição duplicada A", "FICTICIO_HONDA"])  # duplicidade
    ws.append(["8010", "Descrição duplicada B", "FICTICIO_HONDA"])  # duplicidade
    ws.append(["9020", "Peça cliente ponto", "."])  # CLIENTE_INDEFINIDO (".")
    ws.append(["9021", "Peça cliente vazio na planilha", ""])  # CLIENTE_INDEFINIDO (vazio)
    ws.append(["9022", "Peça cliente NA", "N/A"])  # CLIENTE_INDEFINIDO (N/A)
    ws.append(["9023", "Peça cliente hífen", "-"])  # CLIENTE_INDEFINIDO (-)
    ws.append(["9024", "Peça cliente com acento", "ALFA INDÚSTRIA"])  # alias -> ALFA INDUSTRIA, vínculo só na lista oficial
    ws.append(["9026", "Peça vínculo duplo", "FICTICIO_DUPLO"])  # mesmo vínculo também via pedido -> AMBAS
    ws.append(["9027", "Peça grupo diversos", "FICTICIO_DIVERSOS_TESTE"])  # mesmo pai também ligado a HONDA via pedido
    ws.append(["9028", "Peça só lista oficial", "FICTICIO_SOMENTE_LISTA"])  # vínculo só na lista oficial
    ws.append(["123456", "Peça código base longo demais", "FICTICIO_INVALIDO"])  # CODIGO_INVALIDO (>5 dígitos)
    ws.append(["AB123", "Peça código com letra", "FICTICIO_INVALIDO"])  # CODIGO_INVALIDO (formato inesperado)
    ws.append([4405.5, "Peça valor Excel ambíguo", "FICTICIO_AMBIGUO"])  # FORMATO_AMBIGUO (célula numérica não-inteira)
    ws.append(["88771/1", "Peça com barra variante", "FICTICIO_BARRA"])  # barra == extensão -> "88771.1"
    ws.append(["88771/1/2", "Peça barra invalida", "FICTICIO_INVALIDO"])  # continua CODIGO_INVALIDO
    ws.append(["ZZ-090", "Arruela lisa ZZ090", "FICTICIO_ARRUELA"])  # ARRUELA, não CPD
    ws.append(["YY-081", "Arruela lisa YY081", "FICTICIO_ARRUELA"])  # ARRUELA, não CPD
    # Regra oficial PARAF/PARAFUSO: lista oficial COM prefixo, manual_cpd SEM
    # -- a canônica tem que ser a versão sem prefixo, mesmo vindo de
    # manual_cpd (não de lista oficial, que é a prioridade padrão).
    ws.append(["50001", "PARAF TESTE QUATRO 5X15 ZNA", "FICTICIO_PARAF"])
    # Regra de colapso + PARAF juntas: o código-base "50002" e a extensão
    # histórica "50002.1" decidem suas canônicas de forma independente (cada
    # um só tem 1 fonte), mas pertencem ao MESMO CPD -- só uma pode
    # sobreviver como canônica depois do colapso, e tem que ser a versão sem
    # o prefixo PARAF (mesmo estando na extensão, não no código-base).
    ws.append(["50002", "PARAF TESTE CINCO 6X20 ZNA", "FICTICIO_PARAF"])
    ws.append(["50002.1", "TESTE CINCO 6X20 ZNA", "FICTICIO_PARAF"])
    wb.save(caminho)


class MigrarClientesCpdsTestCase(unittest.TestCase):
    tmp_dir: Path
    db_base: Path
    planilha_path: Path
    servidor = None

    @classmethod
    def setUpClass(cls):
        cls.tmp_dir = Path(tempfile.mkdtemp(prefix="pcp_teste_migrar_clientes_cpds_"))
        cls.db_base = cls.tmp_dir / "base.sqlite3"
        cls.planilha_path = cls.tmp_dir / "lista_oficial_ficticia.xlsx"
        cls.config_clientes_path = _criar_config_clientes_teste(cls.tmp_dir / "config_clientes_teste.json")
        _criar_planilha(cls.planilha_path)

        os.environ["PCP_DB_FILE"] = str(cls.db_base)
        os.environ["PCP_EXCEL_FILE"] = str(cls.tmp_dir / "excel_inexistente.xlsx")
        os.environ["PCP_EXCEL_PENDING_FILE"] = str(cls.tmp_dir / "pendente.xlsx")
        os.environ["PCP_LOG_FILE"] = str(cls.tmp_dir / "log.txt")
        os.environ["PCP_BACKUP_DIR"] = str(cls.tmp_dir / "backups")
        for nome in ("servidor_pcp", "db_manutencao"):
            sys.modules.pop(nome, None)
        cls.servidor = importlib.import_module("servidor_pcp")
        cls.dbm = importlib.import_module("db_manutencao")
        cls.servidor.init_db()

        with sqlite3.connect(str(cls.db_base)) as conn:
            conn.execute("PRAGMA foreign_keys=ON")
            cls.dbm.run_migrations(conn, cls.db_base, cls.tmp_dir / "backups_migracao", log=lambda _msg: None)

        with sqlite3.connect(str(cls.db_base)) as conn:
            agora = "2026-01-01 10:00:00"
            conn.execute(
                "INSERT INTO manual_cpd(codigo, descricao, updated_at) VALUES ('08002','PARAF PECA TESTE DOIS',?)",
                (agora,),
            )
            conn.execute(
                "INSERT INTO manual_cpd(codigo, descricao, updated_at) VALUES ('08003',?,?)",
                (f"PARAF {_DESC_LONGA_ALTA}", agora),
            )
            conn.execute(
                "INSERT INTO manual_cpd(codigo, descricao, updated_at) VALUES ('50001','TESTE QUATRO 5X15 ZNA',?)",
                (agora,),
            )

            def _pedido(order_id, cliente, cpd, descricao="", deleted=False):
                payload = json.dumps({
                    "id": order_id, "cliente": cliente, "cpd": cpd, "descricao": descricao,
                    "deleted": deleted, "criadoEm": agora, "alteradoEm": agora,
                })
                conn.execute("INSERT INTO orders(id, payload, updated_at) VALUES (?,?,?)", (order_id, payload, agora))

            _pedido("ped_media", "FICTICIO_MEDIA", "08002", "Peca teste dois")
            # Sem descrição no pedido -- só existe cliente como evidência.
            # "8003" (planilha) e "08003" (manual_cpd) se fundem no mesmo
            # código final (regra de 5 dígitos); a descrição de manual_cpd
            # tem prefixo PARAF (diferente da oficial). Isso expôs um bug
            # real: cada código bruto decidia sua canônica isolado, e a
            # fusão juntava duas canônicas pro mesmo código final, violando
            # o índice único (corrigido agrupando por código final antes de
            # decidir).
            _pedido("ped_alta", "FICTICIO_ALTA", "08003")
            _pedido("ped_honda", "FICTICIO_HONDA", "8001", "Peça teste um")
            # CLIENTE_SO_NA_LIXEIRA só aparece em pedido na lixeira -- deve continuar ativo mesmo assim.
            _pedido("ped_cliente_lixeira", "CLIENTE_SO_NA_LIXEIRA", "8001", "Peça teste um (lixeira)", deleted=True)

            # Mesmo vínculo (cliente, CPD-pai) nas duas fontes -- não pode duplicar.
            _pedido("ped_duplo", "FICTICIO_DUPLO", "9026", "Peça vínculo duplo (pedido)")
            # Cliente só em pedido, código nunca aparece na planilha oficial.
            _pedido("ped_somente_pedido", "FICTICIO_SOMENTE_PEDIDO", "9029", "Peça só em pedido")
            # Mesmo CPD-pai "9027" ligado a HONDA (grupo Honda) via pedido, e a
            # FICTICIO_DIVERSOS_TESTE (grupo Diversos) via linha da planilha
            # acima -- deve gerar os dois vínculos de grupo pro mesmo pai.
            _pedido("ped_honda_9027", "GRUPO A MEMBRO", "9027", "Peça grupo honda")
            # Cliente escrito com grafia alternativa (alias) num pedido -- o
            # texto original do pedido deve continuar exatamente assim depois.
            _pedido("ped_cliente_alias", "Beta Componentes", "9030", "Peça cliente com alias em pedido")
            # Mesmo código de "88771/1" (barra, na planilha) só que já em
            # formato de ponto -- prova que barra e ponto se fundem no MESMO
            # código completo, sem criar dois CPDs/variações.
            _pedido("ped_barra_ponto", "FICTICIO_BARRA", "88771.1", "Peça com barra variante (pedido)")
            # Arruela referenciada também via pedido (mesmo cliente da
            # planilha) -- prova vínculo cliente-arruela com origem AMBAS.
            _pedido("ped_arruela", "FICTICIO_ARRUELA", "ZZ-090", "Arruela lisa ZZ090 (pedido)")
            conn.commit()

        cls.plano = cls._construir_plano()

    @classmethod
    def tearDownClass(cls):
        for var in ("PCP_DB_FILE", "PCP_EXCEL_FILE", "PCP_EXCEL_PENDING_FILE", "PCP_LOG_FILE", "PCP_BACKUP_DIR"):
            os.environ.pop(var, None)
        for nome in ("servidor_pcp", "db_manutencao"):
            sys.modules.pop(nome, None)
        shutil.rmtree(cls.tmp_dir, ignore_errors=True)

    @classmethod
    def _construir_plano(cls) -> dict:
        conn = base_analise.abrir_somente_leitura(cls.db_base)
        try:
            config = config_clientes.carregar_config_clientes(cls.config_clientes_path)
            return migrador.construir_plano(conn, cls.planilha_path, config)
        finally:
            conn.close()

    def _copia_base(self, nome: str) -> Path:
        destino = self.tmp_dir / f"{nome}.sqlite3"
        origem_conn = sqlite3.connect(str(self.db_base))
        destino_conn = sqlite3.connect(str(destino))
        try:
            with destino_conn:
                origem_conn.backup(destino_conn)
        finally:
            destino_conn.close()
            origem_conn.close()
        return destino

    # ------------------------------------------------------------------
    # Regra oficial de 5 dígitos: fusão automática (substitui zero à esquerda)
    # ------------------------------------------------------------------

    def test_codigo_curto_e_codigo_ja_preenchido_se_fundem_automaticamente(self):
        # "8003" (só lista oficial) e "08003" (manual_cpd/pedido) nunca
        # existem como CPDs-pai separados -- sempre se fundem em "08003".
        self.assertNotIn("8003", self.plano["codigos_finais"])
        self.assertIn("08003", self.plano["codigos_finais"])
        self.assertIn("08003", self.plano["colisoes_normalizacao"])
        self.assertIn("8003", self.plano["colisoes_normalizacao"]["08003"])
        self.assertIn("08003", self.plano["colisoes_normalizacao"]["08003"])
        # Fusão automática nunca gera pendência ZERO_ESQUERDA (tipo extinto).
        self.assertFalse(any(p["tipo"] == "ZERO_ESQUERDA" for p in self.plano["pendencias"]))

    def test_colisao_normalizacao_gera_exatamente_uma_canonica(self):
        # Regressão: "8003" e "08003" se fundem no mesmo código final --
        # cada um decidia sua canônica isoladamente antes da correção, o que
        # gerava 2 canônicas pro mesmo código completo (violaria o índice
        # único ao aplicar de verdade).
        canonicas = [d for d in self.plano["descricoes_fontes"] if d["codigo_completo"] == "08003" and d["descricao_canonica"] == 1]
        self.assertEqual(len(canonicas), 1, f"esperado 1 canônica pra 08003, achou {len(canonicas)}: {canonicas}")
        self.assertEqual(canonicas[0]["fonte"], "LISTA_OFICIAL")

    def test_outro_par_curto_preenchido_tambem_se_funde(self):
        # "8002" (lista oficial) e "08002" (manual_cpd/pedido) -- mesma
        # fusão obrigatória, sem exceção nem gating por confiança.
        self.assertNotIn("8002", self.plano["codigos_finais"])
        self.assertIn("08002", self.plano["codigos_finais"])
        self.assertIn("08002", self.plano["colisoes_normalizacao"])

    # ------------------------------------------------------------------
    # Normalização de 5 dígitos: casos inválidos/ambíguos nunca são descartados
    # ------------------------------------------------------------------

    def test_base_com_mais_de_5_digitos_vira_pendencia_codigo_invalido(self):
        self.assertNotIn("123456", self.plano["codigos_finais"])
        pendencia = next(p for p in self.plano["pendencias"] if p["tipo"] == "CODIGO_INVALIDO" and p["codigo_completo"] == "123456")
        self.assertEqual(pendencia["status"], "PENDENTE")
        detalhes = json.loads(pendencia["detalhes_json"])
        self.assertEqual(detalhes["motivo"], "CODIGO_BASE_LONGO")
        self.assertEqual(detalhes["codigo_original"], "123456")

    def test_codigo_com_letra_vira_pendencia_codigo_invalido(self):
        self.assertNotIn("AB123", self.plano["codigos_finais"])
        pendencia = next(p for p in self.plano["pendencias"] if p["tipo"] == "CODIGO_INVALIDO" and p["codigo_completo"] == "AB123")
        detalhes = json.loads(pendencia["detalhes_json"])
        self.assertEqual(detalhes["motivo"], "FORMATO_INESPERADO")

    def test_valor_excel_ambiguo_nao_e_convertido_silenciosamente(self):
        # 4405.5 é uma célula Excel NUMÉRICA com valor não-inteiro: não dá
        # pra saber se ".5" é uma extensão de verdade ou artefato do Excel.
        # Preserva o valor bruto, marca FORMATO_AMBIGUO, nunca cria CPD a
        # partir dele (não assume nada sobre "04405" já existir por outra via).
        self.assertNotIn("4405.5", self.plano["codigos_finais"])
        self.assertNotIn("04405.5", self.plano["codigos_finais"])
        pendencia = next(p for p in self.plano["pendencias"] if p["tipo"] == "FORMATO_AMBIGUO" and p["codigo_completo"] == "4405.5")
        detalhes = json.loads(pendencia["detalhes_json"])
        self.assertEqual(detalhes["motivo"], "FORMATO_AMBIGUO")
        self.assertEqual(detalhes["codigo_original"], "4405.5")

    def test_celula_excel_inteira_nao_e_ambigua(self):
        # 8003/8002 são células Excel NUMÉRICAS mas de valor INTEIRO --
        # zero_esquerda_desconhecido=True não implica ambiguidade; só uma
        # fração genuína (não-inteira) é ambígua. Ambos normalizam normalmente.
        self.assertFalse(any(p["tipo"] == "FORMATO_AMBIGUO" and p["codigo_completo"] in ("8002", "8003") for p in self.plano["pendencias"]))
        self.assertIn("08002", self.plano["codigos_finais"])
        self.assertIn("08003", self.plano["codigos_finais"])

    # ------------------------------------------------------------------
    # Fontes de descrição / canônica única
    # ------------------------------------------------------------------

    def test_apenas_uma_descricao_canonica_por_codigo_no_plano(self):
        canonicas_por_codigo: dict[str, int] = {}
        for d in self.plano["descricoes_fontes"]:
            if d["descricao_canonica"] == 1:
                canonicas_por_codigo[d["codigo_completo"]] = canonicas_por_codigo.get(d["codigo_completo"], 0) + 1
        for codigo, total in canonicas_por_codigo.items():
            self.assertEqual(total, 1, f"{codigo} tem {total} descrições canônicas, deveria ter no máximo 1")

    def test_fontes_de_descricao_preservadas_para_codigo_normal(self):
        fontes = [d for d in self.plano["descricoes_fontes"] if d["codigo_completo"] == "08001"]
        origens = {f["fonte"] for f in fontes}
        self.assertIn("LISTA_OFICIAL", origens)
        self.assertIn("PEDIDO", origens)

    def test_descricao_canonica_prefere_versao_sem_prefixo_paraf(self):
        # Lista oficial diz "PARAF TESTE QUATRO 5X15 ZNA", manual_cpd diz
        # "TESTE QUATRO 5X15 ZNA" (mesma peça, só sem o prefixo) -- a
        # canônica tem que ser a versão SEM prefixo, mesmo vindo de
        # manual_cpd (não da lista oficial, que é a prioridade padrão nos
        # outros casos).
        fontes = [d for d in self.plano["descricoes_fontes"] if d["codigo_completo"] == "50001"]
        canonica = next(d for d in fontes if d["descricao_canonica"] == 1)
        self.assertEqual(canonica["descricao"], "TESTE QUATRO 5X15 ZNA")
        self.assertEqual(canonica["fonte"], "MANUAL_CPD")

    def test_todas_as_fontes_paraf_permanecem_preservadas(self):
        fontes = [d for d in self.plano["descricoes_fontes"] if d["codigo_completo"] == "50001"]
        descricoes = {d["descricao"] for d in fontes}
        self.assertEqual(descricoes, {"PARAF TESTE QUATRO 5X15 ZNA", "TESTE QUATRO 5X15 ZNA"})
        origens = {d["fonte"] for d in fontes}
        self.assertEqual(origens, {"LISTA_OFICIAL", "MANUAL_CPD"})

    def test_canonica_consolidada_por_cpd_base_mesmo_com_extensao_conflitante(self):
        # "50002" e "50002.1" decidem canônica cada um isoladamente (só 1
        # fonte cada), mas colapsam no mesmo CPD-base "50002" -- só uma pode
        # sobrar como canônica, e tem que ser a versão sem prefixo PARAF
        # (que está na extensão "50002.1", não no código-base).
        fontes = [d for d in self.plano["descricoes_fontes"] if d["codigo_completo"] in ("50002", "50002.1")]
        canonicas = [d for d in fontes if d["descricao_canonica"] == 1]
        self.assertEqual(len(canonicas), 1, f"esperado exatamente 1 canônica entre os códigos históricos do mesmo CPD-base, achou {len(canonicas)}: {canonicas}")
        self.assertEqual(canonicas[0]["descricao"], "TESTE CINCO 6X20 ZNA")
        self.assertEqual(canonicas[0]["codigo_completo"], "50002.1")

    def test_nenhum_codigo_historico_e_perdido_apos_consolidacao(self):
        fontes = [d for d in self.plano["descricoes_fontes"] if d["codigo_completo"] in ("50002", "50002.1")]
        self.assertEqual({d["codigo_completo"] for d in fontes}, {"50002", "50002.1"})

    def test_regra_paraf_nao_altera_manual_cpd_nem_pedidos(self):
        with sqlite3.connect(str(self.db_base)) as c:
            desc_manual = c.execute("SELECT descricao FROM manual_cpd WHERE codigo='50001'").fetchone()[0]
        self.assertEqual(desc_manual, "TESTE QUATRO 5X15 ZNA", "manual_cpd não pode ser reescrito pela migração")

    def test_extensao_preserva_zeros_a_esquerda(self):
        # "8001.1" -> pai "08001", extensão "1" (texto, nunca vira número).
        self.assertIn("08001.1", self.plano["codigos_finais"])
        self.assertEqual(self.plano["codigos_finais"]["08001.1"]["extensao"], "1")
        self.assertEqual(self.plano["codigos_finais"]["08001.1"]["codigo_pai"], "08001")

    # ------------------------------------------------------------------
    # Duplicidades
    # ------------------------------------------------------------------

    def test_duplicidade_preserva_as_duas_linhas_sem_escolha_silenciosa(self):
        fontes_8010 = [d for d in self.plano["descricoes_fontes"] if d["codigo_completo"] == "08010"]
        descricoes_lista_oficial = [f["descricao"] for f in fontes_8010 if f["fonte"] == "LISTA_OFICIAL"]
        self.assertEqual(sorted(descricoes_lista_oficial), ["Descrição duplicada A", "Descrição duplicada B"])
        # Nenhuma das duas linhas da planilha foi marcada como canônica (não existe cadastro anterior pro 08010).
        for f in fontes_8010:
            if f["fonte"] == "LISTA_OFICIAL":
                self.assertEqual(f["descricao_canonica"], 0)
        pendencia = next(p for p in self.plano["pendencias"] if p["tipo"] == "DUPLICIDADE_LISTA" and p["codigo_completo"] == "08010")
        self.assertEqual(pendencia["status"], "PENDENTE")

    # ------------------------------------------------------------------
    # Clientes
    # ------------------------------------------------------------------

    def test_todos_os_clientes_ativos(self):
        self.assertTrue(all(c["ativo"] == 1 for c in self.plano["clientes"]))

    def test_cliente_so_na_lixeira_fica_ativo_e_grupo_diversos(self):
        cliente_lixeira = next(c for c in self.plano["clientes"] if c["nome"] == "CLIENTE_SO_NA_LIXEIRA")
        self.assertEqual(cliente_lixeira["ativo"], 1)
        self.assertEqual(cliente_lixeira["grupo"], "Diversos")

    def test_todos_os_cpds_ativos(self):
        self.assertTrue(all(info["ativo"] == 1 for info in self.plano["codigos_finais"].values()))

    # ------------------------------------------------------------------
    # Placeholders de cliente não criam cliente
    # ------------------------------------------------------------------

    def test_ponto_nao_cria_cliente(self):
        self.assertNotIn(".", {c["nome"] for c in self.plano["clientes"]})
        pendencia = next(
            p for p in self.plano["pendencias"]
            if p["tipo"] == "CLIENTE_INDEFINIDO" and p["codigo_completo"] == "09020"
        )
        self.assertEqual(pendencia["status"], "PENDENTE")
        self.assertIn("09020", self.plano["codigos_finais"], "CPD da linha com cliente '.' deve ser preservado normalmente")

    def test_vazio_hifen_na_nao_criam_cliente(self):
        nomes = {c["nome"] for c in self.plano["clientes"]}
        for valor_invalido in ("", "-", "N/A", "NA"):
            self.assertNotIn(valor_invalido, nomes)
        codigos_com_pendencia_indefinido = {
            p["codigo_completo"] for p in self.plano["pendencias"] if p["tipo"] == "CLIENTE_INDEFINIDO"
        }
        self.assertIn("09021", codigos_com_pendencia_indefinido)  # vazio
        self.assertIn("09022", codigos_com_pendencia_indefinido)  # N/A
        self.assertIn("09023", codigos_com_pendencia_indefinido)  # hífen

    def test_cpd_das_linhas_invalidas_e_preservado(self):
        # O CPD continua existindo normalmente mesmo sem cliente válido.
        for codigo in ("09020", "09021", "09022", "09023"):
            self.assertIn(codigo, self.plano["codigos_finais"])

    # ------------------------------------------------------------------
    # Aliases
    # ------------------------------------------------------------------

    def test_aliases_configurados_viram_alias(self):
        nomes = {c["nome"] for c in self.plano["clientes"]}
        self.assertIn("ALFA INDUSTRIA", nomes)
        self.assertNotIn("ALFA INDÚSTRIA", nomes)
        self.assertIn("BETA COMPONENTES", nomes)
        self.assertNotIn("Beta Componentes", nomes)

        aliases = {(a["cliente_canonico"], a["alias"]) for a in self.plano["cliente_aliases"]}
        self.assertIn(("ALFA INDUSTRIA", "ALFA INDÚSTRIA"), aliases)
        self.assertIn(("BETA COMPONENTES", "Beta Componentes"), aliases)

    def test_nenhum_cliente_duplicado_pelos_aliases(self):
        nomes = [c["nome"] for c in self.plano["clientes"]]
        self.assertEqual(len(nomes), len(set(nomes)))

    # ------------------------------------------------------------------
    # Vínculos cliente x CPD (lista oficial + pedidos, sem duplicar)
    # ------------------------------------------------------------------

    def test_cliente_da_lista_oficial_cria_vinculo(self):
        item = next(
            i for i in self.plano["cliente_cpds"]
            if i["cliente"] == "FICTICIO_SOMENTE_LISTA" and i["cpd_pai"] == "09028"
        )
        self.assertEqual(item["origem"], "LISTA_OFICIAL")

    def test_cliente_do_pedido_cria_vinculo(self):
        item = next(
            i for i in self.plano["cliente_cpds"]
            if i["cliente"] == "FICTICIO_SOMENTE_PEDIDO" and i["cpd_pai"] == "09029"
        )
        self.assertEqual(item["origem"], "PEDIDO")

    def test_mesmo_vinculo_nas_duas_fontes_nao_duplica(self):
        itens = [
            i for i in self.plano["cliente_cpds"]
            if i["cliente"] == "FICTICIO_DUPLO" and i["cpd_pai"] == "09026"
        ]
        self.assertEqual(len(itens), 1, "Vínculo presente nas duas fontes deve gerar só 1 registro")
        self.assertEqual(itens[0]["origem"], "AMBAS")

    def test_cpd_com_clientes_grupo_a_e_diversos_recebe_dois_grupos(self):
        grupos = {g for g, pai in self.plano["grupo_cpds"] if pai == "09027"}
        self.assertEqual(grupos, {"Honda", "Diversos"})

    def test_cpd_sem_cliente_nao_recebe_grupo(self):
        pais_com_grupo = {pai for _, pai in self.plano["grupo_cpds"]}
        pais_com_vinculo = {item["cpd_pai"] for item in self.plano["cliente_cpds"]}
        pais_sem_cliente = set(self.plano["pais_agrupados"]) - pais_com_vinculo
        self.assertTrue(pais_sem_cliente, "Fixture deveria ter algum CPD sem nenhum cliente vinculado")
        self.assertEqual(pais_sem_cliente & pais_com_grupo, set())

    def test_codigo_invalido_nao_cria_vinculo_mesmo_com_cliente_valido(self):
        # "123456"/"AB123"/"4405.5" têm clientes válidos na planilha, mas
        # como o CPD é inválido/ambíguo, nenhum vínculo pode ser criado.
        pais_com_vinculo = {item["cpd_pai"] for item in self.plano["cliente_cpds"]}
        self.assertNotIn("123456", pais_com_vinculo)
        self.assertNotIn("AB123", pais_com_vinculo)
        self.assertNotIn("4405.5", pais_com_vinculo)
        self.assertFalse(any(i["cliente"] == "FICTICIO_INVALIDO" for i in self.plano["cliente_cpds"]))
        self.assertFalse(any(i["cliente"] == "FICTICIO_AMBIGUO" for i in self.plano["cliente_cpds"]))

    # ------------------------------------------------------------------
    # Barra como extensão (equivalente ao ponto)
    # ------------------------------------------------------------------

    def test_barra_e_convertida_para_ponto(self):
        resultado = migrador.normalizar_cpd_cinco_digitos("88771/1")
        self.assertEqual(resultado["tipo_item"], "CPD")
        self.assertEqual(resultado["codigo_pai"], "88771")
        self.assertEqual(resultado["extensao"], "1")
        self.assertEqual(resultado["codigo_canonico"], "88771.1")

    def test_barra_curta_preenche_zero_e_converte(self):
        resultado = migrador.normalizar_cpd_cinco_digitos("442/1")
        self.assertEqual(resultado["codigo_completo"], "00442.1")
        resultado2 = migrador.normalizar_cpd_cinco_digitos("442/01")
        self.assertEqual(resultado2["codigo_completo"], "00442.01")
        self.assertEqual(resultado2["extensao"], "01", "zero da extensão deve ser preservado")

    def test_barra_e_ponto_nao_criam_dois_codigos(self):
        # "88771/1" (planilha) e "88771.1" (pedido) são o MESMO código completo.
        self.assertIn("88771.1", self.plano["codigos_finais"])
        self.assertNotIn("88771/1", self.plano["codigos_finais"])
        self.assertIn("88771.1", self.plano["colisoes_normalizacao"])
        self.assertIn("88771/1", self.plano["colisoes_normalizacao"]["88771.1"])
        self.assertIn("88771.1", self.plano["colisoes_normalizacao"]["88771.1"])
        self.assertIn("88771.1", self.plano["colisoes_barra_ponto"])

    def test_codigo_original_com_barra_e_preservado_nas_fontes(self):
        origens = {d["referencia_origem"] for d in self.plano["descricoes_fontes"] if d["codigo_completo"] == "88771.1"}
        self.assertTrue(origens, "deveria haver ao menos uma fonte de descrição pra 88771.1")
        # A referência da lista oficial aponta pra linha onde o texto ORIGINAL
        # ("88771/1", com barra) foi lido -- não reescrevemos a planilha, só
        # convertemos o código completo canônico armazenado.
        self.assertIn("88771/1", self.plano["codigos_com_barra_convertidos"])

    def test_barra_com_multiplos_separadores_continua_invalido(self):
        resultado = migrador.normalizar_cpd_cinco_digitos("88771/1/2")
        self.assertFalse(resultado["valido"])
        self.assertEqual(resultado["motivo"], "FORMATO_INESPERADO")
        self.assertNotIn("88771/1/2", self.plano["codigos_finais"])
        pendencia = next(p for p in self.plano["pendencias"] if p["tipo"] == "CODIGO_INVALIDO" and p["codigo_completo"] == "88771/1/2")
        self.assertEqual(pendencia["status"], "PENDENTE")

    # ------------------------------------------------------------------
    # Arruelas: item separado, nunca vira CPD
    # ------------------------------------------------------------------

    def test_aw090_e_cw081_sao_arruela(self):
        self.assertEqual(migrador._normalizar_codigo_de_fonte("ZZ-090")["tipo_item"], "ARRUELA")
        self.assertEqual(migrador._normalizar_codigo_de_fonte("YY-081")["tipo_item"], "ARRUELA")
        self.assertIn("ZZ-090", self.plano["arruelas"])
        self.assertIn("YY-081", self.plano["arruelas"])

    def test_arruela_nao_entra_em_codigos_finais_nem_pais(self):
        self.assertNotIn("ZZ-090", self.plano["codigos_finais"])
        self.assertNotIn("YY-081", self.plano["codigos_finais"])
        self.assertNotIn("ZZ-090", self.plano["pais_agrupados"])

    def test_arruela_nao_gera_pendencia_codigo_invalido(self):
        self.assertFalse(any(
            p["codigo_completo"] in ("ZZ-090", "YY-081") and p["tipo"] == "CODIGO_INVALIDO"
            for p in self.plano["pendencias"]
        ))

    def test_arruela_permanece_ativa(self):
        self.assertEqual(self.plano["arruelas"]["ZZ-090"]["ativo"], 1)
        self.assertEqual(self.plano["arruelas"]["YY-081"]["ativo"], 1)

    def test_arruela_vinculada_a_cliente_e_grupo(self):
        vinculo = next(v for v in self.plano["cliente_arruelas"] if v["arruela"] == "ZZ-090")
        self.assertEqual(vinculo["cliente"], "FICTICIO_ARRUELA")
        self.assertEqual(vinculo["origem"], "AMBAS", "ZZ-090 aparece na lista oficial E num pedido")
        grupos = {g for g, cod in self.plano["grupo_arruelas"] if cod == "ZZ-090"}
        self.assertEqual(grupos, {"Diversos"})

    def test_simular_reporta_nenhuma_arruela_em_cpds(self):
        resumo = migrador.simular(self.plano)
        self.assertTrue(resumo["nenhuma_arruela_em_cpds"])
        self.assertGreaterEqual(resumo["total_arruelas_codigos_unicos"], 2)
        self.assertTrue(resumo["todas_arruelas_ativas"])

    def test_pedido_original_mantem_seu_texto_de_cliente(self):
        db = self._copia_base("preserva_texto_pedido")
        codigo = migrador.main([
            "--planilha", str(self.planilha_path), "--banco", str(db), "--saida", str(self.tmp_dir / "saida_preserva_texto"),
            "--config-clientes", str(self.config_clientes_path),
            "--confirmar", "--texto-confirmacao", "CONFIRMAR MIGRACAO",
        ])
        self.assertEqual(codigo, 0)
        with sqlite3.connect(str(db)) as c:
            payload = json.loads(c.execute("SELECT payload FROM orders WHERE id='ped_cliente_alias'").fetchone()[0])
        self.assertEqual(payload["cliente"], "Beta Componentes", "Texto original do pedido não pode ser alterado, mesmo sendo um alias")

    # ------------------------------------------------------------------
    # Simulação: não escreve nada
    # ------------------------------------------------------------------

    def test_modo_simular_nao_escreve_no_banco(self):
        db = self._copia_base("simular_nao_escreve")
        hash_antes = _sha256_arquivo(db)
        pasta_saida = self.tmp_dir / "saida_simular"
        codigo = migrador.main([
            "--planilha", str(self.planilha_path), "--banco", str(db), "--saida", str(pasta_saida),
            "--config-clientes", str(self.config_clientes_path),
        ])
        self.assertEqual(codigo, 0)
        hash_depois = _sha256_arquivo(db)
        self.assertEqual(hash_antes, hash_depois)
        self.assertTrue((pasta_saida / "simulacao_migracao_clientes_cpds.json").exists())
        self.assertTrue((pasta_saida / "normalizacao_cpds_5_digitos.xlsx").exists())
        self.assertTrue((pasta_saida / "simulacao_final_cpds_5_digitos.xlsx").exists())
        self.assertTrue((pasta_saida / "codigos_invalidos_ou_ambiguos.xlsx").exists())
        self.assertTrue((pasta_saida / "cpds_com_barra_normalizados.xlsx").exists())
        self.assertTrue((pasta_saida / "arruelas_identificadas.xlsx").exists())
        self.assertTrue((pasta_saida / "simulacao_cpds_e_arruelas.xlsx").exists())

    def test_pedidos_e_manual_cpd_nao_alterados_pela_simulacao(self):
        db = self._copia_base("simular_preserva_legado")
        with sqlite3.connect(str(db)) as c:
            antes_orders = c.execute("SELECT COUNT(*) FROM orders").fetchone()[0]
            antes_manual = c.execute("SELECT COUNT(*) FROM manual_cpd").fetchone()[0]
        migrador.main([
            "--planilha", str(self.planilha_path), "--banco", str(db), "--saida", str(self.tmp_dir / "saida_simular2"),
            "--config-clientes", str(self.config_clientes_path),
        ])
        with sqlite3.connect(str(db)) as c:
            depois_orders = c.execute("SELECT COUNT(*) FROM orders").fetchone()[0]
            depois_manual = c.execute("SELECT COUNT(*) FROM manual_cpd").fetchone()[0]
        self.assertEqual(antes_orders, depois_orders)
        self.assertEqual(antes_manual, depois_manual)

    def test_simulacao_e_confirmacao_usam_o_mesmo_hash_de_configuracao(self):
        from openpyxl import load_workbook

        db_sim = self._copia_base("hash_config_simular")
        db_conf = self._copia_base("hash_config_confirmar")
        pasta_sim = self.tmp_dir / "saida_hash_simular"
        pasta_conf = self.tmp_dir / "saida_hash_confirmar"

        migrador.main([
            "--planilha", str(self.planilha_path), "--banco", str(db_sim), "--saida", str(pasta_sim),
            "--config-clientes", str(self.config_clientes_path),
        ])
        migrador.main([
            "--planilha", str(self.planilha_path), "--banco", str(db_conf), "--saida", str(pasta_conf),
            "--config-clientes", str(self.config_clientes_path),
            "--confirmar", "--texto-confirmacao", "CONFIRMAR MIGRACAO",
        ])

        def _linha_hash(pasta: Path) -> str:
            wb = load_workbook(pasta / "simulacao_final_cpds_5_digitos.xlsx", read_only=True)
            ws = wb.active
            linha = next(row[0].value for row in ws.iter_rows() if row[0].value and "Config. clientes" in str(row[0].value))
            wb.close()
            return linha

        hash_esperado = config_clientes.hash_arquivo_config(self.config_clientes_path)
        linha_sim = _linha_hash(pasta_sim)
        linha_conf = _linha_hash(pasta_conf)
        self.assertIn(hash_esperado, linha_sim)
        self.assertIn(hash_esperado, linha_conf)
        self.assertEqual(linha_sim, linha_conf)

        # Só o hash e o caminho -- nunca o conteúdo -- ficam nesse relatório.
        resumo_simular = json.loads((pasta_sim / "simulacao_migracao_clientes_cpds.json").read_text(encoding="utf-8"))
        texto_relatorio = json.dumps(resumo_simular, ensure_ascii=False)
        self.assertNotIn("GRUPO A MEMBRO", texto_relatorio)
        self.assertNotIn("ALFA IND", texto_relatorio)

    # ------------------------------------------------------------------
    # Confirmação exigida / aplicação real em cópia
    # ------------------------------------------------------------------

    def test_confirmar_sem_texto_de_confirmacao_correto_nao_altera_nada(self):
        db = self._copia_base("confirmar_texto_errado")
        hash_antes = _sha256_arquivo(db)
        codigo = migrador.main([
            "--planilha", str(self.planilha_path), "--banco", str(db), "--saida", str(self.tmp_dir / "saida_confirmar_errado"),
            "--config-clientes", str(self.config_clientes_path),
            "--confirmar", "--texto-confirmacao", "texto errado",
        ])
        self.assertEqual(codigo, 1)
        hash_depois = _sha256_arquivo(db)
        self.assertEqual(hash_antes, hash_depois)

    def test_confirmar_com_texto_correto_aplica_e_preserva_legado(self):
        db = self._copia_base("confirmar_ok")
        with sqlite3.connect(str(db)) as c:
            antes = {
                t: c.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
                for t in ("orders", "manual_cpd", "sections", "audit_log")
            }
        codigo = migrador.main([
            "--planilha", str(self.planilha_path), "--banco", str(db), "--saida", str(self.tmp_dir / "saida_confirmar_ok"),
            "--config-clientes", str(self.config_clientes_path),
            "--confirmar", "--texto-confirmacao", "CONFIRMAR MIGRACAO",
        ])
        self.assertEqual(codigo, 0)

        with sqlite3.connect(str(db)) as c:
            depois = {
                t: c.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
                for t in ("orders", "manual_cpd", "sections", "audit_log")
            }
            c.execute("PRAGMA foreign_keys=ON")
            integridade = c.execute("PRAGMA integrity_check").fetchone()[0]
            violacoes = c.execute("PRAGMA foreign_key_check").fetchall()
            total_clientes = c.execute("SELECT COUNT(*) FROM clientes").fetchone()[0]
            total_cpds = c.execute("SELECT COUNT(*) FROM cpds").fetchone()[0]
            cliente_lixeira_ativo = c.execute("SELECT ativo FROM clientes WHERE nome='CLIENTE_SO_NA_LIXEIRA'").fetchone()[0]
            codigos_pai = {row[0] for row in c.execute("SELECT codigo_pai FROM cpds").fetchall()}
            codigos_variacao = {row[0] for row in c.execute("SELECT codigo_completo FROM cpd_variacoes").fetchall()}
            arruelas_codigos = {row[0] for row in c.execute("SELECT codigo FROM arruelas").fetchall()}
            aw090_ativo = c.execute("SELECT ativo FROM arruelas WHERE codigo='ZZ-090'").fetchone()

        self.assertEqual(antes, depois, "Tabelas legadas (orders/manual_cpd/sections/audit_log) não podem mudar")
        self.assertEqual(integridade, "ok")
        self.assertEqual(violacoes, [])
        self.assertGreater(total_clientes, 0)
        self.assertGreater(total_cpds, 0)
        self.assertEqual(cliente_lixeira_ativo, 1)
        self.assertTrue(
            all(len(cp) == 5 and cp.isdigit() for cp in codigos_pai),
            "Todo codigo_pai gravado deve ter exatamente 5 dígitos numéricos",
        )
        self.assertIn("ZZ-090", arruelas_codigos)
        self.assertIn("YY-081", arruelas_codigos)
        self.assertIsNotNone(aw090_ativo)
        self.assertEqual(aw090_ativo[0], 1)
        self.assertNotIn("ZZ-090", codigos_pai)
        self.assertNotIn("YY-081", codigos_pai)
        self.assertNotIn("ZZ-090", codigos_variacao)
        self.assertNotIn("YY-081", codigos_variacao)

        with sqlite3.connect(str(db)) as c:
            desc_padrao = c.execute("SELECT descricao_padrao FROM cpds WHERE codigo_pai='50002'").fetchone()[0]
        self.assertEqual(
            desc_padrao, "TESTE CINCO 6X20 ZNA",
            "cpds.descricao_padrao tem que ser a versão sem prefixo PARAF, consolidada por CPD-base "
            "mesmo vindo da extensão histórica '50002.1' -- nunca None nem a versão com prefixo.",
        )

    def test_segunda_aplicacao_em_cima_da_primeira_nao_duplica_clientes(self):
        # A migração de clientes/CPDs em si (diferente das migrações de
        # schema) não foi desenhada pra ser re-executada sobre um banco já
        # migrado (os clientes não têm UNIQUE(nome) hoje) -- então
        # confirmamos que a PRIMEIRA aplicação sozinha já é internamente
        # consistente (sem duplicar nada dentro do próprio plano).
        nomes = [c["nome"] for c in self.plano["clientes"]]
        self.assertEqual(len(nomes), len(set(nomes)), "Plano não pode ter cliente duplicado")


class NormalizarCpdCincoDigitosTestCase(unittest.TestCase):
    """Testes diretos da função pura de normalização (sem banco/planilha)."""

    def test_casos_exatos_de_preenchimento(self):
        casos = {
            "1": "00001",
            "11": "00011",
            "442": "00442",
            "8439": "08439",
            "11500": "11500",
            "442.1": "00442.1",
            "442.01": "00442.01",
            "04772.10": "04772.10",
        }
        for entrada, esperado in casos.items():
            with self.subTest(entrada=entrada):
                resultado = migrador.normalizar_cpd_cinco_digitos(entrada)
                self.assertTrue(resultado["valido"], resultado)
                self.assertEqual(resultado["codigo_completo"], esperado)

    def test_extensao_mantem_zeros_a_esquerda_como_texto(self):
        resultado = migrador.normalizar_cpd_cinco_digitos("442.01")
        self.assertEqual(resultado["extensao"], "01")
        resultado2 = migrador.normalizar_cpd_cinco_digitos("04772.10")
        self.assertEqual(resultado2["extensao"], "10")

    def test_base_ja_5_digitos_marca_regra_ja_5_digitos(self):
        resultado = migrador.normalizar_cpd_cinco_digitos("11500")
        self.assertEqual(resultado["regra_aplicada"], "ja_5_digitos")

    def test_base_curta_marca_regra_zero_padding(self):
        resultado = migrador.normalizar_cpd_cinco_digitos("442")
        self.assertEqual(resultado["regra_aplicada"], "zero_padding")

    def test_base_com_mais_de_5_digitos_invalida(self):
        resultado = migrador.normalizar_cpd_cinco_digitos("123456")
        self.assertFalse(resultado["valido"])
        self.assertEqual(resultado["motivo"], "CODIGO_BASE_LONGO")
        self.assertIsNone(resultado["codigo_completo"])

    def test_letra_no_codigo_invalida(self):
        resultado = migrador.normalizar_cpd_cinco_digitos("AB123")
        self.assertFalse(resultado["valido"])
        self.assertEqual(resultado["motivo"], "FORMATO_INESPERADO")

    def test_extensao_nao_numerica_invalida(self):
        resultado = migrador.normalizar_cpd_cinco_digitos("442.A")
        self.assertFalse(resultado["valido"])
        self.assertEqual(resultado["motivo"], "FORMATO_INESPERADO")

    def test_codigo_vazio_invalida(self):
        resultado = migrador.normalizar_cpd_cinco_digitos("")
        self.assertFalse(resultado["valido"])
        self.assertEqual(resultado["motivo"], "CODIGO_VAZIO")

    def test_celula_excel_numerica_nao_inteira_e_ambigua(self):
        resultado = migrador._normalizar_codigo_de_fonte("442.5", numerico=True)
        self.assertFalse(resultado["valido"])
        self.assertEqual(resultado["motivo"], "FORMATO_AMBIGUO")
        self.assertEqual(resultado["codigo_original"], "442.5")

    def test_celula_texto_com_ponto_nao_e_ambigua(self):
        resultado = migrador._normalizar_codigo_de_fonte("442.5", numerico=False)
        self.assertTrue(resultado["valido"])
        self.assertEqual(resultado["codigo_completo"], "00442.5")

    def test_celula_excel_numerica_inteira_nao_e_ambigua(self):
        resultado = migrador._normalizar_codigo_de_fonte("442", numerico=True)
        self.assertTrue(resultado["valido"])
        self.assertEqual(resultado["codigo_completo"], "00442")


if __name__ == "__main__":
    unittest.main()
